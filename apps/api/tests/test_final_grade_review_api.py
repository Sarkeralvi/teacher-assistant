from collections.abc import Iterator
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from PIL import Image
from sqlalchemy import delete
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.db.session import SessionLocal
from app.main import app
from app.models import (
    AnswerRegion,
    Assessment,
    AuditLog,
    Course,
    FinalGrade,
    GradeSuggestion,
    GradingJob,
    Question,
    Rubric,
    Submission,
    SubmissionPage,
    User,
)

CLEANUP_MODELS = (
    AuditLog,
    FinalGrade,
    GradeSuggestion,
    GradingJob,
    AnswerRegion,
    SubmissionPage,
    Submission,
    Rubric,
    Question,
    Assessment,
    Course,
    User,
)


@pytest.fixture()
def db_session() -> Iterator[Session]:
    db = SessionLocal()
    try:
        for model in CLEANUP_MODELS:
            db.execute(delete(model))
        db.commit()
        yield db
    finally:
        for model in CLEANUP_MODELS:
            db.execute(delete(model))
        db.commit()
        db.close()


@pytest.fixture()
def client(
    db_session: Session, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> Iterator[TestClient]:
    get_settings.cache_clear()
    monkeypatch.setenv("LOCAL_STORAGE_ROOT", str(tmp_path / "storage"))
    monkeypatch.setenv("UPLOADS_DIR", str(tmp_path / "storage" / "uploads"))
    monkeypatch.setenv("ARTIFACTS_DIR", str(tmp_path / "storage" / "artifacts"))
    try:
        yield TestClient(app)
    finally:
        get_settings.cache_clear()



def make_png(path: Path, size: tuple[int, int] = (100, 80)) -> None:
    Image.new("RGB", size, color="white").save(path, format="PNG")


def strict_rubric() -> dict[str, object]:
    return {
        "total_marks": "5.00",
        "criteria": [
            {
                "id": "concept",
                "name": "Core concept",
                "description": "Identifies the correct principle or idea.",
                "max_marks": "3.00",
            },
            {
                "id": "clarity",
                "name": "Clarity",
                "description": "Explains the answer clearly.",
                "max_marks": "2.00",
            },
        ],
    }


def create_region_and_suggestion(client: TestClient, tmp_path: Path) -> dict[str, object]:
    email = f"review-{len(list(tmp_path.glob('*.png')))}@example.com"
    auth = register_auth_teacher(client, email)
    headers = auth_header(str(auth["access_token"]))
    teacher = auth["user"]
    course_response = client.post(
        "/courses",
        headers=headers,
        json={"code": "REV101", "title": "Review"},
    )
    assert course_response.status_code == 201
    assessment_response = client.post(
        f"/courses/{course_response.json()['id']}/assessments",
        headers=headers,
        json={"title": "Quiz", "assessment_type": "quiz", "total_marks": "5.00"},
    )
    assert assessment_response.status_code == 201
    assessment = assessment_response.json()
    question_response = client.post(
        f"/assessments/{assessment['id']}/questions",
        headers=headers,
        json={
            "question_no": "1(a)(i)",
            "question_text": "Explain.",
            "model_answer": "A complete explanation earns full marks.",
            "total_marks": "5.00",
        },
    )
    assert question_response.status_code == 201
    question = question_response.json()
    rubric_response = client.post(
        f"/questions/{question['id']}/rubrics",
        headers=headers,
        json={"version": 1, "rubric_json": strict_rubric(), "is_active": True},
    )
    assert rubric_response.status_code == 201

    image_path = tmp_path / f"review-source-{len(list(tmp_path.glob('*.png')))}.png"
    make_png(image_path)
    with image_path.open("rb") as file_obj:
        submission_response = client.post(
            f"/assessments/{assessment['id']}/submissions/upload",
            headers=headers,
            data={"student_identifier": "S-001", "student_name": "Student One"},
            files={"file": ("answer.png", file_obj, "image/png")},
        )
    assert submission_response.status_code == 201
    submission = submission_response.json()
    page = submission["pages"][0]
    region_response = client.post(
        f"/submission-pages/{page['id']}/answer-regions",
        headers=headers,
        json={
            "question_id": question["id"],
            "x": 1,
            "y": 2,
            "width": 20,
            "height": 25,
            "manual_answer_text": "2 + 2 = 4",
        },
    )
    assert region_response.status_code == 201
    region = region_response.json()
    grade_response = client.post(f"/answer-regions/{region['id']}/grade", headers=headers)
    assert grade_response.status_code == 201
    suggestion = grade_response.json()["suggestion"]
    return {
        "teacher": teacher,
        "assessment": assessment,
        "question": question,
        "submission": submission,
        "region": region,
        "suggestion": suggestion,
        "auth": auth,
        "headers": headers,
    }


def create_extra_region_and_suggestion(
    client: TestClient, data: dict[str, object], *, x: int
) -> dict[str, object]:
    submission = data["submission"]
    question = data["question"]
    headers = data["headers"]
    page = submission["pages"][0]
    region_response = client.post(
        f"/submission-pages/{page['id']}/answer-regions",
        headers=headers,
        json={
            "question_id": question["id"],
            "x": x,
            "y": 2,
            "width": 20,
            "height": 25,
            "manual_answer_text": "2 + 2 = 4",
        },
    )
    assert region_response.status_code == 201
    grade_response = client.post(
        f"/answer-regions/{region_response.json()['id']}/grade", headers=headers
    )
    assert grade_response.status_code == 201
    return {"region": region_response.json(), "suggestion": grade_response.json()["suggestion"]}


def auth_header(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def register_auth_teacher(
    client: TestClient, email: str = "review-auth@example.com"
) -> dict[str, object]:
    response = client.post(
        "/auth/register",
        json={"name": "Review Auth Teacher", "email": email, "password": "review-password"},
    )
    assert response.status_code == 201
    return response.json()


def create_owned_region_and_suggestion(
    client: TestClient, tmp_path: Path, email: str
) -> dict[str, object]:
    auth = register_auth_teacher(client, email)
    headers = auth_header(str(auth["access_token"]))
    course_response = client.post(
        "/courses",
        headers=headers,
        json={"code": "OWN101", "title": "Owned Review"},
    )
    assert course_response.status_code == 201
    assessment_response = client.post(
        f"/courses/{course_response.json()['id']}/assessments",
        headers=headers,
        json={"title": "Owned Quiz", "assessment_type": "quiz", "total_marks": "5.00"},
    )
    assert assessment_response.status_code == 201
    assessment = assessment_response.json()
    question_response = client.post(
        f"/assessments/{assessment['id']}/questions",
        headers=headers,
        json={
            "question_no": "1(a)(i)",
            "question_text": "Explain.",
            "model_answer": "A complete explanation earns full marks.",
            "total_marks": "5.00",
        },
    )
    assert question_response.status_code == 201
    question = question_response.json()
    rubric_response = client.post(
        f"/questions/{question['id']}/rubrics",
        headers=headers,
        json={"version": 1, "rubric_json": strict_rubric(), "is_active": True},
    )
    assert rubric_response.status_code == 201
    image_path = tmp_path / f"owned-{email}.png"
    make_png(image_path)
    with image_path.open("rb") as file_obj:
        submission_response = client.post(
            f"/assessments/{assessment['id']}/submissions/upload",
            headers=headers,
            data={"student_identifier": "S-OWN", "student_name": "Owned Student"},
            files={"file": ("answer.png", file_obj, "image/png")},
        )
    assert submission_response.status_code == 201
    submission = submission_response.json()
    page = submission["pages"][0]
    region_response = client.post(
        f"/submission-pages/{page['id']}/answer-regions",
        headers=headers,
        json={
            "question_id": question["id"],
            "x": 1,
            "y": 2,
            "width": 20,
            "height": 25,
            "manual_answer_text": "2 + 2 = 4",
        },
    )
    assert region_response.status_code == 201
    grade_response = client.post(
        f"/answer-regions/{region_response.json()['id']}/grade", headers=headers
    )
    assert grade_response.status_code == 201
    return {
        "auth": auth,
        "headers": headers,
        "assessment": assessment,
        "submission": submission,
        "question": question,
        "region": region_response.json(),
        "suggestion": grade_response.json()["suggestion"],
    }


def test_owner_required_for_review_queue_summary_export_and_final_grade_actions(
    client: TestClient, tmp_path: Path, db_session: Session
) -> None:
    owner = create_owned_region_and_suggestion(client, tmp_path, "owner-review@example.com")
    intruder = register_auth_teacher(client, "intruder-review@example.com")
    intruder_headers = auth_header(str(intruder["access_token"]))

    assert client.get(f"/assessments/{owner['assessment']['id']}/review-queue").status_code == 401
    assert client.get(
        f"/assessments/{owner['assessment']['id']}/review-queue", headers=intruder_headers
    ).status_code == 404
    assert client.get(
        f"/assessments/{owner['assessment']['id']}/summary", headers=intruder_headers
    ).status_code == 404
    assert client.get(
        f"/assessments/{owner['assessment']['id']}/export/final-grades.xlsx",
        headers=intruder_headers,
    ).status_code == 404
    assert client.post(
        f"/grade-suggestions/{owner['suggestion']['id']}/approve",
        headers=intruder_headers,
        json={"teacher_comment": "intruder"},
    ).status_code == 404

    owner_approve = client.post(
        f"/grade-suggestions/{owner['suggestion']['id']}/approve",
        headers=owner["headers"],
        json={"teacher_id": intruder["user"]["id"], "teacher_comment": "owner approval"},
    )
    assert owner_approve.status_code == 201
    assert owner_approve.json()["teacher_id"] == owner["auth"]["user"]["id"]
    logs = db_session.query(AuditLog).filter(AuditLog.event_type == "final_grade.approved").all()
    assert {log.actor_id for log in logs} == {owner["auth"]["user"]["id"]}


def test_owner_export_includes_only_approved_final_grades(
    client: TestClient, tmp_path: Path
) -> None:
    owner = create_owned_region_and_suggestion(client, tmp_path, "owner-export@example.com")
    second = create_extra_region_and_suggestion(client, owner, x=30)

    pending_export = client.get(
        f"/assessments/{owner['assessment']['id']}/export/final-grades.xlsx",
        headers=owner["headers"],
    )
    assert pending_export.status_code == 200
    pending_rows = list(
        load_workbook(BytesIO(pending_export.content)).active.iter_rows(values_only=True)
    )
    assert len(pending_rows) == 1

    approve = client.post(
        f"/grade-suggestions/{owner['suggestion']['id']}/approve",
        headers=owner["headers"],
        json={"teacher_comment": "approved"},
    )
    assert approve.status_code == 201

    response = client.get(
        f"/assessments/{owner['assessment']['id']}/export/final-grades.xlsx",
        headers=owner["headers"],
    )
    assert response.status_code == 200
    rows = list(load_workbook(BytesIO(response.content)).active.iter_rows(values_only=True))
    assert len(rows) == 2
    headers = list(rows[0])
    row = dict(zip(headers, rows[1], strict=True))
    assert row["grade_suggestion_id"] == owner["suggestion"]["id"]
    assert row["answer_region_id"] == owner["region"]["id"]
    assert second["suggestion"]["id"] not in [cell for row in rows for cell in row]


def test_approve_grade_suggestion_creates_final_grade(client: TestClient, tmp_path: Path) -> None:
    data = create_region_and_suggestion(client, tmp_path)

    response = client.post(
        f"/grade-suggestions/{data['suggestion']['id']}/approve",
        headers=data["headers"],
        json={
            "teacher_id": data["teacher"]["id"],
            "teacher_comment": "Approved after review.",
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["answer_region_id"] == data["region"]["id"]
    assert payload["suggestion_id"] == data["suggestion"]["id"]
    assert payload["teacher_id"] == data["teacher"]["id"]
    assert payload["final_score"] == data["suggestion"]["score"]
    assert payload["approval_status"] == "approved"
    assert payload["teacher_comment"] == "Approved after review."
    assert "password_hash" not in payload


def test_auth_aware_review_actions_use_current_teacher_and_require_valid_token(
    client: TestClient, tmp_path: Path
) -> None:
    data = create_region_and_suggestion(client, tmp_path)
    suggestion_id = data["suggestion"]["id"]

    missing = client.post(
        f"/grade-suggestions/{suggestion_id}/approve",
        json={"teacher_comment": "No token."},
    )
    invalid = client.post(
        f"/grade-suggestions/{suggestion_id}/approve",
        headers=auth_header("bad-token"),
        json={"teacher_comment": "Bad token."},
    )
    approved = client.post(
        f"/grade-suggestions/{suggestion_id}/approve",
        headers=data["headers"],
        json={"teacher_comment": "Reviewed as logged-in teacher."},
    )

    assert missing.status_code == 401
    assert invalid.status_code == 401
    assert approved.status_code == 201
    assert approved.json()["teacher_id"] == data["teacher"]["id"]
    assert approved.json()["teacher_comment"] == "Reviewed as logged-in teacher."


def test_edit_grade_suggestion_creates_final_grade_with_teacher_score(
    client: TestClient, tmp_path: Path
) -> None:
    data = create_region_and_suggestion(client, tmp_path)

    response = client.post(
        f"/grade-suggestions/{data['suggestion']['id']}/edit",
        headers=data["headers"],
        json={
            "teacher_id": data["teacher"]["id"],
            "final_score": "4.00",
            "teacher_comment": "Adjusted after inspecting work.",
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["final_score"] == "4.00"
    assert payload["approval_status"] == "edited"
    assert payload["teacher_comment"] == "Adjusted after inspecting work."


def test_reject_grade_suggestion_creates_rejected_final_grade(
    client: TestClient, tmp_path: Path
) -> None:
    data = create_region_and_suggestion(client, tmp_path)

    response = client.post(
        f"/grade-suggestions/{data['suggestion']['id']}/reject",
        headers=data["headers"],
        json={
            "teacher_id": data["teacher"]["id"],
            "teacher_comment": "Suggestion is not usable.",
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["final_score"] == "0.00"
    assert payload["approval_status"] == "rejected"
    assert payload["teacher_comment"] == "Suggestion is not usable."


def test_finalize_validation_failures(client: TestClient, tmp_path: Path) -> None:
    data = create_region_and_suggestion(client, tmp_path)
    suggestion_id = data["suggestion"]["id"]

    too_high = client.post(
        f"/grade-suggestions/{suggestion_id}/edit",
        headers=data["headers"],
        json={
            "teacher_id": data["teacher"]["id"],
            "final_score": "6.00",
        },
    )
    assert too_high.status_code == 422
    assert "max_score" in too_high.text

    spoofed_teacher = client.post(
        f"/grade-suggestions/{suggestion_id}/approve",
        headers=data["headers"],
        json={"teacher_id": 999999},
    )
    assert spoofed_teacher.status_code == 201
    assert spoofed_teacher.json()["teacher_id"] == data["teacher"]["id"]

    missing_suggestion = client.post(
        "/grade-suggestions/999999/approve",
        headers=data["headers"],
        json={"teacher_id": data["teacher"]["id"]},
    )
    assert missing_suggestion.status_code == 404


def test_get_final_grade_and_review_queue_states(client: TestClient, tmp_path: Path) -> None:
    data = create_region_and_suggestion(client, tmp_path)
    assessment_id = data["assessment"]["id"]
    region_id = data["region"]["id"]

    before_final = client.get(f"/assessments/{assessment_id}/review-queue", headers=data["headers"])
    assert before_final.status_code == 200
    queue = before_final.json()
    assert len(queue) == 1
    assert queue[0]["review_status"] == "suggested"
    assert queue[0]["latest_grade_suggestion"]["model_provider"] == "mock"
    assert queue[0]["final_grade"] is None
    assert queue[0]["submission"]["student_identifier"] == "S-001"

    not_found = client.get(f"/answer-regions/{region_id}/final-grade", headers=data["headers"])
    assert not_found.status_code == 404

    finalize = client.post(
        f"/grade-suggestions/{data['suggestion']['id']}/approve",
        headers=data["headers"],
        json={"teacher_id": data["teacher"]["id"]},
    )
    assert finalize.status_code == 201

    final_grade = client.get(f"/answer-regions/{region_id}/final-grade", headers=data["headers"])
    assert final_grade.status_code == 200
    assert final_grade.json()["id"] == finalize.json()["id"]

    after_final = client.get(f"/assessments/{assessment_id}/review-queue", headers=data["headers"])
    assert after_final.status_code == 200
    assert after_final.json()[0]["review_status"] == "finalized"


def test_review_queue_includes_ungraded_regions(client: TestClient, tmp_path: Path) -> None:
    data = create_region_and_suggestion(client, tmp_path)
    assessment_id = data["assessment"]["id"]
    page_id = data["submission"]["pages"][0]["id"]
    question_id = data["question"]["id"]
    second_region = client.post(
        f"/submission-pages/{page_id}/answer-regions",
        headers=data["headers"],
        json={"question_id": question_id, "x": 30, "y": 2, "width": 20, "height": 25},
    )
    assert second_region.status_code == 201

    queue = client.get(f"/assessments/{assessment_id}/review-queue", headers=data["headers"])

    assert queue.status_code == 200
    by_region = {item["answer_region"]["id"]: item for item in queue.json()}
    assert by_region[data["region"]["id"]]["review_status"] == "suggested"
    assert by_region[second_region.json()["id"]]["review_status"] == "ungraded"


def test_teacher_actions_replace_existing_current_final_grade(
    client: TestClient, tmp_path: Path
) -> None:
    data = create_region_and_suggestion(client, tmp_path)
    suggestion_id = data["suggestion"]["id"]
    first = client.post(
        f"/grade-suggestions/{suggestion_id}/approve",
        headers=data["headers"],
        json={"teacher_id": data["teacher"]["id"]},
    )
    assert first.status_code == 201

    second = client.post(
        f"/grade-suggestions/{suggestion_id}/edit",
        headers=data["headers"],
        json={
            "teacher_id": data["teacher"]["id"],
            "final_score": "3.00",
        },
    )

    assert second.status_code == 200
    assert second.json()["id"] == first.json()["id"]
    assert second.json()["final_score"] == "3.00"


def test_teacher_review_action_writes_audit_log(
    client: TestClient, tmp_path: Path, db_session: Session
) -> None:
    data = create_region_and_suggestion(client, tmp_path)

    response = client.post(
        f"/grade-suggestions/{data['suggestion']['id']}/approve",
        headers=data["headers"],
        json={"teacher_id": data["teacher"]["id"]},
    )

    assert response.status_code == 201
    logs = db_session.query(AuditLog).filter(AuditLog.event_type == "final_grade.approved").all()
    assert len(logs) == 1
    assert logs[0].actor_id == data["teacher"]["id"]
    assert logs[0].entity_type == "final_grade"


def test_assessment_summary_returns_review_counts(client: TestClient, tmp_path: Path) -> None:
    data = create_region_and_suggestion(client, tmp_path)
    second = create_extra_region_and_suggestion(client, data, x=30)
    third = create_extra_region_and_suggestion(client, data, x=60)
    page_id = data["submission"]["pages"][0]["id"]
    pending_region = client.post(
        f"/submission-pages/{page_id}/answer-regions",
        headers=data["headers"],
        json={"question_id": data["question"]["id"], "x": 70, "y": 2, "width": 20, "height": 25},
    )
    assert pending_region.status_code == 201
    assert client.post(
        f"/grade-suggestions/{data['suggestion']['id']}/approve",
        headers=data["headers"],
        json={"teacher_id": data["teacher"]["id"]},
    ).status_code == 201
    assert client.post(
        f"/grade-suggestions/{second['suggestion']['id']}/edit",
        headers=data["headers"],
        json={"teacher_id": data["teacher"]["id"], "final_score": "3.00"},
    ).status_code == 201
    assert client.post(
        f"/grade-suggestions/{third['suggestion']['id']}/reject",
        headers=data["headers"],
        json={"teacher_id": data["teacher"]["id"], "teacher_comment": "Reject"},
    ).status_code == 201

    response = client.get(
        f"/assessments/{data['assessment']['id']}/summary", headers=data["headers"]
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["assessment_id"] == data["assessment"]["id"]
    assert payload["course_id"] == data["assessment"]["course_id"]
    assert payload["total_submissions"] == 1
    assert payload["total_answer_regions"] == 4
    assert payload["total_grade_suggestions"] == 3
    assert payload["total_final_grades"] == 3
    assert payload["approved_count"] == 1
    assert payload["edited_count"] == 1
    assert payload["rejected_count"] == 1
    assert payload["pending_review_count"] == 1
    assert payload["average_final_score"] == "1.00"
    assert payload["max_possible_score"] == "20.00"
    assert payload["submission_totals"] == [
        {
            "submission_id": data["submission"]["id"],
            "student_identifier": "S-001",
            "student_name": "Student One",
            "approved_score": "0.00",
            "approved_max_score": "5.00",
            "assessment_max_score": "5.00",
            "approved_question_count": 1,
            "expected_question_count": 1,
            "is_complete": True,
        }
    ]
    assert "generated_at" in payload


def test_export_xlsx_contains_headers_rows_and_safe_fields(
    client: TestClient, tmp_path: Path
) -> None:
    data = create_region_and_suggestion(client, tmp_path)
    second = create_extra_region_and_suggestion(client, data, x=30)
    third = create_extra_region_and_suggestion(client, data, x=60)
    assert client.post(
        f"/grade-suggestions/{data['suggestion']['id']}/approve",
        headers=data["headers"],
        json={"teacher_id": data["teacher"]["id"], "teacher_comment": "Approved"},
    ).status_code == 201
    assert client.post(
        f"/grade-suggestions/{second['suggestion']['id']}/edit",
        headers=data["headers"],
        json={"teacher_id": data["teacher"]["id"], "final_score": "3.00"},
    ).status_code == 201
    assert client.post(
        f"/grade-suggestions/{third['suggestion']['id']}/reject",
        headers=data["headers"],
        json={"teacher_id": data["teacher"]["id"], "teacher_comment": "Rejected"},
    ).status_code == 201

    response = client.get(
        f"/assessments/{data['assessment']['id']}/export/final-grades.xlsx",
        headers=data["headers"],
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Final Grades", "Submission Totals"]
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = list(rows[0])
    assert headers == [
        "assessment_id",
        "course_id",
        "submission_id",
        "student_identifier",
        "student_name",
        "question_id",
        "question_no",
        "grading_unit_label",
        "grading_unit_max_marks",
        "answer_region_id",
        "grade_suggestion_id",
        "final_grade_id",
        "ai_score",
        "ai_max_score",
        "ai_confidence",
        "ai_needs_review",
        "marking_policy",
        "final_score",
        "approval_status",
        "teacher_comment",
        "reviewed_at",
        "feedback_to_student",
    ]
    assert len(rows) == 2
    exported_text = " ".join(str(cell) for row in rows for cell in row if cell is not None)
    first_row = dict(zip(headers, rows[1], strict=True))
    assert first_row["grading_unit_label"] == "1(a)(i)"
    assert first_row["grading_unit_max_marks"] == 5
    assert first_row["approval_status"] == "approved"
    assert "edited" not in exported_text
    assert "rejected" not in exported_text
    assert "raw_response_json" not in exported_text
    assert "password_hash" not in exported_text
    totals_rows = list(workbook["Submission Totals"].iter_rows(values_only=True))
    assert totals_rows[0] == (
        "assessment_id",
        "course_id",
        "submission_id",
        "student_identifier",
        "student_name",
        "approved_score",
        "approved_max_score",
        "assessment_max_score",
        "approved_question_count",
        "expected_question_count",
        "is_complete",
    )
    assert totals_rows[1][2] == data["submission"]["id"]
    assert totals_rows[1][5:11] == (0, 5, 5, 1, 1, True)


def test_assessment_summary_keeps_new_ungraded_submission_separate(
    client: TestClient, tmp_path: Path
) -> None:
    data = create_region_and_suggestion(client, tmp_path)
    assert client.post(
        f"/grade-suggestions/{data['suggestion']['id']}/approve",
        headers=data["headers"],
        json={"teacher_id": data["teacher"]["id"]},
    ).status_code == 201

    image_path = tmp_path / "second-submission.png"
    make_png(image_path)
    with image_path.open("rb") as file_obj:
        response = client.post(
            f"/assessments/{data['assessment']['id']}/submissions/upload",
            headers=data["headers"],
            data={"student_identifier": "S-002", "student_name": "Student Two"},
            files={"file": ("second.png", file_obj, "image/png")},
        )
    assert response.status_code == 201
    second_submission_id = response.json()["id"]

    summary_response = client.get(
        f"/assessments/{data['assessment']['id']}/summary", headers=data["headers"]
    )
    assert summary_response.status_code == 200
    totals = {
        item["submission_id"]: item
        for item in summary_response.json()["submission_totals"]
    }
    assert totals[data["submission"]["id"]]["approved_score"] == "0.00"
    assert totals[data["submission"]["id"]]["is_complete"] is True
    assert totals[second_submission_id]["approved_score"] == "0.00"
    assert totals[second_submission_id]["approved_question_count"] == 0
    assert totals[second_submission_id]["expected_question_count"] == 1
    assert totals[second_submission_id]["is_complete"] is False


def test_export_xlsx_excludes_pending_region_without_final_grade(
    client: TestClient, tmp_path: Path
) -> None:
    data = create_region_and_suggestion(client, tmp_path)

    response = client.get(
        f"/assessments/{data['assessment']['id']}/export/final-grades.xlsx",
        headers=data["headers"],
    )

    assert response.status_code == 200
    rows = list(load_workbook(BytesIO(response.content)).active.iter_rows(values_only=True))
    assert len(rows) == 1


def test_batch_approve_selected_suggestions_uses_auth_teacher_and_writes_summary(
    client: TestClient, tmp_path: Path, db_session: Session
) -> None:
    data = create_region_and_suggestion(client, tmp_path)
    second = create_extra_region_and_suggestion(client, data, x=30)
    response = client.post(
        f"/assessments/{data['assessment']['id']}/final-grades/approve-selected",
        headers=data["headers"],
        json={
            "grade_suggestion_ids": [
                data["suggestion"]["id"],
                second["suggestion"]["id"],
            ]
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["requested_count"] == 2
    assert payload["approved_count"] == 2
    assert payload["skipped_count"] == 0
    assert payload["failed_count"] == 0
    assert len(payload["final_grade_ids"]) == 2
    assert payload["errors"] == []
    final_grades = db_session.query(FinalGrade).order_by(FinalGrade.id).all()
    assert len(final_grades) == 2
    assert {grade.teacher_id for grade in final_grades} == {data["teacher"]["id"]}
    assert {grade.approval_status for grade in final_grades} == {"approved"}
    audit_count = db_session.query(AuditLog).filter(
        AuditLog.event_type == "final_grade.approved"
    ).count()
    assert audit_count == 2
    assert "raw_response_json" not in response.text
    assert "password_hash" not in response.text


def test_batch_approve_selected_skips_missing_and_outside_assessment_suggestions(
    client: TestClient, tmp_path: Path
) -> None:
    data = create_region_and_suggestion(client, tmp_path)
    outside = create_region_and_suggestion(client, tmp_path)
    response = client.post(
        f"/assessments/{data['assessment']['id']}/final-grades/approve-selected",
        headers=data["headers"],
        json={
            "grade_suggestion_ids": [
                data["suggestion"]["id"],
                outside["suggestion"]["id"],
                999999,
            ]
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["requested_count"] == 3
    assert payload["approved_count"] == 1
    assert payload["skipped_count"] == 2
    assert payload["failed_count"] == 0
    assert len(payload["final_grade_ids"]) == 1
    assert any("not found" in error for error in payload["errors"])
    assert any("does not belong to assessment" in error for error in payload["errors"])


def test_batch_approve_selected_requires_auth_and_does_not_duplicate_final_grade(
    client: TestClient, tmp_path: Path, db_session: Session
) -> None:
    data = create_region_and_suggestion(client, tmp_path)
    suggestion_id = data["suggestion"]["id"]
    missing_auth = client.post(
        f"/assessments/{data['assessment']['id']}/final-grades/approve-selected",
        json={"grade_suggestion_ids": [suggestion_id]},
    )
    assert missing_auth.status_code == 401

    first = client.post(
        f"/assessments/{data['assessment']['id']}/final-grades/approve-selected",
        headers=data["headers"],
        json={"grade_suggestion_ids": [suggestion_id]},
    )
    second = client.post(
        f"/assessments/{data['assessment']['id']}/final-grades/approve-selected",
        headers=data["headers"],
        json={"grade_suggestion_ids": [suggestion_id]},
    )

    assert first.status_code == 200
    assert second.status_code == 200
    assert first.json()["final_grade_ids"] == second.json()["final_grade_ids"]
    duplicate_count = db_session.query(FinalGrade).filter(
        FinalGrade.answer_region_id == data["region"]["id"]
    ).count()
    assert duplicate_count == 1


def test_summary_and_export_missing_assessment_returns_404(client: TestClient) -> None:
    auth = register_auth_teacher(client, "missing-summary@example.com")
    headers = auth_header(str(auth["access_token"]))
    assert client.get("/assessments/999999/summary", headers=headers).status_code == 404
    missing_export = client.get(
        "/assessments/999999/export/final-grades.xlsx", headers=headers
    )
    assert missing_export.status_code == 404
