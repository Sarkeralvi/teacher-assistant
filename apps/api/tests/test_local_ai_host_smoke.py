from __future__ import annotations

import json
import os
from collections.abc import Iterator
from io import BytesIO
from pathlib import Path
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.db.session import SessionLocal
from app.main import app
from app.models import AuditLog
from app.worker.jobs import run_grading_dispatch_job
from tests.test_cohort_grading_api import CLEANUP_MODELS, FakeQueue

pytestmark = pytest.mark.skipif(
    os.environ.get("RUN_LOCAL_AI_HOST_SMOKE") != "1",
    reason="set RUN_LOCAL_AI_HOST_SMOKE=1 to exercise installed host models",
)


@pytest.fixture()
def db_session() -> Iterator[Session]:
    db = SessionLocal()
    database_name = db.get_bind().url.database or ""
    if not database_name.endswith("_test"):
        db.close()
        pytest.fail("The local AI host smoke requires a disposable *_test database")
    try:
        for model in CLEANUP_MODELS:
            db.execute(delete(model))
        db.commit()
        yield db
    finally:
        db.rollback()
        for model in CLEANUP_MODELS:
            db.execute(delete(model))
        db.commit()
        db.close()


def _local_config() -> dict[str, str]:
    repository_root = Path(__file__).resolve().parents[3]
    config_path = repository_root / ".env.local-ai"
    if not config_path.is_file():
        pytest.fail("Run Initialize-LocalAiConfig.ps1 before the host smoke")
    config: dict[str, str] = {}
    for line in config_path.read_text(encoding="utf-8").splitlines():
        if line and not line.startswith("#"):
            key, value = line.split("=", 1)
            config[key] = value
    return config


def _expect(response: Any, status_code: int) -> dict[str, Any]:
    assert response.status_code == status_code, response.text[:2000]
    return response.json()


def _synthetic_answer(path: Path, answer: str) -> None:
    image = Image.new("RGB", (1600, 700), color="white")
    draw = ImageDraw.Draw(image)
    font_path = Path(os.environ.get("WINDIR", "C:\\Windows")) / "Fonts" / "arial.ttf"
    font = ImageFont.truetype(str(font_path), 58)
    draw.multiline_text(
        (80, 90),
        f"Question 1\n{answer}",
        fill="black",
        font=font,
        spacing=28,
    )
    image.save(path, format="PNG")


def test_two_student_local_ocr_qwen_review_and_export_host_smoke(
    db_session: Session,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    config = _local_config()
    for key in (
        "BRAIN_ALLOW_REAL_PROVIDERS",
        "LOCAL_QWEN_ENABLED",
        "LOCAL_QWEN_BASE_URL",
        "LOCAL_QWEN_MODEL",
        "LOCAL_QWEN_API_KEY",
        "LOCAL_OCR_ENABLED",
        "LOCAL_OCR_BASE_URL",
        "LOCAL_OCR_API_KEY",
    ):
        monkeypatch.setenv(key, config[key])
    monkeypatch.setenv("BRAIN_PROVIDER", "mock")
    monkeypatch.setenv("LOCAL_QWEN_TIMEOUT_SECONDS", "480")
    monkeypatch.setenv("LOCAL_OCR_TIMEOUT_SECONDS", "480")
    monkeypatch.setenv("COHORT_MODEL_GRADING_ENABLED", "true")
    monkeypatch.setenv("COHORT_MAX_PROVIDER_CALLS", "25")
    monkeypatch.setenv("COHORT_PROVIDER_RETRY_COUNT", "0")
    monkeypatch.setenv("LOCAL_STORAGE_ROOT", str(tmp_path / "storage"))
    monkeypatch.setenv("UPLOADS_DIR", str(tmp_path / "storage" / "uploads"))
    monkeypatch.setenv("ARTIFACTS_DIR", str(tmp_path / "storage" / "artifacts"))
    fake_queue = FakeQueue()
    monkeypatch.setattr("app.api.routes.grading.get_default_queue", lambda: fake_queue)
    get_settings.cache_clear()

    with TestClient(app) as client:
        registration = _expect(
            client.post(
                "/auth/register",
                json={
                    "name": "Local AI Smoke Teacher",
                    "email": "local-ai-host-smoke@example.com",
                    "password": "host-smoke-password",
                },
            ),
            201,
        )
        headers = {"Authorization": f"Bearer {registration['access_token']}"}

        local_status = _expect(client.get("/local-ai/status", headers=headers), 200)
        assert local_status["qwen"]["available"] is True
        assert local_status["ocr"]["available"] is True
        serialized_status = json.dumps(local_status)
        assert config["LOCAL_QWEN_API_KEY"] not in serialized_status
        assert config["LOCAL_OCR_API_KEY"] not in serialized_status
        assert config["LOCAL_QWEN_MODEL_PATH"] not in serialized_status

        course = _expect(
            client.post(
                "/courses",
                headers=headers,
                json={"code": "LOCAL101", "title": "Local AI smoke"},
            ),
            201,
        )
        assessment = _expect(
            client.post(
                f"/courses/{course['id']}/assessments",
                headers=headers,
                json={
                    "title": "Synthetic local assessment",
                    "assessment_type": "exam",
                    "total_marks": "5.00",
                },
            ),
            201,
        )
        grading_run = _expect(
            client.post(
                f"/assessments/{assessment['id']}/grading-runs/custom",
                headers=headers,
                json={"marking_policy": "general"},
            ),
            201,
        )
        question = _expect(
            client.post(
                f"/assessments/{assessment['id']}/questions",
                headers=headers,
                json={
                    "question_no": "1",
                    "question_text": "Calculate 12 + 8 and state the final answer.",
                    "model_answer": "12 + 8 = 20.",
                    "total_marks": "5.00",
                },
            ),
            201,
        )
        _expect(
            client.post(
                f"/questions/{question['id']}/rubrics",
                headers=headers,
                json={
                    "version": 1,
                    "is_active": True,
                    "rubric_json": {
                        "total_marks": "5.00",
                        "criteria": [
                            {
                                "id": "calculation",
                                "name": "Correct calculation",
                                "description": "Adds 12 and 8 and states the result.",
                                "max_marks": "5.00",
                            }
                        ],
                    },
                },
            ),
            201,
        )

        region_ids: list[int] = []
        intended_answers = (
            "12 + 8 = 20. Therefore, the answer is 20.",
            "12 + 8 = 18. Therefore, the answer is 18.",
        )
        for index, intended_answer in enumerate(intended_answers, start=1):
            image_path = tmp_path / f"synthetic-student-{index}.png"
            _synthetic_answer(image_path, intended_answer)
            with image_path.open("rb") as image_file:
                submission = _expect(
                    client.post(
                        f"/assessments/{assessment['id']}/submissions/upload",
                        headers=headers,
                        data={"student_identifier": f"SYN-{index:03d}"},
                        files={"file": (image_path.name, image_file, "image/png")},
                    ),
                    201,
                )
            region = _expect(
                client.post(
                    f"/submission-pages/{submission['pages'][0]['id']}/answer-regions",
                    headers=headers,
                    json={
                        "question_id": question["id"],
                        "x": 0,
                        "y": 0,
                        "width": 1600,
                        "height": 700,
                    },
                ),
                201,
            )
            region_ids.append(region["id"])
            ocr_run = _expect(
                client.post(f"/answer-regions/{region['id']}/ocr-runs", headers=headers),
                201,
            )
            assert ocr_run["status"] == "succeeded"
            assert ocr_run["draft_text"]
            _expect(
                client.post(
                    f"/answer-regions/{region['id']}/ocr-runs/{ocr_run['id']}/confirm",
                    headers=headers,
                    json={"confirmed_text": intended_answer},
                ),
                200,
            )
            confirmation = _expect(
                client.patch(
                    f"/answer-regions/{region['id']}/corrections/full-answer-confirmation",
                    headers=headers,
                    json={
                        "full_answer_confirmed": True,
                        "packet_status": "complete",
                        "manual_answer_text": intended_answer,
                    },
                ),
                200,
            )
            assert confirmation["answer_region"]["full_answer_confirmed"] is True

        queue_run = _expect(
            client.post(
                f"/assessments/{assessment['id']}/grading-queue-runs",
                headers=headers,
                json={},
            ),
            201,
        )
        assert queue_run["queued_item_count"] == 2
        dispatch_url = (
            f"/assessments/{assessment['id']}/questions/{question['id']}/grade-cohort"
        )
        dispatch_payload = {
            "queue_run_id": queue_run["id"],
            "grading_run_id": grading_run["id"],
            "provider": "llama_cpp_qwen",
            "expected_model": config["LOCAL_QWEN_MODEL"],
            "call_limit": 2,
            "draft_only_confirmed": True,
        }
        preflight = _expect(
            client.post(
                f"{dispatch_url}/preflight", headers=headers, json=dispatch_payload
            ),
            200,
        )
        assert preflight["selected_call_count"] == 2
        dispatch = _expect(
            client.post(dispatch_url, headers=headers, json=dispatch_payload), 202
        )
        assert len(fake_queue.enqueued) == 1

        run_grading_dispatch_job(dispatch["id"])

        finished = _expect(
            client.get(f"/grading-dispatch-runs/{dispatch['id']}", headers=headers),
            200,
        )
        assert finished["status"] == "completed"
        assert finished["succeeded_count"] == 2
        assert finished["calls_started"] == 2
        assert finished["uncertain_count"] == 0

        review_queue = _expect(
            client.get(
                f"/assessments/{assessment['id']}/review-queue", headers=headers
            ),
            200,
        )
        assert len(review_queue) == 2
        assert {item["answer_region"]["id"] for item in review_queue} == set(region_ids)
        for item in review_queue:
            suggestion = item["latest_grade_suggestion"]
            assert suggestion["model_provider"] == "llama_cpp_qwen"
            assert suggestion["needs_review"] is True
            _expect(
                client.post(
                    f"/grade-suggestions/{suggestion['id']}/approve",
                    headers=headers,
                    json={"teacher_comment": "Synthetic host smoke approval"},
                ),
                201,
            )

        exported = client.get(
            f"/assessments/{assessment['id']}/export/final-grades.xlsx",
            headers=headers,
        )
        assert exported.status_code == 200
        rows = list(
            load_workbook(BytesIO(exported.content)).active.iter_rows(values_only=True)
        )
        assert len(rows) == 3
        assert {row[3] for row in rows[1:]} == {"SYN-001", "SYN-002"}

        audit_payloads = db_session.scalars(select(AuditLog.payload_json)).all()
        audit_text = json.dumps(audit_payloads, default=str)
        assert all(answer not in audit_text for answer in intended_answers)

    get_settings.cache_clear()
