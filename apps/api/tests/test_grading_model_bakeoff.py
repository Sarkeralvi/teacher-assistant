from __future__ import annotations

from collections.abc import Iterator
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.db.session import SessionLocal
from app.models import AnswerRegionOcrRun, AuditLog, GradeSuggestion, GradingJob
from packages.evaluation import grading_model_bakeoff as bakeoff
from packages.evaluation import local_curated_evaluation as evaluation
from tests.test_cohort_grading_api import CLEANUP_MODELS

_INTEGRATION_COMMIT = "d" * 40
_HARNESS_COMMIT = "e" * 40


@pytest.fixture()
def clean_database() -> Iterator[Session]:
    db = SessionLocal()
    try:
        for model in CLEANUP_MODELS:
            db.execute(delete(model))
        db.commit()
        yield db
    finally:
        db.rollback()
        for model in CLEANUP_MODELS:
            db.execute(delete(model))
        db.commit()
        db.close()
        get_settings.cache_clear()


def _assets(model_alias: str) -> evaluation.OperatorAssetMetadata:
    model_hash = "1" * 64 if model_alias.startswith("qwen3.6") else "2" * 64
    return evaluation.OperatorAssetMetadata.model_validate(
        {
            "llama_cpp": {
                "build": "10622",
                "model_alias": model_alias,
                "model_sha256": model_hash,
                "model_size_bytes": 1024,
                "device": (
                    "gpu_hybrid" if model_alias.startswith("qwen3.6") else "gpu_hybrid_single_slot"
                ),
            },
            "qwen38_vision": {
                "build": "10622",
                "model_alias": "qwen3.8-27b-q4km",
                "model_sha256": "2" * 64,
                "model_size_bytes": 2048,
                "device": "gpu_hybrid_single_slot",
                "mmproj_sha256": "3" * 64,
                "mmproj_size_bytes": 1024,
            },
        }
    )


def _complete_ground_truth(run_dir: Path) -> evaluation.GroundTruthLock:
    manifest = evaluation.load_manifest(run_dir)
    workbook_path = run_dir / "ground_truth_review.xlsx"
    workbook = load_workbook(workbook_path)
    sheet = workbook["Cases"]
    headers = {str(cell.value): index for index, cell in enumerate(sheet[1], start=1)}
    for row, case in enumerate(manifest.cases, start=2):
        sheet.cell(row, headers["teacher_transcription"], case.authored_transcription)
        sheet.cell(row, headers["teacher_score"], str(case.expected_score))
        sheet.cell(row, headers["teacher_notes"], "Teacher-locked bake-off source")
        if case.primary_category == evaluation.PrimaryCategory.DIFFICULT_HANDWRITING:
            sheet.cell(row, headers["handwriting_acceptable"], "yes")
        sheet.cell(row, headers["teacher_approved"], "yes")
    workbook.save(workbook_path)
    return evaluation.lock_ground_truth(
        run_dir,
        reviewer_id="teacher-1",
        confirm_teacher_signoff=True,
    )


def _write_confirmed_visual_source(
    run_dir: Path,
    ground_truth: evaluation.GroundTruthLock,
) -> None:
    manifest = evaluation.load_manifest(run_dir)
    first_call_at = datetime.now(UTC) + timedelta(seconds=1)
    completed_at = first_call_at + timedelta(seconds=1)
    ocr_cases = [
        evaluation.OcrCaseResult(
            case_id=case.case_id,
            answer_region_id=index,
            ocr_run_id=100 + index,
            status="succeeded",
            draft_text=case.authored_transcription,
            markdown=case.authored_transcription,
            blocks=[],
            warnings=[],
            latency_ms=1000,
            provider="llama_cpp_qwen38",
            model="qwen3.8-27b-q4km",
            profile="qwen38_verbatim_visual",
            reasoning_mode="off",
            source_image_sha256=evaluation.sha256_text(case.image_sha256),
            source_image_hashes=[case.image_sha256],
            is_blank=case.answer_quality == evaluation.AnswerQuality.BLANK,
            draft_text_sha256=evaluation.sha256_text(case.authored_transcription),
        )
        for index, case in enumerate(manifest.cases, start=1)
    ]
    result = evaluation.OcrRunResult(
        run_id=manifest.run_id,
        first_call_at=first_call_at,
        completed_at=completed_at,
        call_count=20,
        service_status_before={"qwen38": {"available": True}},
        service_status_after={"qwen38": {"available": True}},
        database_name=f"teacher_assistant_eval_{manifest.run_id}",
        assessment_id=1,
        grading_run_id=1,
        owner_teacher_id=1,
        intruder_teacher_id=2,
        question_ids={pack: index for index, pack in enumerate("ABCDE", start=1)},
        rubric_ids={pack: index for index, pack in enumerate("ABCDE", start=11)},
        cases=ocr_cases,
    )
    result_path = run_dir / "ocr_results.json"
    runtime_path = run_dir / "ocr_runtime.json"
    evaluation.write_json(result_path, result)
    evaluation.write_json(
        runtime_path,
        {
            "visual_reasoning_mode": "off",
            "gpu_safety_before": {"single_slot": True},
            "gpu_safety_after": {"single_slot": True},
        },
    )
    workbook_path = evaluation.create_ocr_review_workbook(
        manifest,
        ground_truth,
        result,
        run_dir,
    )
    workbook = load_workbook(workbook_path)
    sheet = workbook["OCR Review"]
    headers = {str(cell.value): index for index, cell in enumerate(sheet[1], start=1)}
    for row in range(2, 22):
        sheet.cell(row, headers["teacher_confirms_faithful"], "yes")
        sheet.cell(row, headers["teacher_approved"], "yes")
    workbook.save(workbook_path)
    confirmations = [
        evaluation.OcrConfirmationCase(
            case_id=case.case_id,
            answer_region_id=ocr_case.answer_region_id,
            ocr_run_id=ocr_case.ocr_run_id,
            confirmed_text=ocr_case.draft_text,
            confirmed_text_sha256=ocr_case.draft_text_sha256,
            teacher_approved=True,
            evidence_status=(
                "blank" if case.answer_quality == evaluation.AnswerQuality.BLANK else "complete"
            ),
            full_answer_confirmed=case.answer_quality != evaluation.AnswerQuality.BLANK,
        )
        for case, ocr_case in zip(manifest.cases, result.cases, strict=True)
    ]
    confirmation_lock = evaluation.OcrConfirmationLock(
        run_id=manifest.run_id,
        reviewer_id="teacher-1",
        signed_at=completed_at + timedelta(seconds=1),
        ocr_results_sha256=evaluation.sha256_file(result_path),
        workbook_sha256=evaluation.sha256_file(workbook_path),
        cases=confirmations,
    )
    confirmation_path = run_dir / "ocr_confirmation_lock.json"
    evaluation.write_json(confirmation_path, confirmation_lock)
    evaluation.append_state(
        run_dir,
        "ocr_completed",
        locked_artifacts={
            "ocr_results.json": evaluation.sha256_file(result_path),
            "ocr_runtime.json": evaluation.sha256_file(runtime_path),
        },
    )
    evaluation.append_state(
        run_dir,
        "ocr_confirmed",
        locked_artifacts={
            "ocr_review.xlsx": evaluation.sha256_file(workbook_path),
            "ocr_confirmation_lock.json": evaluation.sha256_file(confirmation_path),
        },
    )


def test_fork_replays_one_teacher_confirmed_visual_source_without_model_calls(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source_run_dir = evaluation.prepare_evaluation(
        run_id="bakeoff_source",
        output_root=tmp_path,
        integration_commit=_INTEGRATION_COMMIT,
        harness_commit=_HARNESS_COMMIT,
        operator_assets=_assets("qwen3.6-35b-a3b-q4km"),
        seed=12345,
    )
    ground_truth = _complete_ground_truth(source_run_dir)
    _write_confirmed_visual_source(source_run_dir, ground_truth)
    monkeypatch.setattr(evaluation, "require_clean_git_worktree", lambda: None)
    monkeypatch.setattr(evaluation, "current_git_commit", lambda: _HARNESS_COMMIT)

    manifest = bakeoff.create_grading_bakeoff(
        source_run_dir=source_run_dir,
        operator_assets_for_model=_assets,
    )

    assert manifest.source.source_run_id == "bakeoff_source"
    assert manifest.source.replay_visual_call_count == 0
    assert [candidate.model_alias for candidate in manifest.candidates] == [
        "qwen3.6-35b-a3b-q4km",
        "qwen3.8-27b-q4km",
    ]
    assert evaluation.current_state(source_run_dir) == "ocr_confirmed"
    for candidate in manifest.candidates:
        candidate_dir = tmp_path / candidate.run_id
        assert evaluation.current_state(candidate_dir) == "ground_truth_locked"
        assert (candidate_dir / "source_evidence_lock.json").is_file()
        assert (candidate_dir / "source_ocr_results.json").is_file()
        assert not (candidate_dir / "grading_results.json").exists()
        candidate_manifest = evaluation.load_manifest(candidate_dir)
        assert candidate_manifest.expected_qwen_model == candidate.model_alias
        assert [case.image_sha256 for case in candidate_manifest.cases] == [
            case.image_sha256 for case in evaluation.load_manifest(source_run_dir).cases
        ]


def test_qwen38_promotion_requires_material_gain_severity_and_latency() -> None:
    threshold = evaluation.EvaluationThresholds()
    baseline = bakeoff.CandidateSummary(
        label="qwen36",
        model_alias="qwen3.6-35b-a3b-q4km",
        grading_results_sha256="1" * 64,
        review_lock_sha256="2" * 64,
        reference_bundle_sha256="3" * 64,
        confirmed_text_hashes={case_id: "4" * 64 for case_id in evaluation._EXPECTED_CASE_IDS},
        process_checks_pass=True,
        teacher_review_pass=True,
        exact_count=10,
        within_one_count=17,
        mean_absolute_error=Decimal("0.60"),
        mean_normalized_absolute_error=Decimal("0.10"),
        severe_false_confident_count=0,
        severe_low_confidence_count=0,
        formula_multistep_within_one=True,
        irrelevant_over_limit_count=0,
        wrong_over_half_count=0,
        zero_reference_overscore_count=0,
        p95_latency_ms=Decimal("10000"),
    )
    qwen38 = baseline.model_copy(
        update={
            "label": "qwen38",
            "model_alias": "qwen3.8-27b-q4km",
            "exact_count": 12,
            "mean_absolute_error": Decimal("0.45"),
            "p95_latency_ms": Decimal("20000"),
        }
    )

    assert bakeoff._candidate_quality_passes(baseline, threshold)[0] is True
    assert bakeoff._candidate_quality_passes(qwen38, threshold)[0] is True
    assert qwen38.exact_count >= baseline.exact_count + 2
    assert qwen38.p95_latency_ms <= baseline.p95_latency_ms * 2

    too_slow = qwen38.model_copy(update={"p95_latency_ms": Decimal("20001")})
    assert too_slow.p95_latency_ms > baseline.p95_latency_ms * 2


def test_seed_replays_locked_visual_evidence_without_a_model_call(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    clean_database: Session,
) -> None:
    del clean_database
    source_run_dir = evaluation.prepare_evaluation(
        run_id="bakeoff_seed_source",
        output_root=tmp_path,
        integration_commit=_INTEGRATION_COMMIT,
        harness_commit=_HARNESS_COMMIT,
        operator_assets=_assets("qwen3.6-35b-a3b-q4km"),
        seed=12345,
    )
    ground_truth = _complete_ground_truth(source_run_dir)
    _write_confirmed_visual_source(source_run_dir, ground_truth)
    monkeypatch.setattr(evaluation, "require_clean_git_worktree", lambda: None)
    monkeypatch.setattr(evaluation, "current_git_commit", lambda: _HARNESS_COMMIT)
    manifest = bakeoff.create_grading_bakeoff(
        source_run_dir=source_run_dir,
        operator_assets_for_model=_assets,
    )
    candidate = next(item for item in manifest.candidates if item.label == "qwen36")
    candidate_dir = tmp_path / candidate.run_id
    storage_root = tmp_path / "candidate_storage"
    monkeypatch.setenv("BRAIN_PROVIDER", "mock")
    monkeypatch.setenv("BRAIN_ALLOW_REAL_PROVIDERS", "true")
    monkeypatch.setenv("LOCAL_QWEN_ENABLED", "true")
    monkeypatch.setenv("LOCAL_QWEN_MODEL", "qwen3.6-35b-a3b-q4km")
    monkeypatch.setenv("LOCAL_QWEN38_ENABLED", "true")
    monkeypatch.setenv("LOCAL_QWEN38_MODEL", "qwen3.8-27b-q4km")
    monkeypatch.setenv("LOCAL_QWEN38_VISUAL_PREPARATION_ENABLED", "true")
    monkeypatch.setenv("LOCAL_QWEN38_GRADING_ENABLED", "true")
    monkeypatch.setenv("COHORT_MODEL_GRADING_ENABLED", "true")
    monkeypatch.setenv("COHORT_MAX_PROVIDER_CALLS", "25")
    monkeypatch.setenv("COHORT_PROVIDER_RETRY_COUNT", "0")
    monkeypatch.setenv("LOCAL_AI_PHASE_SWITCH_ENABLED", "false")
    monkeypatch.setenv("LOCAL_STORAGE_ROOT", str(storage_root))
    monkeypatch.setenv("UPLOADS_DIR", str(storage_root / "uploads"))
    monkeypatch.setenv("ARTIFACTS_DIR", str(storage_root / "artifacts"))
    get_settings.cache_clear()
    settings = get_settings()
    monkeypatch.setattr(
        evaluation,
        "_configure_runtime",
        lambda *_args, **_kwargs: (settings, SessionLocal, storage_root),
    )
    database_url = (
        "postgresql+psycopg://postgres@127.0.0.1:55432/"
        f"teacher_assistant_eval_{candidate.run_id}"
    )

    confirmation_lock = bakeoff.seed_grading_bakeoff_candidate(
        candidate_run_dir=candidate_dir,
        database_url=database_url,
    )

    assert evaluation.current_state(candidate_dir) == "ocr_confirmed"
    candidate_ocr = evaluation.OcrRunResult.model_validate(
        evaluation.read_json(candidate_dir / "ocr_results.json")
    )
    assert candidate_ocr.evidence_origin == "locked_bakeoff_replay"
    assert candidate_ocr.source_evaluation_run_id == "bakeoff_seed_source"
    assert candidate_ocr.call_count == 20
    assert candidate_ocr.new_provider_call_count == 0
    assert len(confirmation_lock.cases) == 20
    assert all(
        case.confirmed_text_sha256 == manifest.source.confirmed_text_hashes[case.case_id]
        for case in confirmation_lock.cases
    )
    with SessionLocal() as db:
        assert int(db.scalar(select(func.count(AnswerRegionOcrRun.id))) or 0) == 20
        assert int(db.scalar(select(func.count(GradingJob.id))) or 0) == 0
        assert int(db.scalar(select(func.count(GradeSuggestion.id))) or 0) == 0
        replay_runs = list(db.scalars(select(AnswerRegionOcrRun)).all())
        assert all(run.status == "confirmed" and run.calls_used == 0 for run in replay_runs)
        audits = list(db.scalars(select(AuditLog.payload_json)).all())
    audit_text = str(audits)
    assert "teacher_confirmed_source_evidence_replay" not in audit_text
    assert not any(
        case.confirmed_text in audit_text
        for case in confirmation_lock.cases
        if case.confirmed_text
    )


def _advance_candidate_to_teacher_review(
    candidate_dir: Path,
    candidate: bakeoff.GradingBakeoffCandidate,
) -> None:
    manifest = evaluation.load_manifest(candidate_dir)
    source_ocr = evaluation.OcrRunResult.model_validate(
        evaluation.read_json(candidate_dir / "source_ocr_results.json")
    )
    source_confirmations = evaluation.OcrConfirmationLock.model_validate(
        evaluation.read_json(candidate_dir / "source_ocr_confirmation_lock.json")
    )
    now = datetime.now(UTC)
    ocr = source_ocr.model_copy(
        update={
            "run_id": manifest.run_id,
            "evidence_origin": "locked_bakeoff_replay",
            "source_evaluation_run_id": source_ocr.run_id,
            "new_provider_call_count": 0,
            "database_name": f"teacher_assistant_eval_{manifest.run_id}",
            "first_call_at": now - timedelta(minutes=4),
            "completed_at": now - timedelta(minutes=3),
        }
    )
    ocr_path = candidate_dir / "ocr_results.json"
    runtime_path = candidate_dir / "ocr_runtime.json"
    evaluation.write_json(ocr_path, ocr)
    evaluation.write_json(runtime_path, {"visual_reasoning_mode": "off"})
    confirmations = source_confirmations.model_copy(
        update={
            "run_id": manifest.run_id,
            "signed_at": now - timedelta(minutes=2),
            "ocr_results_sha256": evaluation.sha256_file(ocr_path),
        }
    )
    confirmation_path = candidate_dir / "ocr_confirmation_lock.json"
    evaluation.write_json(confirmation_path, confirmations)
    evaluation.append_state(
        candidate_dir,
        "ocr_completed",
        locked_artifacts={
            "ocr_results.json": evaluation.sha256_file(ocr_path),
            "ocr_runtime.json": evaluation.sha256_file(runtime_path),
        },
    )
    evaluation.append_state(
        candidate_dir,
        "ocr_confirmed",
        locked_artifacts={
            "ocr_confirmation_lock.json": evaluation.sha256_file(confirmation_path),
        },
    )
    next_item_id = 1
    grading_cases: list[evaluation.GradingCaseResult] = []
    for index, case in enumerate(manifest.cases, start=1):
        if case.answer_quality == evaluation.AnswerQuality.BLANK:
            grading_cases.append(
                evaluation.GradingCaseResult(
                    case_id=case.case_id,
                    answer_region_id=index,
                    outcome="not_called_blank_safety_gate",
                    max_score=case.max_score,
                    confirmed_text_sha256=evaluation.sha256_text(case.authored_transcription),
                )
            )
            continue
        grading_cases.append(
            evaluation.GradingCaseResult(
                case_id=case.case_id,
                answer_region_id=index,
                outcome="suggested",
                dispatch_run_id=ord(case.pack_id) - ord("A") + 1,
                dispatch_item_id=next_item_id,
                grading_job_id=next_item_id,
                grade_suggestion_id=next_item_id,
                ai_score=case.expected_score,
                max_score=case.max_score,
                confidence=Decimal("0.60"),
                needs_review=True,
                rubric_breakdown=[
                    {
                        "criterion_id": "synthetic",
                        "criterion": "Synthetic criterion",
                        "max_marks": str(case.max_score),
                        "awarded_marks": str(case.expected_score),
                    }
                ],
                review_flags=[
                    "image_input_disabled",
                    "local_provider",
                    "teacher_review_required",
                ],
                model_provider=(
                    "llama_cpp_qwen"
                    if candidate.label == "qwen36"
                    else "llama_cpp_qwen38"
                ),
                model_name=candidate.model_alias,
                prompt_version="real-grading-v3",
                marking_policy="general",
                token_usage={
                    "prompt_tokens": 10,
                    "completion_tokens": 10,
                    "total_tokens": 20,
                },
                latency_ms=1000 if candidate.label == "qwen36" else 2000,
                cost_estimate=Decimal("0"),
                confirmed_text_sha256=evaluation.sha256_text(case.authored_transcription),
            )
        )
        next_item_id += 1
    grading_result = evaluation.GradingRunResult(
        run_id=manifest.run_id,
        first_call_at=now - timedelta(minutes=1),
        completed_at=now,
        qwen_call_count=18,
        blank_refusal_count=2,
        dispatch_run_ids=[1, 2, 3, 4, 5],
        cases=grading_cases,
        safety_checks={key: True for key in evaluation._REQUIRED_GRADING_SAFETY_CHECKS},
    )
    grading_path = candidate_dir / "grading_results.json"
    grading_runtime_path = candidate_dir / "grading_runtime.json"
    evaluation.write_json(grading_path, grading_result)
    evaluation.write_json(grading_runtime_path, {"image_input_to_grading": False})
    evaluation.append_state(
        candidate_dir,
        "grading_completed",
        locked_artifacts={
            "grading_results.json": evaluation.sha256_file(grading_path),
            "grading_runtime.json": evaluation.sha256_file(grading_runtime_path),
        },
    )
    grading_workbook_path = evaluation.create_grading_review_workbook(
        manifest,
        grading_result,
        candidate_dir,
    )
    review = evaluation.ReviewLock(
        run_id=manifest.run_id,
        reviewer_id="teacher-1",
        signed_at=now + timedelta(seconds=1),
        grading_results_sha256=evaluation.sha256_file(grading_path),
        workbook_sha256=evaluation.sha256_file(grading_workbook_path),
        cases=[
            evaluation.GradingReviewCase(
                case_id=case.case_id,
                disagreement_reason="none",
                useful_draft=case.answer_quality != evaluation.AnswerQuality.BLANK,
                approved_review=True,
            )
            for case in manifest.cases
        ],
    )
    review_path = candidate_dir / "review_lock.json"
    evaluation.write_json(review_path, review)
    evaluation.append_state(
        candidate_dir,
        "review_completed",
        locked_artifacts={
            "grading_review.xlsx": evaluation.sha256_file(grading_workbook_path),
            "review_lock.json": evaluation.sha256_file(review_path),
        },
    )


def test_compare_requires_reviewed_candidates_and_keeps_qwen36_without_material_gain(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source_run_dir = evaluation.prepare_evaluation(
        run_id="bakeoff_compare_source",
        output_root=tmp_path,
        integration_commit=_INTEGRATION_COMMIT,
        harness_commit=_HARNESS_COMMIT,
        operator_assets=_assets("qwen3.6-35b-a3b-q4km"),
        seed=12345,
    )
    ground_truth = _complete_ground_truth(source_run_dir)
    _write_confirmed_visual_source(source_run_dir, ground_truth)
    monkeypatch.setattr(evaluation, "require_clean_git_worktree", lambda: None)
    monkeypatch.setattr(evaluation, "current_git_commit", lambda: _HARNESS_COMMIT)
    manifest = bakeoff.create_grading_bakeoff(
        source_run_dir=source_run_dir,
        operator_assets_for_model=_assets,
    )
    for candidate in manifest.candidates:
        _advance_candidate_to_teacher_review(tmp_path / candidate.run_id, candidate)

    report = bakeoff.compare_grading_bakeoff(source_run_dir=source_run_dir)

    assert report.verdict == bakeoff.GradingBakeoffVerdict.QWEN36_RETAINED
    assert report.recommended_grading_model == "qwen3.6-35b-a3b-q4km"
    assert [candidate.model_alias for candidate in report.candidates] == [
        "qwen3.6-35b-a3b-q4km",
        "qwen3.8-27b-q4km",
    ]
    assert (source_run_dir / "grading_bakeoff" / "report.md").is_file()
