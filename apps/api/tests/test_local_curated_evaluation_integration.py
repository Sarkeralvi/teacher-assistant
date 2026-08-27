from __future__ import annotations

from collections.abc import Iterator
from decimal import Decimal
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy import delete
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.db.session import SessionLocal
from packages.brain.adapter import BrainAdapter
from packages.brain.provider_base import BrainProvider
from packages.brain.schemas import GradeSuggestionOutput, RubricBreakdownItem
from packages.brain.schemas_qwen38 import VisualTranscriptionOutput
from packages.evaluation import local_curated_evaluation as evaluation
from tests.test_cohort_grading_api import CLEANUP_MODELS


class LocalOcrResult(BaseModel):
    request_id: str
    mode: str
    text: str
    normalized_text: str
    markdown: str
    blocks: list[dict[str, Any]]
    warnings: list[str]
    provider: str
    model: str
    layout_model: str
    version: str
    device: str
    latency_ms: int


class FakeLocalQwenProvider(BrainProvider):
    provider_name = "llama_cpp_qwen"
    model_name = "qwen3.6-35b-a3b-q4km"

    def __init__(self) -> None:
        self.calls: list[str] = []

    def grade(
        self,
        *,
        question_text: str,
        question_total_marks: Decimal,
        rubric_json: dict[str, Any],
        answer_image_path: str,
        prompt_version: str,
        student_answer_text: str | None = None,
        task_name: str = "answer_region_grading",
        model_policy: object | None = None,
        messages: list[dict[str, Any]] | None = None,
        image_data_url: str | None = None,
        marking_policy: str = "general",
    ) -> GradeSuggestionOutput:
        del question_text, task_name, model_policy, messages
        assert answer_image_path == "[image input disabled]"
        assert image_data_url is None
        assert student_answer_text
        assert prompt_version == "real-grading-v3"
        assert marking_policy == "general"
        self.calls.append(student_answer_text)
        criteria = rubric_json["criteria"]
        breakdown = [
            RubricBreakdownItem(
                criterion_id=str(criterion["id"]),
                criterion=str(criterion["name"]),
                max_marks=Decimal(str(criterion["max_marks"])),
                awarded_marks=Decimal("0"),
                reason="Fake provider deliberately awards zero for pipeline rehearsal.",
                evidence=None,
                confidence=Decimal("0.50"),
            )
            for criterion in criteria
        ]
        return GradeSuggestionOutput(
            score=Decimal("0"),
            max_score=Decimal(str(rubric_json.get("total_marks", question_total_marks))),
            confidence=Decimal("0.50"),
            needs_review=True,
            rubric_breakdown=breakdown,
            detected_answer_summary="Synthetic fake-provider result.",
            major_errors=[],
            feedback_to_student="Teacher review is required.",
            review_flags=[
                "image_input_disabled",
                "local_provider",
                "teacher_review_required",
            ],
            model_provider=self.provider_name,
            model_name=self.model_name,
            prompt_version=prompt_version,
            cost_estimate=Decimal("0"),
            latency_ms=1,
            prompt_tokens=10,
            completion_tokens=10,
            total_tokens=20,
        )


class FakeQwen38VisualProvider(BrainProvider):
    """Fake only the local transport; preserve the production visual job path."""

    provider_name = "llama_cpp_qwen38"
    model_name = "qwen3.8-27b-q4km"

    def __init__(self, transcription_by_image_sha256: dict[str, str]) -> None:
        self.transcription_by_image_sha256 = transcription_by_image_sha256
        self.calls: list[list[str]] = []

    def grade(
        self,
        **_kwargs: object,
    ) -> GradeSuggestionOutput:
        raise AssertionError("The Qwen3.8 visual fake must never be used for grading")

    def transcribe_images(
        self,
        *,
        images: list[tuple[bytes, str]],
        label: str = "answer",
    ) -> VisualTranscriptionOutput:
        import hashlib

        del label
        image_hashes = [hashlib.sha256(image_bytes).hexdigest() for image_bytes, _ in images]
        self.calls.append(image_hashes)
        text = self.transcription_by_image_sha256[image_hashes[0]]
        return VisualTranscriptionOutput(
            draft_text=text,
            uncertain_glyphs=[],
            is_blank=not text,
            is_irrelevant=False,
            confidence=Decimal("0.90"),
            needs_review=True,
            model_provider=self.provider_name,
            model_name=self.model_name,
            image_sha256=image_hashes[0],
            latency_ms=1,
            prompt_tokens=10,
            completion_tokens=10,
        )


class FakeQwen38VisualAndGradingProvider(FakeQwen38VisualProvider):
    """Qwen3.8 fake used to prove visual work and text grading stay separate."""

    def __init__(self, transcription_by_image_sha256: dict[str, str]) -> None:
        super().__init__(transcription_by_image_sha256)
        self.grading_calls: list[str] = []

    def grade(
        self,
        *,
        question_text: str,
        question_total_marks: Decimal,
        rubric_json: dict[str, Any],
        answer_image_path: str,
        prompt_version: str,
        student_answer_text: str | None = None,
        task_name: str = "answer_region_grading",
        model_policy: object | None = None,
        messages: list[dict[str, Any]] | None = None,
        image_data_url: str | None = None,
        marking_policy: str = "general",
    ) -> GradeSuggestionOutput:
        del question_text, task_name, model_policy, messages
        assert answer_image_path == "[image input disabled]"
        assert image_data_url is None
        assert student_answer_text
        assert prompt_version == "real-grading-v3"
        assert marking_policy == "general"
        self.grading_calls.append(student_answer_text)
        breakdown = [
            RubricBreakdownItem(
                criterion_id=str(criterion["id"]),
                criterion=str(criterion["name"]),
                max_marks=Decimal(str(criterion["max_marks"])),
                awarded_marks=Decimal("0"),
                reason="Fake provider deliberately awards zero for pipeline rehearsal.",
                evidence=None,
                confidence=Decimal("0.50"),
            )
            for criterion in rubric_json["criteria"]
        ]
        return GradeSuggestionOutput(
            score=Decimal("0"),
            max_score=Decimal(str(rubric_json.get("total_marks", question_total_marks))),
            confidence=Decimal("0.50"),
            needs_review=True,
            rubric_breakdown=breakdown,
            detected_answer_summary="Synthetic fake-provider result.",
            major_errors=[],
            feedback_to_student="Teacher review is required.",
            review_flags=[
                "image_input_disabled",
                "local_provider",
                "teacher_review_required",
            ],
            model_provider=self.provider_name,
            model_name=self.model_name,
            prompt_version=prompt_version,
            cost_estimate=Decimal("0"),
            latency_ms=1,
            prompt_tokens=10,
            completion_tokens=10,
            total_tokens=20,
        )


@pytest.fixture()
def clean_database() -> Iterator[Session]:
    db = SessionLocal()
    try:
        for model in CLEANUP_MODELS:
            db.execute(delete(model))
        db.commit()
        yield db
    finally:
        db.rollback()
        for model in CLEANUP_MODELS:
            db.execute(delete(model))
        db.commit()
        db.close()
        get_settings.cache_clear()


def _assets(grading_model: str) -> evaluation.OperatorAssetMetadata:
    return evaluation.OperatorAssetMetadata.model_validate(
        {
            "llama_cpp": {
                "build": "10249",
                "model_alias": grading_model,
                "model_sha256": (
                    "1" * 64 if grading_model == "qwen3.6-35b-a3b-q4km" else "2" * 64
                ),
                "model_size_bytes": 1024 if grading_model == "qwen3.6-35b-a3b-q4km" else 2048,
                "device": (
                    "gpu_hybrid"
                    if grading_model == "qwen3.6-35b-a3b-q4km"
                    else "gpu_hybrid_single_slot"
                ),
            },
            "qwen38_vision": {
                "build": "10249",
                "model_alias": "qwen3.8-27b-q4km",
                "model_sha256": "2" * 64,
                "model_size_bytes": 2048,
                "device": "gpu_hybrid_single_slot",
                "mmproj_sha256": "3" * 64,
                "mmproj_size_bytes": 1024,
            },
        }
    )


def _complete_ground_truth(run_dir: Path) -> None:
    manifest = evaluation.load_manifest(run_dir)
    workbook_path = run_dir / "ground_truth_review.xlsx"
    workbook = load_workbook(workbook_path)
    sheet = workbook["Cases"]
    headers = {str(cell.value): index for index, cell in enumerate(sheet[1], start=1)}
    for row, case in enumerate(manifest.cases, start=2):
        sheet.cell(row, headers["teacher_transcription"], case.authored_transcription)
        sheet.cell(row, headers["teacher_score"], str(case.expected_score))
        sheet.cell(row, headers["teacher_notes"], "Independent fake-provider rehearsal")
        if case.primary_category == evaluation.PrimaryCategory.DIFFICULT_HANDWRITING:
            sheet.cell(row, headers["handwriting_acceptable"], "yes")
        sheet.cell(row, headers["teacher_approved"], "yes")
    workbook.save(workbook_path)


def _complete_ocr_review(run_dir: Path) -> None:
    workbook_path = run_dir / "ocr_review.xlsx"
    workbook = load_workbook(workbook_path)
    sheet = workbook["OCR Review"]
    headers = {str(cell.value): index for index, cell in enumerate(sheet[1], start=1)}
    for row in range(2, 22):
        sheet.cell(row, headers["teacher_confirms_faithful"], "yes")
        sheet.cell(row, headers["teacher_notes"], "Confirmed in fake-provider rehearsal")
        sheet.cell(row, headers["teacher_approved"], "yes")
    workbook.save(workbook_path)


def _complete_grading_review(run_dir: Path) -> None:
    manifest = evaluation.load_manifest(run_dir)
    result = evaluation.GradingRunResult.model_validate(
        evaluation.read_json(run_dir / "grading_results.json")
    )
    definitions = {case.case_id: case for case in manifest.cases}
    results = {case.case_id: case for case in result.cases}
    workbook_path = run_dir / "grading_review.xlsx"
    workbook = load_workbook(workbook_path)
    sheet = workbook["Grading Review"]
    headers = {str(cell.value): index for index, cell in enumerate(sheet[1], start=1)}
    for row in range(2, 22):
        case_id = str(sheet.cell(row, headers["case_id"]).value)
        case_result = results[case_id]
        definition = definitions[case_id]
        if case_result.outcome == "not_called_blank_safety_gate":
            reason = "none"
            useful = "no"
        else:
            reason = "none" if case_result.ai_score == definition.expected_score else "model_error"
            useful = "yes"
        sheet.cell(row, headers["disagreement_reason"], reason)
        sheet.cell(row, headers["teacher_notes"], "Fake-provider rehearsal review")
        sheet.cell(row, headers["useful_draft"], useful)
        sheet.cell(row, headers["approved_review"], "yes")
    workbook.save(workbook_path)


def _status(**_kwargs: Any) -> dict[str, Any]:
    return {
        "real_providers_allowed": True,
        "cohort_model_grading_enabled": True,
        "qwen": {
            "enabled": True,
            "available": True,
            "provider": "llama_cpp_qwen",
            "model": "qwen3.6-35b-a3b-q4km",
            "layout_model": None,
            "device": "gpu_hybrid",
            "detail": "ready",
        },
        "qwen38": {
            "enabled": True,
            "available": True,
            "provider": "llama_cpp_qwen38",
            "model": "qwen3.8-27b-q4km",
            "layout_model": None,
            "device": "gpu_hybrid_single_slot",
            "detail": "ready",
            "visual_preparation_enabled": True,
            "grading_enabled": True,
        },
    }


@pytest.mark.parametrize(
    "grading_model",
    ["qwen3.6-35b-a3b-q4km"],
)
def test_full_harness_rehearsal_uses_only_fake_providers(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    clean_database: Session,
    grading_model: str,
) -> None:
    del clean_database
    run_id = f"fake_full_rehearsal_{'q36' if grading_model.startswith('qwen3.6') else 'q38'}"
    assets = _assets(grading_model)
    run_dir = evaluation.prepare_evaluation(
        run_id=run_id,
        output_root=tmp_path / "evaluation",
        integration_commit="d" * 40,
        harness_commit="e" * 40,
        operator_assets=assets,
        seed=12345,
    )
    _complete_ground_truth(run_dir)
    evaluation.lock_ground_truth(
        run_dir,
        reviewer_id="fake-teacher",
        confirm_teacher_signoff=True,
    )

    runtime_storage = run_dir / "runtime_storage"
    monkeypatch.setenv("BRAIN_PROVIDER", "mock")
    monkeypatch.setenv("BRAIN_ALLOW_REAL_PROVIDERS", "true")
    monkeypatch.setenv("LOCAL_QWEN_ENABLED", "true")
    monkeypatch.setenv("LOCAL_QWEN_API_KEY", "fake-qwen-key")
    monkeypatch.setenv("LOCAL_QWEN_MODEL", "qwen3.6-35b-a3b-q4km")
    monkeypatch.setenv("LOCAL_QWEN38_ENABLED", "true")
    monkeypatch.setenv("LOCAL_QWEN38_API_KEY", "fake-qwen38-key")
    monkeypatch.setenv("LOCAL_QWEN38_MODEL", "qwen3.8-27b-q4km")
    monkeypatch.setenv("LOCAL_QWEN38_TRANSCRIPTION_ENABLED", "true")
    monkeypatch.setenv("LOCAL_QWEN38_VISUAL_PREPARATION_ENABLED", "true")
    monkeypatch.setenv("LOCAL_QWEN38_GRADING_ENABLED", "true")
    monkeypatch.setenv("LOCAL_AI_PHASE_SWITCH_ENABLED", "false")
    monkeypatch.setenv("COHORT_MODEL_GRADING_ENABLED", "true")
    monkeypatch.setenv("LOCAL_CURATED_EVALUATION_RUN_ID", run_id)
    monkeypatch.setenv("COHORT_MAX_PROVIDER_CALLS", "25")
    monkeypatch.setenv("COHORT_PROVIDER_RETRY_COUNT", "0")
    monkeypatch.setenv("LOCAL_STORAGE_ROOT", str(runtime_storage))
    monkeypatch.setenv("UPLOADS_DIR", str(runtime_storage / "uploads"))
    monkeypatch.setenv("ARTIFACTS_DIR", str(runtime_storage / "artifacts"))
    get_settings.cache_clear()
    settings = get_settings()
    monkeypatch.setattr(evaluation, "require_clean_git_worktree", lambda: None)
    monkeypatch.setattr(evaluation, "current_git_commit", lambda: "e" * 40)
    monkeypatch.setattr(evaluation, "_sanitized_status", _status)
    monkeypatch.setattr(
        evaluation,
        "_gpu_safety_snapshot",
        lambda **_kwargs: {
            "active_model_present_in_gpu_compute_clients": True,
            "other_managed_model_absent_from_gpu_compute_clients": True,
            "gpu_compute_client_names": ["fake-llama-server.exe"],
        },
    )
    monkeypatch.setattr(evaluation, "_operator_asset_metadata", lambda _expected: assets)
    monkeypatch.setattr(
        evaluation,
        "_configure_runtime",
        lambda *_args, **_kwargs: (settings, SessionLocal, runtime_storage),
    )
    database_url = (
        "postgresql+psycopg://postgres@127.0.0.1:55432/"
        f"teacher_assistant_eval_{run_id}"
    )

    manifest = evaluation.load_manifest(run_dir)
    transcription_by_image_sha256 = {
        evaluation.sha256_file(run_dir / case.image_relative_path): case.authored_transcription
        for case in manifest.cases
    }
    fake_visual: FakeQwen38VisualProvider
    fake_qwen: FakeLocalQwenProvider | None = None
    if grading_model == "qwen3.8-27b-q4km":
        fake_visual = FakeQwen38VisualAndGradingProvider(transcription_by_image_sha256)
    else:
        fake_visual = FakeQwen38VisualProvider(transcription_by_image_sha256)
        fake_qwen = FakeLocalQwenProvider()
    visual_adapter = BrainAdapter(
        fake_visual,
        image_input_enabled=False,
        storage_root=str(runtime_storage),
    )
    grading_adapter = (
        visual_adapter
        if grading_model == "qwen3.8-27b-q4km"
        else BrainAdapter(
            fake_qwen,
            image_input_enabled=False,
            storage_root=str(runtime_storage),
        )
    )

    def fake_for_provider(
        cls: type[BrainAdapter],
        provider_settings: Any,
        requested_provider: str,
    ) -> BrainAdapter:
        del cls, provider_settings
        if requested_provider == "llama_cpp_qwen38":
            return grading_adapter if grading_model == "qwen3.8-27b-q4km" else visual_adapter
        if requested_provider == "llama_cpp_qwen":
            return grading_adapter
        raise AssertionError(f"Unexpected provider: {requested_provider}")

    monkeypatch.setattr(BrainAdapter, "for_provider", classmethod(fake_for_provider))

    ocr_result = evaluation.run_ocr_stage(
        run_dir,
        allow_local_ocr=True,
        max_ocr_calls=20,
        database_url=database_url,
    )
    assert ocr_result.retry_count == 0
    assert len(fake_visual.calls) == 20
    assert evaluation.current_state(run_dir) == "ocr_completed"

    _complete_ocr_review(run_dir)
    confirmations = evaluation.lock_ocr_confirmations(
        run_dir,
        reviewer_id="fake-teacher",
        confirm_teacher_signoff=True,
        database_url=database_url,
    )
    assert len([case for case in confirmations.cases if case.evidence_status == "complete"]) == 18
    assert len([case for case in confirmations.cases if case.evidence_status == "blank"]) == 2

    monkeypatch.delenv("LOCAL_CURATED_EVALUATION_RUN_ID")
    with pytest.raises(evaluation.LocalCuratedEvaluationError, match="explicitly scoped"):
        evaluation.run_grading_stage(
            run_dir,
            allow_local_qwen=True,
            max_qwen_calls=18,
            expected_model=grading_model,
            database_url=database_url,
        )
    assert evaluation.current_state(run_dir) == "ocr_confirmed"
    assert fake_qwen is None or fake_qwen.calls == []
    if isinstance(fake_visual, FakeQwen38VisualAndGradingProvider):
        assert fake_visual.grading_calls == []
    monkeypatch.setenv("LOCAL_CURATED_EVALUATION_RUN_ID", run_id)

    grading_result = evaluation.run_grading_stage(
        run_dir,
        allow_local_qwen=True,
        max_qwen_calls=18,
        expected_model=grading_model,
        database_url=database_url,
    )
    if grading_model == "qwen3.8-27b-q4km":
        assert isinstance(fake_visual, FakeQwen38VisualAndGradingProvider)
        assert len(fake_visual.grading_calls) == 18
        assert all(fake_visual.grading_calls)
    else:
        assert fake_qwen is not None
        assert len(fake_qwen.calls) == 18
        assert all(fake_qwen.calls)
    assert grading_result.qwen_call_count == 18
    assert grading_result.blank_refusal_count == 2
    assert grading_result.retry_count == 0
    # Local Qwen grading is text-only: it must not create unused cropped or
    # composite student-image artifacts before dispatching the provider call.
    assert not (runtime_storage / "artifacts" / "grading_context").exists()
    assert evaluation.current_state(run_dir) == "grading_completed"

    _complete_grading_review(run_dir)
    evaluation.lock_grading_review(
        run_dir,
        reviewer_id="fake-teacher",
        confirm_teacher_signoff=True,
    )
    report = evaluation.generate_report(run_dir)
    assert report["verdict"] == evaluation.EvaluationVerdict.NO_GO_QUALITY.value
    assert evaluation.current_state(run_dir) == "reported"
    assert report["process_checks"]["exactly_eighteen_qwen_calls"] is True
    assert report["process_checks"]["two_blank_policy_refusals"] is True
    assert not (run_dir / "invalid_run.json").exists()
    get_settings.cache_clear()
