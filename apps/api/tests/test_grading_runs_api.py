from collections.abc import Iterator
from pathlib import Path
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from PIL import Image
from sqlalchemy import delete
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.db.session import SessionLocal
from app.main import app
from app.models import (
    AnswerRegion,
    Assessment,
    Course,
    ExtractionRun,
    FinalGrade,
    GradeSuggestion,
    GradingJob,
    GradingRun,
    Question,
    QuestionImportJob,
    QuestionNode,
    Rubric,
    RubricExtractionCriterion,
    Submission,
    SubmissionPage,
    User,
)

CLEANUP_MODELS = (
    FinalGrade,
    GradeSuggestion,
    GradingJob,
    AnswerRegion,
    SubmissionPage,
    Submission,
    Rubric,
    Question,
    RubricExtractionCriterion,
    QuestionNode,
    ExtractionRun,
    QuestionImportJob,
    GradingRun,
    Assessment,
    Course,
    User,
)


@pytest.fixture()
def db_session() -> Iterator[Session]:
    db = SessionLocal()
    try:
        for model in CLEANUP_MODELS:
            db.execute(delete(model))
        db.commit()
        yield db
    finally:
        for model in CLEANUP_MODELS:
            db.execute(delete(model))
        db.commit()
        db.close()


@pytest.fixture()
def client(
    db_session: Session, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> Iterator[TestClient]:
    get_settings.cache_clear()
    monkeypatch.setenv("LOCAL_STORAGE_ROOT", str(tmp_path / "storage"))
    monkeypatch.setenv("UPLOADS_DIR", str(tmp_path / "storage" / "uploads"))
    monkeypatch.setenv("ARTIFACTS_DIR", str(tmp_path / "storage" / "artifacts"))
    try:
        yield TestClient(app)
    finally:
        get_settings.cache_clear()


def register_teacher(
    client: TestClient, email_prefix: str = "run"
) -> tuple[dict[str, object], str]:
    response = client.post(
        "/auth/register",
        json={
            "name": "Teacher",
            "email": f"{email_prefix}-{uuid4().hex}@example.com",
            "password": "correct-horse-battery",
        },
    )
    assert response.status_code == 201
    body = response.json()
    return body["user"], body["access_token"]


def create_assessment_for_teacher(
    client: TestClient, teacher_id: int, token: str
) -> dict[str, object]:
    headers = {"Authorization": f"Bearer {token}"}
    course_response = client.post(
        "/courses",
        headers=headers,
        json={"teacher_id": teacher_id, "code": "MATH101", "title": "Math"},
    )
    assert course_response.status_code == 201
    assessment_response = client.post(
        f"/courses/{course_response.json()['id']}/assessments",
        headers=headers,
        json={"title": "Midterm", "assessment_type": "exam", "total_marks": "50.00"},
    )
    assert assessment_response.status_code == 201
    return assessment_response.json()


def make_question_paper_pdf(path: Path) -> None:
    Image.new("RGB", (160, 120), color="white").save(path, format="PDF")


class CapturingQueue:
    def __init__(self) -> None:
        self.calls: list[tuple[object, tuple[object, ...], dict[str, object]]] = []

    def enqueue(self, function: object, *args: object, **kwargs: object) -> None:
        self.calls.append((function, args, kwargs))


def test_create_and_list_custom_grading_run_requires_auth_and_assessment_scope(
    client: TestClient,
) -> None:
    teacher, token = register_teacher(client)
    assessment = create_assessment_for_teacher(client, int(teacher["id"]), token)

    unauthenticated = client.post(f"/assessments/{assessment['id']}/grading-runs/custom")
    assert unauthenticated.status_code == 401

    response = client.post(
        f"/assessments/{assessment['id']}/grading-runs/custom",
        headers={"Authorization": f"Bearer {token}"},
        json={"notes": "Teacher will provide solution and rubric PDFs."},
    )

    assert response.status_code == 201
    run = response.json()
    assert run["assessment_id"] == assessment["id"]
    assert run["created_by_teacher_id"] == teacher["id"]
    assert run["mode"] == "custom_controlled"
    assert run["status"] == "draft"
    assert run["notes"] == "Teacher will provide solution and rubric PDFs."
    assert run["question_pdf_path"] is None
    assert run["solution_pdf_path"] is None
    assert run["rubric_pdf_path"] is None

    listed = client.get(
        f"/assessments/{assessment['id']}/grading-runs",
        headers={"Authorization": f"Bearer {token}"},
    ).json()
    assert [item["id"] for item in listed] == [run["id"]]

    detail = client.get(f"/grading-runs/{run['id']}", headers={"Authorization": f"Bearer {token}"})
    assert detail.status_code == 200
    assert detail.json()["id"] == run["id"]


def test_create_custom_grading_run_missing_assessment_returns_404(client: TestClient) -> None:
    _, token = register_teacher(client)

    response = client.post(
        "/assessments/999999/grading-runs/custom",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 404


def test_upload_materials_stores_safe_relative_pdf_paths_and_updates_status(
    client: TestClient,
) -> None:
    teacher, token = register_teacher(client)
    assessment = create_assessment_for_teacher(client, int(teacher["id"]), token)
    run_response = client.post(
        f"/assessments/{assessment['id']}/grading-runs/custom",
        headers={"Authorization": f"Bearer {token}"},
    )
    run_id = run_response.json()["id"]

    response = client.post(
        f"/grading-runs/{run_id}/materials",
        headers={"Authorization": f"Bearer {token}"},
        files={
            "question_pdf": ("question.pdf", b"%PDF-1.4\n%question", "application/pdf"),
            "solution_pdf": ("solution.pdf", b"%PDF-1.4\n%solution", "application/pdf"),
            "rubric_pdf": ("rubric.pdf", b"%PDF-1.4\n%rubric", "application/pdf"),
        },
    )

    assert response.status_code == 200
    run = response.json()
    assert run["status"] == "materials_uploaded"
    for field in ("question_pdf_path", "solution_pdf_path", "rubric_pdf_path"):
        value = run[field]
        assert isinstance(value, str)
        assert value.endswith(".pdf")
        assert not value.startswith("/")
        assert ".." not in Path(value).parts
        stored_path = Path(get_settings().local_storage_root) / value
        assert stored_path.exists()


def test_material_upload_rejects_non_pdf_and_wrong_teacher(client: TestClient) -> None:
    owner, owner_token = register_teacher(client, "owner")
    assessment = create_assessment_for_teacher(client, int(owner["id"]), owner_token)
    run = client.post(
        f"/assessments/{assessment['id']}/grading-runs/custom",
        headers={"Authorization": f"Bearer {owner_token}"},
    ).json()
    _, other_token = register_teacher(client, "other")

    wrong_teacher = client.post(
        f"/grading-runs/{run['id']}/materials",
        headers={"Authorization": f"Bearer {other_token}"},
        files={"question_pdf": ("question.pdf", b"%PDF-1.4", "application/pdf")},
    )
    assert wrong_teacher.status_code == 404

    unsupported = client.post(
        f"/grading-runs/{run['id']}/materials",
        headers={"Authorization": f"Bearer {owner_token}"},
        files={"question_pdf": ("question.txt", b"not a pdf", "text/plain")},
    )
    assert unsupported.status_code == 415

    traversal_name = client.post(
        f"/grading-runs/{run['id']}/materials",
        headers={"Authorization": f"Bearer {owner_token}"},
        files={"question_pdf": ("../../question.pdf", b"%PDF-1.4", "application/pdf")},
    )
    assert traversal_name.status_code == 200
    assert ".." not in Path(traversal_name.json()["question_pdf_path"]).parts


def test_reference_extraction_route_uses_confirmed_bundle_and_enqueues_once(
    client: TestClient, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setenv("BRAIN_ALLOW_REAL_PROVIDERS", "true")
    monkeypatch.setenv("LOCAL_QWEN38_ENABLED", "true")
    monkeypatch.setenv("LOCAL_QWEN38_API_KEY", "key-local-test")
    monkeypatch.setenv("LOCAL_QWEN38_VISUAL_PREPARATION_ENABLED", "true")
    monkeypatch.setenv("LOCAL_REFERENCE_EXTRACTION_ENABLED", "true")
    get_settings.cache_clear()
    queue = CapturingQueue()
    monkeypatch.setattr(
        "app.api.routes.grading_runs.get_default_queue", lambda: queue
    )

    teacher, token = register_teacher(client, "reference-owner")
    assessment = create_assessment_for_teacher(client, int(teacher["id"]), token)
    run = client.post(
        f"/assessments/{assessment['id']}/grading-runs/custom",
        headers={"Authorization": f"Bearer {token}"},
    ).json()
    upload = client.post(
        f"/grading-runs/{run['id']}/materials",
        headers={"Authorization": f"Bearer {token}"},
        files={
            "question_pdf": ("Question.pdf", b"%PDF question", "application/pdf"),
            "solution_pdf": ("Solution.pdf", b"%PDF solution", "application/pdf"),
            "rubric_pdf": ("rubric.pdf", b"%PDF rubric", "application/pdf"),
        },
    )
    assert upload.status_code == 200
    assert upload.json()["question_pdf_name"] == "Question.pdf"

    mismatch = client.post(
        f"/grading-runs/{run['id']}/reference-extraction",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "provider": "llama_cpp_qwen38",
            "expected_model": "wrong-model",
            "materials_confirmed": True,
            "draft_only_confirmed": True,
        },
    )
    assert mismatch.status_code == 409
    assert queue.calls == []

    started = client.post(
        f"/grading-runs/{run['id']}/reference-extraction",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "provider": "llama_cpp_qwen38",
            "expected_model": "qwen3.8-27b-q4km",
            "materials_confirmed": True,
            "draft_only_confirmed": True,
        },
    )
    assert started.status_code == 202
    assert started.json()["status"] == "queued"
    assert started.json()["questions"] == []
    assert len(queue.calls) == 1
    assert queue.calls[0][1] == (run["id"],)
    assert queue.calls[0][2] == {"job_timeout": 2400}

    _, intruder_token = register_teacher(client, "reference-intruder")
    hidden = client.get(
        f"/grading-runs/{run['id']}/reference-extraction",
        headers={"Authorization": f"Bearer {intruder_token}"},
    )
    assert hidden.status_code == 404


def test_status_update_allows_controlled_workflow_statuses(client: TestClient) -> None:
    teacher, token = register_teacher(client)
    assessment = create_assessment_for_teacher(client, int(teacher["id"]), token)
    run = client.post(
        f"/assessments/{assessment['id']}/grading-runs/custom",
        headers={"Authorization": f"Bearer {token}"},
    ).json()

    response = client.patch(
        f"/grading-runs/{run['id']}",
        headers={"Authorization": f"Bearer {token}"},
        json={"status": "questions_ready", "notes": "Questions and rubrics confirmed."},
    )

    assert response.status_code == 200
    assert response.json()["status"] == "questions_ready"
    assert response.json()["notes"] == "Questions and rubrics confirmed."

    invalid = client.patch(
        f"/grading-runs/{run['id']}",
        headers={"Authorization": f"Bearer {token}"},
        json={"status": "fully_automated"},
    )
    assert invalid.status_code == 422


def upload_all_materials(client: TestClient, token: str, run_id: int) -> dict[str, object]:
    response = client.post(
        f"/grading-runs/{run_id}/materials",
        headers={"Authorization": f"Bearer {token}"},
        files={
            "question_pdf": ("question.pdf", b"%PDF-1.4\n%question", "application/pdf"),
            "solution_pdf": ("solution.pdf", b"%PDF-1.4\n%solution", "application/pdf"),
            "rubric_pdf": ("rubric.pdf", b"%PDF-1.4\n%rubric", "application/pdf"),
        },
    )
    assert response.status_code == 200
    return response.json()


def create_question_and_active_rubric(
    client: TestClient, assessment_id: int, token: str
) -> dict[str, object]:
    headers = {"Authorization": f"Bearer {token}"}
    question_response = client.post(
        f"/assessments/{assessment_id}/questions",
        headers=headers,
        json={
            "question_no": "1",
            "question_text": "Explain the method.",
            "model_answer": "A complete answer with reasoning.",
            "total_marks": "5.00",
        },
    )
    assert question_response.status_code == 201
    question = question_response.json()
    rubric_response = client.post(
        f"/questions/{question['id']}/rubrics",
        headers=headers,
        json={
            "version": 1,
            "is_active": True,
            "rubric_json": {
                "total_marks": "5.00",
                "criteria": [
                    {
                        "id": "reasoning",
                        "name": "Reasoning",
                        "description": "Shows valid reasoning.",
                        "max_marks": "5.00",
                    }
                ],
            },
        },
    )
    assert rubric_response.status_code == 201
    return {"question": question, "rubric": rubric_response.json()}


def upload_script_and_create_region(
    client: TestClient, token: str, tmp_path: Path, assessment_id: int, question_id: int
) -> dict[str, object]:
    image_path = tmp_path / "script.png"
    Image.new("RGB", (160, 120), color="white").save(image_path, format="PNG")
    headers = {"Authorization": f"Bearer {token}"}
    with image_path.open("rb") as file_obj:
        submission_response = client.post(
            f"/assessments/{assessment_id}/submissions/upload",
            data={"student_identifier": "S-001"},
            files={"file": ("script.png", file_obj, "image/png")},
            headers=headers,
        )
    assert submission_response.status_code == 201
    submission = submission_response.json()
    page = submission["pages"][0]
    region_response = client.post(
        f"/submission-pages/{page['id']}/answer-regions",
        json={"question_id": question_id, "x": 1, "y": 1, "width": 80, "height": 80},
        headers=headers,
    )
    assert region_response.status_code == 201
    region = region_response.json()
    confirmation_response = client.patch(
        f"/answer-regions/{region['id']}/full-answer-confirmation",
        json={
            "full_answer_confirmed": True,
            "continuation_not_needed": True,
            "manual_answer_text": "Synthetic confirmed answer text.",
            "packet_status": "complete",
        },
        headers=headers,
    )
    assert confirmation_response.status_code == 200
    return {"submission": submission, "page": page, "region": confirmation_response.json()}


def get_run(client: TestClient, token: str, run_id: int) -> dict[str, object]:
    response = client.get(f"/grading-runs/{run_id}", headers={"Authorization": f"Bearer {token}"})
    assert response.status_code == 200
    return response.json()


def test_derived_checklist_shows_material_blockers_before_upload(client: TestClient) -> None:
    teacher, token = register_teacher(client)
    assessment = create_assessment_for_teacher(client, int(teacher["id"]), token)
    run = client.post(
        f"/assessments/{assessment['id']}/grading-runs/custom",
        headers={"Authorization": f"Bearer {token}"},
    ).json()

    detail = get_run(client, token, int(run["id"]))

    workflow = detail["workflow_state"]
    assert workflow["materials_uploaded"] is False
    assert workflow["grading_ready"] is False
    assert workflow["question_count"] == 0
    assert "Upload question PDF." in workflow["blockers"]
    assert "Upload required materials." in workflow["next_actions"]


def test_derived_checklist_updates_after_material_upload_and_confirmation(
    client: TestClient,
) -> None:
    teacher, token = register_teacher(client)
    assessment = create_assessment_for_teacher(client, int(teacher["id"]), token)
    run = client.post(
        f"/assessments/{assessment['id']}/grading-runs/custom",
        headers={"Authorization": f"Bearer {token}"},
    ).json()
    run_id = int(run["id"])

    upload_all_materials(client, token, run_id)
    after_upload = get_run(client, token, run_id)["workflow_state"]
    assert after_upload["materials_uploaded"] is True
    assert after_upload["materials_confirmed"] is False
    assert "Confirm uploaded materials." in after_upload["blockers"]

    confirm_response = client.post(
        f"/grading-runs/{run_id}/confirm-materials",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert confirm_response.status_code == 200
    confirmed = confirm_response.json()
    assert confirmed["materials_confirmed_at"] is not None
    assert confirmed["workflow_state"]["materials_confirmed"] is True


def test_derived_checklist_requires_confirmed_questions_rubrics_scripts_and_regions(
    client: TestClient, tmp_path: Path
) -> None:
    teacher, token = register_teacher(client)
    assessment = create_assessment_for_teacher(client, int(teacher["id"]), token)
    run_id = int(
        client.post(
            f"/assessments/{assessment['id']}/grading-runs/custom",
            headers={"Authorization": f"Bearer {token}"},
        ).json()["id"]
    )
    upload_all_materials(client, token, run_id)
    assert (
        client.post(
            f"/grading-runs/{run_id}/confirm-materials",
            headers={"Authorization": f"Bearer {token}"},
        ).status_code
        == 200
    )
    question_data = create_question_and_active_rubric(client, int(assessment["id"]), token)

    before_confirm = get_run(client, token, run_id)["workflow_state"]
    assert before_confirm["question_count"] == 1
    assert before_confirm["rubric_count"] == 1
    assert before_confirm["mapped_question_count"] == 0
    assert before_confirm["unmapped_question_count"] == 1
    assert before_confirm["mapped_page_count"] == 0
    assert before_confirm["unmapped_page_count"] == 0
    assert before_confirm["mapped_submission_count"] == 0
    assert before_confirm["unmapped_submission_count"] == 0

    confirm = client.post(
        f"/grading-runs/{run_id}/confirm-questions-rubrics",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert confirm.status_code == 200
    after_confirm = confirm.json()["workflow_state"]
    assert after_confirm["questions_confirmed"] is True
    assert after_confirm["rubrics_confirmed"] is True
    assert "Upload at least one script." in after_confirm["blockers"]

    created = upload_script_and_create_region(
        client, token, tmp_path, int(assessment["id"]), int(question_data["question"]["id"])
    )
    ready = get_run(client, token, run_id)["workflow_state"]
    assert ready["submission_count"] == 1
    assert ready["submission_page_count"] == 1
    assert ready["answer_region_count"] == 1
    assert ready["mapped_question_count"] == 1
    assert ready["unmapped_question_count"] == 0
    assert ready["mapped_page_count"] == 1
    assert ready["unmapped_page_count"] == 0
    assert ready["mapped_submission_count"] == 1
    assert ready["unmapped_submission_count"] == 0
    assert ready["grading_ready"] is True
    assert ready["review_ready"] is False
    assert created["region"]["id"]



def test_confirm_questions_rubrics_blocks_unconfirmed_extracted_nodes_and_rubrics(
    client: TestClient,
) -> None:
    teacher, token = register_teacher(client, "extract-gate")
    assessment = create_assessment_for_teacher(client, int(teacher["id"]), token)
    run_id = int(
        client.post(
            f"/assessments/{assessment['id']}/grading-runs/custom",
            headers={"Authorization": f"Bearer {token}"},
        ).json()["id"]
    )
    upload_all_materials(client, token, run_id)
    assert client.post(
        f"/grading-runs/{run_id}/confirm-materials",
        headers={"Authorization": f"Bearer {token}"},
    ).status_code == 200
    create_question_and_active_rubric(client, int(assessment["id"]), token)

    extraction_run = ExtractionRun(
        assessment_id=int(assessment["id"]),
        artifact_file_path="uploads/extractions/question.pdf",
        original_filename="question.pdf",
        content_type="application/pdf",
        extraction_type="question_paper",
        provider="mock",
        status="succeeded",
        blockers=[],
    )
    db = SessionLocal()
    try:
        db.add(extraction_run)
        db.flush()
        db.add(
            QuestionNode(
                assessment_id=int(assessment["id"]),
                extraction_run_id=extraction_run.id,
                question_number="Q1",
                parent_question_number=None,
                label="Q1",
                text="Synthetic extracted question",
                marks="10.00",
                node_type="question",
                source_page=1,
                source_reference={"page": 1},
                confidence="0.90",
                teacher_confirmed=False,
            )
        )
        db.add(
            RubricExtractionCriterion(
                assessment_id=int(assessment["id"]),
                extraction_run_id=extraction_run.id,
                question_number="Q1",
                criterion_label="Criterion",
                description="Synthetic extracted rubric",
                max_marks="10.00",
                confidence="0.80",
                blocker=None,
                teacher_confirmed=False,
            )
        )
        db.commit()
    finally:
        db.close()

    blocked = client.post(
        f"/grading-runs/{run_id}/confirm-questions-rubrics",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert blocked.status_code == 409
    assert "Confirm every extracted question node" in blocked.json()["detail"]

    auth_headers = {"Authorization": f"Bearer {token}"}
    node = client.get(
        f"/assessments/{assessment['id']}/question-nodes", headers=auth_headers
    ).json()[0]
    criterion = client.get(
        f"/assessments/{assessment['id']}/rubric-extraction-criteria", headers=auth_headers
    ).json()[0]
    assert client.patch(
        f"/question-nodes/{node['id']}",
        headers=auth_headers,
        json={"teacher_confirmed": True},
    ).status_code == 200
    assert client.patch(
        f"/rubric-extraction-criteria/{criterion['id']}",
        headers=auth_headers,
        json={"teacher_confirmed": True, "blocker": None},
    ).status_code == 200

    confirmed = client.post(
        f"/grading-runs/{run_id}/confirm-questions-rubrics",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert confirmed.status_code == 200
    workflow = confirmed.json()["workflow_state"]
    assert workflow["extracted_questions_confirmed"] is True
    assert workflow["extracted_rubrics_confirmed"] is True
    assert workflow["question_extraction_run_id"] is not None



def test_semi_automated_run_rejected_by_default(client: TestClient) -> None:
    teacher, token = register_teacher(client, "semi")
    assessment = create_assessment_for_teacher(client, int(teacher["id"]), token)

    response = client.post(
        f"/assessments/{assessment['id']}/grading-runs/custom",
        headers={"Authorization": f"Bearer {token}"},
        json={"mode": "semi_automated", "notes": "Semi-automated mode", "marking_policy": "easy"},
    )

    assert response.status_code == 400
    assert response.json()["detail"] == (
        "Mode 'semi_automated' is not available for teacher workflow yet. Use Custom Controlled."
    )


def test_semi_automated_run_can_be_enabled_for_explicit_experimentation(
    client: TestClient, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setenv("SEMI_AUTOMATED_MODE_ENABLED", "true")
    get_settings.cache_clear()
    teacher, token = register_teacher(client, "semi-enabled")
    assessment = create_assessment_for_teacher(client, int(teacher["id"]), token)

    response = client.post(
        f"/assessments/{assessment['id']}/grading-runs/custom",
        headers={"Authorization": f"Bearer {token}"},
        json={"mode": "semi_automated", "notes": "Semi-automated mode", "marking_policy": "easy"},
    )

    assert response.status_code == 201
    run = response.json()
    assert run["mode"] == "semi_automated"
    assert run["marking_policy"] == "easy"


def test_fully_automated_run_is_rejected_with_a_clear_message(client: TestClient) -> None:
    teacher, token = register_teacher(client, "fully")
    assessment = create_assessment_for_teacher(client, int(teacher["id"]), token)

    response = client.post(
        f"/assessments/{assessment['id']}/grading-runs/custom",
        headers={"Authorization": f"Bearer {token}"},
        json={"mode": "fully_automated", "notes": "Fully automated mode"},
    )

    assert response.status_code == 400
    assert response.json()["detail"] == (
        "Mode 'fully_automated' is not available for teacher workflow yet. Use Custom Controlled."
    )


def test_status_update_does_not_imply_finalization(client: TestClient) -> None:
    teacher, token = register_teacher(client)
    assessment = create_assessment_for_teacher(client, int(teacher["id"]), token)
    run = client.post(
        f"/assessments/{assessment['id']}/grading-runs/custom",
        headers={"Authorization": f"Bearer {token}"},
    ).json()

    response = client.patch(
        f"/grading-runs/{run['id']}",
        headers={"Authorization": f"Bearer {token}"},
        json={"status": "completed", "notes": "Manual note only."},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "completed"
    assert body["workflow_state"]["export_ready"] is False
    assert body["workflow_state"]["grading_ready"] is False
    assert body["workflow_state"]["derived_status"] != "completed"
    assert body["workflow_state"]["blockers"]


def test_custom_controlled_full_v0_api_workflow_and_no_auto_finalization(
    client: TestClient, tmp_path: Path
) -> None:
    teacher, token = register_teacher(client, "v0")
    assessment = create_assessment_for_teacher(client, int(teacher["id"]), token)
    run = client.post(
        f"/assessments/{assessment['id']}/grading-runs/custom",
        headers={"Authorization": f"Bearer {token}"},
        json={"notes": "V0 smoke"},
    ).json()
    run_id = int(run["id"])

    upload_all_materials(client, token, run_id)
    assert (
        client.post(
            f"/grading-runs/{run_id}/confirm-materials",
            headers={"Authorization": f"Bearer {token}"},
        ).status_code
        == 200
    )
    question_data = create_question_and_active_rubric(client, int(assessment["id"]), token)
    assert (
        client.post(
            f"/grading-runs/{run_id}/confirm-questions-rubrics",
            headers={"Authorization": f"Bearer {token}"},
        ).status_code
        == 200
    )
    upload_script_and_create_region(
        client, token, tmp_path, int(assessment["id"]), int(question_data["question"]["id"])
    )

    grade_response = client.post(
        f"/grading-runs/{run_id}/grade-all-mock",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert grade_response.status_code == 200
    assert grade_response.json()["graded_count"] == 1

    after_grading = get_run(client, token, run_id)["workflow_state"]
    assert after_grading["suggestions_created"] is True
    assert after_grading["review_ready"] is True
    assert after_grading["final_grade_count"] == 0
    assert after_grading["export_ready"] is False

    review_queue = client.get(
        f"/assessments/{assessment['id']}/review-queue",
        headers={"Authorization": f"Bearer {token}"},
    ).json()
    suggestion_id = review_queue[0]["latest_grade_suggestion"]["id"]
    approve_response = client.post(
        f"/grade-suggestions/{suggestion_id}/approve",
        headers={"Authorization": f"Bearer {token}"},
        json={"teacher_comment": "Approved in V0 smoke."},
    )
    assert approve_response.status_code == 201

    final_state = get_run(client, token, run_id)["workflow_state"]
    assert final_state["final_grades_created"] is True
    assert final_state["export_ready"] is True
    assert final_state["derived_status"] == "completed"

    export_response = client.get(
        f"/assessments/{assessment['id']}/export/final-grades.xlsx",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert export_response.status_code == 200
    assert export_response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_custom_controlled_workflow_state_counts_zip_imported_submissions(
    client: TestClient, tmp_path: Path
) -> None:
    import zipfile

    teacher, token = register_teacher(client, "zip")
    assessment = create_assessment_for_teacher(client, int(teacher["id"]), token)
    run = client.post(
        f"/assessments/{assessment['id']}/grading-runs/custom",
        headers={"Authorization": f"Bearer {token}"},
    ).json()
    image_path = tmp_path / "script.png"
    Image.new("RGB", (160, 120), color="white").save(image_path, format="PNG")
    zip_path = tmp_path / "scripts.zip"
    with zipfile.ZipFile(zip_path, "w") as archive:
        archive.writestr("student_001.png", image_path.read_bytes())
        archive.writestr("student_002.png", image_path.read_bytes())

    with zip_path.open("rb") as file_obj:
        upload = client.post(
            f"/assessments/{assessment['id']}/submissions/upload-zip",
            files={"file": ("scripts.zip", file_obj, "application/zip")},
            headers={"Authorization": f"Bearer {token}"},
        )
    assert upload.status_code == 201
    assert upload.json()["imported_count"] == 2

    detail = get_run(client, token, int(run["id"]))
    workflow = detail["workflow_state"]
    assert workflow["scripts_uploaded"] is True
    assert workflow["submission_count"] == 2
    assert workflow["submission_page_count"] == 2
    assert workflow["answer_region_count"] == 0
    assert workflow["mapped_question_count"] == 0
    assert workflow["unmapped_question_count"] == 0
    assert workflow["mapped_page_count"] == 0
    assert workflow["unmapped_page_count"] == 2
    assert workflow["mapped_submission_count"] == 0
    assert workflow["unmapped_submission_count"] == 2



def test_marking_policy_defaults_validates_and_updates_on_custom_run(client: TestClient) -> None:
    teacher, token = register_teacher(client, "policy")
    assessment = create_assessment_for_teacher(client, int(teacher["id"]), token)

    default_response = client.post(
        f"/assessments/{assessment['id']}/grading-runs/custom",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert default_response.status_code == 201
    assert default_response.json()["marking_policy"] == "general"

    tough_response = client.post(
        f"/assessments/{assessment['id']}/grading-runs/custom",
        headers={"Authorization": f"Bearer {token}"},
        json={"marking_policy": "tough", "notes": "Strict marking"},
    )
    assert tough_response.status_code == 201
    run_id = tough_response.json()["id"]
    assert tough_response.json()["marking_policy"] == "tough"

    update_response = client.patch(
        f"/grading-runs/{run_id}",
        headers={"Authorization": f"Bearer {token}"},
        json={"marking_policy": "easy"},
    )
    assert update_response.status_code == 200
    assert update_response.json()["marking_policy"] == "easy"

    invalid = client.patch(
        f"/grading-runs/{run_id}",
        headers={"Authorization": f"Bearer {token}"},
        json={"marking_policy": "harsh"},
    )
    assert invalid.status_code == 422


def test_custom_controlled_mock_grading_persists_marking_policy_and_exports_it(
    client: TestClient, tmp_path: Path
) -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    teacher, token = register_teacher(client, "policy-grade")
    assessment = create_assessment_for_teacher(client, int(teacher["id"]), token)
    run = client.post(
        f"/assessments/{assessment['id']}/grading-runs/custom",
        headers={"Authorization": f"Bearer {token}"},
        json={"marking_policy": "tough"},
    ).json()
    run_id = int(run["id"])
    upload_all_materials(client, token, run_id)
    assert client.post(
        f"/grading-runs/{run_id}/confirm-materials",
        headers={"Authorization": f"Bearer {token}"},
    ).status_code == 200
    question_data = create_question_and_active_rubric(client, int(assessment["id"]), token)
    assert client.post(
        f"/grading-runs/{run_id}/confirm-questions-rubrics",
        headers={"Authorization": f"Bearer {token}"},
    ).status_code == 200
    upload_script_and_create_region(
        client, token, tmp_path, int(assessment["id"]), int(question_data["question"]["id"])
    )

    grade_response = client.post(
        f"/grading-runs/{run_id}/grade-all-mock",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert grade_response.status_code == 200
    assert grade_response.json()["marking_policy"] == "tough"
    suggestion_id = grade_response.json()["created_grade_suggestion_ids"][0]

    suggestion = client.get(
        f"/grade-suggestions/{suggestion_id}", headers={"Authorization": f"Bearer {token}"}
    )
    assert suggestion.status_code == 200
    suggestion_body = suggestion.json()
    assert suggestion_body["marking_policy"] == "tough"
    assert suggestion_body["raw_response_json"]["marking_policy"] == "tough"
    assert "marking_policy:tough" in suggestion_body["raw_response_json"]["review_flags"]

    review_queue = client.get(
        f"/assessments/{assessment['id']}/review-queue",
        headers={"Authorization": f"Bearer {token}"},
    ).json()
    assert review_queue[0]["latest_grade_suggestion"]["marking_policy"] == "tough"
    approve_response = client.post(
        f"/grade-suggestions/{suggestion_id}/approve",
        headers={"Authorization": f"Bearer {token}"},
        json={"teacher_comment": "Approved with tough policy."},
    )
    assert approve_response.status_code == 201

    export_response = client.get(
        f"/assessments/{assessment['id']}/export/final-grades.xlsx",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert export_response.status_code == 200
    workbook = load_workbook(BytesIO(export_response.content))
    rows = list(workbook.active.iter_rows(values_only=True))
    headers = list(rows[0])
    assert "marking_policy" in headers
    policy_index = headers.index("marking_policy")
    assert rows[1][policy_index] == "tough"
