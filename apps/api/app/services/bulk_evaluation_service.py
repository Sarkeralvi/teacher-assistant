from __future__ import annotations

import csv
import hashlib
import json
import re
import shutil
import zipfile
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path, PurePosixPath
from uuid import uuid4

from fastapi import UploadFile
from openpyxl import Workbook
from PIL import Image, UnidentifiedImageError
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.core.config import Settings, get_settings
from app.models import (
    AnswerRegion,
    AnswerRegionMapping,
    AnswerRegionOcrRun,
    AuditLog,
    BulkEvaluationItem,
    BulkEvaluationRun,
    FinalGrade,
    GradeSuggestion,
    GradingRun,
    Question,
    Rubric,
    Submission,
    SubmissionPage,
    User,
)
from app.services.final_grade_service import FinalGradeService
from app.services.grading_service import GradingService, rubric_snapshot_hash
from app.services.local_ai_phase_manager import LocalAiPhaseManager
from app.services.local_model_lease_service import LocalModelLeaseService
from app.services.local_script_preparation import (
    LocalScriptPreparationError,
    LocalScriptPreparationService,
)
from app.services.qwen38_visual_transcription_service import (
    Qwen38VisualTranscriptionService,
    VisualTranscriptionError,
)
from app.services.storage import LocalStorage
from app.services.submission_processing import extract_page_images
from packages.brain.adapter import (
    BrainAdapter,
    BrainProviderConfigurationError,
    sanitize_provider_error,
)

BULK_POLICY_VERSION = "bulk-supervised-qwen38-v1"
_UNCERTAIN_PROVIDER_WARNING = (
    "Provider work stopped without a safely persisted result; "
    "explicit item retry required"
)
_SUPPORTED = {".pdf", ".png", ".jpg", ".jpeg"}
_KNOWN_METADATA = {".ds_store", "thumbs.db"}
_MAX_MEMBER_BYTES = 100 * 1024 * 1024
_MAX_UNCOMPRESSED_BYTES = 4 * 1024 * 1024 * 1024
_MAX_COMPRESSION_RATIO = 100
_ACTIVE_RUN_STATUSES = {
    "preflighting",
    "queued",
    "mapping",
    "transcribing",
    "grading",
    "review_ready",
    "stopping",
    "paused",
}


class BulkEvaluationError(RuntimeError):
    pass


@dataclass(frozen=True)
class ArchiveUnit:
    source: str
    identifier: str
    student_name: str | None
    members: tuple[zipfile.ZipInfo, ...]
    kind: str


def _natural_key(value: str) -> list[object]:
    return [int(part) if part.isdigit() else part.casefold() for part in re.split(r"(\d+)", value)]


def _safe_member(name: str) -> PurePosixPath:
    normalized = name.replace("\\", "/")
    path = PurePosixPath(normalized)
    if (
        path.is_absolute()
        or not path.parts
        or any(part in {"", ".", ".."} or ":" in part for part in path.parts)
    ):
        raise BulkEvaluationError("ZIP contains an unsafe member path")
    return path


def _ignored_member(path: PurePosixPath) -> bool:
    return "__MACOSX" in path.parts or path.name.casefold() in _KNOWN_METADATA


def _canonical_hash(payload: object) -> str:
    encoded = json.dumps(payload, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


class BulkEvaluationService:
    def __init__(
        self,
        db: Session,
        *,
        settings: Settings | None = None,
        storage: LocalStorage | None = None,
    ) -> None:
        self.db = db
        self.settings = settings or get_settings()
        self.storage = storage or LocalStorage()

    def create_from_zip(
        self,
        *,
        assessment_id: int,
        grading_run: GradingRun,
        teacher: User,
        upload: UploadFile,
        expected_model: str,
        marking_policy: str,
        maximum_provider_calls: int,
    ) -> BulkEvaluationRun:
        self._assert_enabled(expected_model, maximum_provider_calls)
        questions, reference_hash = self._load_references(assessment_id, grading_run)
        staged_path, archive_hash = self._stage_archive(upload)
        created_submission_ids: list[int] = []
        try:
            with zipfile.ZipFile(staged_path) as archive:
                units, manifest_hash, manifest_payload = self._inspect_archive(archive)
                existing_run = self.db.scalar(
                    select(BulkEvaluationRun).where(
                        BulkEvaluationRun.assessment_id == assessment_id,
                        BulkEvaluationRun.archive_sha256 == archive_hash,
                        BulkEvaluationRun.reference_bundle_sha256 == reference_hash,
                        BulkEvaluationRun.status.in_(_ACTIVE_RUN_STATUSES),
                    )
                )
                if existing_run is not None:
                    return self.get_run(existing_run.id)

                self._reject_existing_identifiers(assessment_id, units)
                run = BulkEvaluationRun(
                    assessment_id=assessment_id,
                    grading_run_id=grading_run.id,
                    created_by_teacher_id=teacher.id,
                    provider="llama_cpp_qwen38",
                    model_name=expected_model,
                    marking_policy=marking_policy,
                    policy_version=BULK_POLICY_VERSION,
                    reference_bundle_sha256=reference_hash,
                    archive_sha256=archive_hash,
                    manifest_sha256=manifest_hash,
                    import_manifest={
                        "units": manifest_payload,
                        "submission_ids": [],
                    },
                    status="preflighting",
                    stage="ingestion",
                    authorized_call_limit=maximum_provider_calls,
                    calls_used=0,
                    total_submissions=len(units),
                    started_at=datetime.now(UTC),
                    heartbeat_at=datetime.now(UTC),
                )
                self.db.add(run)
                self.db.flush()

                total_pages = 0
                submissions: list[Submission] = []
                for unit in units:
                    submission = self._import_unit(
                        archive=archive,
                        assessment_id=assessment_id,
                        unit=unit,
                    )
                    created_submission_ids.append(submission.id)
                    submissions.append(submission)
                    total_pages += len(submission.pages)
                    if total_pages > self.settings.bulk_max_pages:
                        raise BulkEvaluationError(
                            f"ZIP exceeds the {self.settings.bulk_max_pages}-page bulk limit"
                        )

                for submission in submissions:
                    for question in questions:
                        self.db.add(
                            BulkEvaluationItem(
                                run_id=run.id,
                                submission_id=submission.id,
                                question_id=question.id,
                                status="pending",
                                stage="mapping",
                                exception_codes=[],
                                warnings=[],
                            )
                        )

                run.total_pages = total_pages
                run.total_items = len(submissions) * len(questions)
                run.status = "queued"
                run.stage = "mapping"
                run.import_manifest = {
                    "units": manifest_payload,
                    "submission_ids": [submission.id for submission in submissions],
                }
                self.db.add(
                    AuditLog(
                        actor_type="teacher",
                        actor_id=teacher.id,
                        event_type="bulk_evaluation_requested",
                        entity_type="bulk_evaluation_run",
                        entity_id=run.id,
                        payload_json={
                            "assessment_id": assessment_id,
                            "grading_run_id": grading_run.id,
                            "provider": run.provider,
                            "model": run.model_name,
                            "policy_version": run.policy_version,
                            "archive_sha256": archive_hash,
                            "reference_bundle_sha256": reference_hash,
                            "submission_count": len(submissions),
                            "page_count": total_pages,
                            "item_count": run.total_items,
                            "authorized_call_limit": maximum_provider_calls,
                            "local_only": True,
                            "draft_only": True,
                        },
                    )
                )
                self.db.commit()
                return self.get_run(run.id)
        except BulkEvaluationError:
            self.db.rollback()
            self._cleanup_submissions(created_submission_ids)
            raise
        except zipfile.BadZipFile as exc:
            self.db.rollback()
            self._cleanup_submissions(created_submission_ids)
            raise BulkEvaluationError("Uploaded file is not a valid ZIP archive") from exc
        except Exception:
            self.db.rollback()
            self._cleanup_submissions(created_submission_ids)
            raise
        finally:
            staged_path.unlink(missing_ok=True)

    def get_run(self, run_id: int) -> BulkEvaluationRun:
        run = self.db.scalar(
            select(BulkEvaluationRun)
            .options(selectinload(BulkEvaluationRun.items))
            .where(BulkEvaluationRun.id == run_id)
        )
        if run is None:
            raise BulkEvaluationError("Bulk evaluation run not found")
        return run

    def reference_bundle_hash(self, assessment_id: int, grading_run: GradingRun) -> str:
        _, value = self._load_references(assessment_id, grading_run)
        return value

    def _assert_enabled(self, expected_model: str, maximum_calls: int) -> None:
        settings = self.settings
        if not settings.bulk_supervised_enabled:
            raise BulkEvaluationError("Bulk supervised evaluation is disabled")
        if not settings.brain_allow_real_providers:
            raise BulkEvaluationError("Real local providers are disabled")
        if not (
            settings.local_qwen38_enabled
            and settings.local_qwen38_visual_preparation_enabled
            and settings.local_qwen38_transcription_enabled
            and settings.local_qwen38_grading_enabled
        ):
            raise BulkEvaluationError("Required local Qwen3.8 capabilities are disabled")
        if expected_model != settings.local_qwen38_model:
            raise BulkEvaluationError("Expected local Qwen3.8 model alias does not match")
        if maximum_calls > settings.bulk_max_provider_calls:
            raise BulkEvaluationError("Requested provider calls exceed the bulk server ceiling")
        holder_id = f"bulk_preflight:{uuid4().hex}"
        lease = LocalModelLeaseService(self.db)
        try:
            with lease.hold(
                model_phase="Qwen38",
                holder_kind="bulk_evaluation_preflight",
                holder_id=holder_id,
            ):
                if settings.local_ai_phase_switch_enabled:
                    LocalAiPhaseManager(settings=settings, db=self.db).switch(
                        "Qwen38", lease_holder_id=holder_id
                    )
                BrainAdapter.for_provider(
                    settings, "llama_cpp_qwen38"
                ).verify_available_model()
                lease.heartbeat(holder_id=holder_id)
        except (BrainProviderConfigurationError, RuntimeError) as exc:
            raise BulkEvaluationError("Local Qwen3.8 is unavailable or mismatched") from exc

    def _load_references(
        self, assessment_id: int, grading_run: GradingRun
    ) -> tuple[list[Question], str]:
        if grading_run.assessment_id != assessment_id:
            raise BulkEvaluationError("Grading run does not belong to this assessment")
        if grading_run.questions_confirmed_at is None or grading_run.rubrics_confirmed_at is None:
            raise BulkEvaluationError("Finalize the questions, solutions, and rubrics first")
        questions = list(
            self.db.scalars(
                select(Question)
                .options(selectinload(Question.rubrics))
                .where(Question.assessment_id == assessment_id)
                .order_by(Question.id)
            ).all()
        )
        if not questions:
            raise BulkEvaluationError("Assessment has no finalized questions")
        snapshot: list[dict[str, object]] = []
        for question in questions:
            rubric = next((item for item in question.rubrics if item.is_active), None)
            if not (question.model_answer or "").strip() or rubric is None:
                raise BulkEvaluationError(
                    f"Question {question.question_no} needs a model answer and active rubric"
                )
            snapshot.append(
                {
                    "id": question.id,
                    "label": question.question_no,
                    "text": question.question_text,
                    "model_answer": question.model_answer,
                    "marks": str(question.total_marks),
                    "rubric_id": rubric.id,
                    "rubric_version": rubric.version,
                    "rubric": rubric.rubric_json,
                }
            )
        return questions, _canonical_hash(snapshot)

    def _stage_archive(self, upload: UploadFile) -> tuple[Path, str]:
        filename = upload.filename or ""
        if PurePosixPath(filename).suffix.casefold() != ".zip":
            raise BulkEvaluationError("Upload a ZIP archive")
        target_dir = self.storage.uploads_dir / "bulk_staging"
        target_dir.mkdir(parents=True, exist_ok=True)
        target = target_dir / f"bulk-{uuid4().hex}.zip"
        digest = hashlib.sha256()
        total = 0
        with target.open("wb") as output:
            while chunk := upload.file.read(1024 * 1024):
                total += len(chunk)
                if total > self.settings.bulk_max_archive_bytes:
                    output.close()
                    target.unlink(missing_ok=True)
                    raise BulkEvaluationError("ZIP exceeds the configured bulk archive limit")
                digest.update(chunk)
                output.write(chunk)
        if total == 0:
            target.unlink(missing_ok=True)
            raise BulkEvaluationError("Uploaded ZIP is empty")
        return target, digest.hexdigest()

    def _inspect_archive(
        self, archive: zipfile.ZipFile
    ) -> tuple[list[ArchiveUnit], str | None, list[dict[str, object]]]:
        infos = [item for item in archive.infolist() if not item.is_dir()]
        if len(infos) > 1000:
            raise BulkEvaluationError("ZIP contains more than 1,000 entries")
        usable: list[tuple[PurePosixPath, zipfile.ZipInfo]] = []
        total_uncompressed = 0
        manifest_info: zipfile.ZipInfo | None = None
        for info in infos:
            path = _safe_member(info.filename)
            if _ignored_member(path):
                continue
            if info.flag_bits & 0x1:
                raise BulkEvaluationError("Encrypted ZIP entries are not supported")
            if info.file_size > _MAX_MEMBER_BYTES:
                raise BulkEvaluationError("A ZIP member exceeds the 100 MiB limit")
            total_uncompressed += info.file_size
            if total_uncompressed > _MAX_UNCOMPRESSED_BYTES:
                raise BulkEvaluationError("ZIP expands beyond the safe uncompressed limit")
            ratio = info.file_size / max(info.compress_size, 1)
            if ratio > _MAX_COMPRESSION_RATIO:
                raise BulkEvaluationError("ZIP compression ratio is unsafe")
            if len(path.parts) == 1 and path.name.casefold() == "manifest.csv":
                manifest_info = info
                continue
            if path.suffix.casefold() not in _SUPPORTED:
                raise BulkEvaluationError(f"Unsupported ZIP entry: {path.name}")
            usable.append((path, info))
        if not usable:
            raise BulkEvaluationError("ZIP contains no supported scripts")

        manifest, manifest_hash = self._read_manifest(archive, manifest_info)
        root_pdfs = [(path, info) for path, info in usable if len(path.parts) == 1]
        root_images = [path for path, _ in root_pdfs if path.suffix.casefold() != ".pdf"]
        if root_images:
            raise BulkEvaluationError(
                "Root-level images are ambiguous; place each script in a folder"
            )
        grouped: dict[str, list[tuple[PurePosixPath, zipfile.ZipInfo]]] = {}
        for path, info in usable:
            if len(path.parts) == 1:
                continue
            if len(path.parts) != 2:
                raise BulkEvaluationError("Student folders may contain files but no nested folders")
            grouped.setdefault(path.parts[0], []).append((path, info))

        sources: dict[str, tuple[str, tuple[zipfile.ZipInfo, ...]]] = {}
        for path, info in root_pdfs:
            sources[path.as_posix()] = ("pdf", (info,))
        for folder, members in grouped.items():
            pdfs = [info for path, info in members if path.suffix.casefold() == ".pdf"]
            images = [info for path, info in members if path.suffix.casefold() != ".pdf"]
            if pdfs and images:
                raise BulkEvaluationError(f"Student folder {folder} mixes PDF and image pages")
            if len(pdfs) > 1:
                # One folder holding several PDFs is ambiguous and must not be
                # guessed: it is either one student whose script was split
                # across files, or -- far more commonly -- a container folder
                # holding one PDF per student. Guessing wrong either merges two
                # students into one submission or splits one student in two, so
                # this stays an error. It says how to resolve it, because the
                # bare rule sent a teacher back to the ZIP with no idea which
                # of the two valid layouts they were supposed to produce.
                raise BulkEvaluationError(
                    f"Student folder {folder} contains {len(pdfs)} PDFs, so it is "
                    "unclear whether this is one student or several. If each PDF "
                    f"is a different student, move them to the top level of the ZIP "
                    f"(delete the {folder} folder) so each PDF becomes one student. "
                    "If they are all one student's script, export that student as a "
                    "single PDF, or put their pages in the folder as numbered images."
                )
            ordered = tuple(
                info
                for _, info in sorted(
                    members, key=lambda item: _natural_key(item[0].name)
                )
            )
            sources[folder] = ("pdf" if pdfs else "images", ordered)

        if len(sources) > self.settings.bulk_max_submissions:
            raise BulkEvaluationError(
                f"ZIP exceeds the {self.settings.bulk_max_submissions}-submission limit"
            )
        unknown_manifest_sources = set(manifest) - set(sources)
        if unknown_manifest_sources:
            raise BulkEvaluationError("manifest.csv refers to a source that is not in the ZIP")

        units: list[ArchiveUnit] = []
        payload: list[dict[str, object]] = []
        seen_identifiers: set[str] = set()
        for source in sorted(sources, key=_natural_key):
            kind, members = sources[source]
            metadata = manifest.get(source, {})
            default_name = PurePosixPath(source).stem if kind == "pdf" else source
            identifier = str(metadata.get("student_identifier") or default_name).strip()[:128]
            if not identifier:
                raise BulkEvaluationError("Every script needs a student identifier")
            folded = identifier.casefold()
            if folded in seen_identifiers:
                raise BulkEvaluationError("ZIP contains duplicate student identifiers")
            seen_identifiers.add(folded)
            student_name = str(metadata.get("student_name") or "").strip()[:255] or None
            units.append(ArchiveUnit(source, identifier, student_name, members, kind))
            payload.append(
                {
                    "source": source,
                    "student_identifier": identifier,
                    "student_name": student_name,
                    "file_count": len(members),
                    "kind": kind,
                }
            )
        return units, manifest_hash, payload

    def _read_manifest(
        self, archive: zipfile.ZipFile, info: zipfile.ZipInfo | None
    ) -> tuple[dict[str, dict[str, str]], str | None]:
        if info is None:
            return {}, None
        raw = archive.read(info)
        digest = hashlib.sha256(raw).hexdigest()
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise BulkEvaluationError("manifest.csv must be UTF-8") from exc
        reader = csv.DictReader(text.splitlines())
        required = {"source", "student_identifier", "student_name"}
        if not reader.fieldnames or not required.issubset(set(reader.fieldnames)):
            raise BulkEvaluationError(
                "manifest.csv needs source, student_identifier, and student_name columns"
            )
        result: dict[str, dict[str, str]] = {}
        for row in reader:
            source = (row.get("source") or "").replace("\\", "/").strip().rstrip("/")
            if not source or source in result:
                raise BulkEvaluationError("manifest.csv contains a blank or duplicate source")
            result[source] = {
                "student_identifier": (row.get("student_identifier") or "").strip(),
                "student_name": (row.get("student_name") or "").strip(),
            }
        return result, digest

    def _reject_existing_identifiers(self, assessment_id: int, units: list[ArchiveUnit]) -> None:
        incoming = {unit.identifier.casefold() for unit in units}
        existing = {
            value.casefold()
            for value in self.db.scalars(
                select(Submission.student_identifier).where(
                    Submission.assessment_id == assessment_id
                )
            ).all()
        }
        if incoming & existing:
            raise BulkEvaluationError(
                "A student identifier already exists in this assessment; "
                "existing scripts were preserved"
            )

    def _import_unit(
        self,
        *,
        archive: zipfile.ZipFile,
        assessment_id: int,
        unit: ArchiveUnit,
    ) -> Submission:
        submission = Submission(
            assessment_id=assessment_id,
            student_identifier=unit.identifier,
            student_name=unit.student_name,
            status="uploaded",
        )
        self.db.add(submission)
        self.db.flush()
        if unit.kind == "pdf":
            info = unit.members[0]
            target_dir = self.storage.uploads_dir / "submissions" / str(submission.id)
            target_dir.mkdir(parents=True, exist_ok=True)
            target = target_dir / f"original-{uuid4().hex}.pdf"
            with archive.open(info) as source, target.open("wb") as output:
                shutil.copyfileobj(source, output)
            page_paths = extract_page_images(
                storage=self.storage,
                submission_id=submission.id,
                uploaded_path=target,
                kind="pdf",
            )
        else:
            page_paths = []
            for page_no, info in enumerate(unit.members, start=1):
                output = self.storage.page_image_path(submission.id, page_no)
                try:
                    with archive.open(info) as source, Image.open(BytesIO(source.read())) as image:
                        image.verify()
                    with archive.open(info) as source, Image.open(BytesIO(source.read())) as image:
                        image.convert("RGB").save(output.absolute_path, format="PNG")
                except (UnidentifiedImageError, OSError) as exc:
                    raise BulkEvaluationError(
                        f"Image page in {unit.source} could not be decoded"
                    ) from exc
                page_paths.append(output.relative_path)
        if not page_paths:
            raise BulkEvaluationError(f"Script {unit.identifier} contains no pages")
        for page_no, image_path in enumerate(page_paths, start=1):
            page = SubmissionPage(
                submission_id=submission.id,
                page_no=page_no,
                image_path=image_path,
                quality_score=None,
            )
            submission.pages.append(page)
            self.db.add(page)
        self.db.flush()
        return submission

    def run_next(self, run_id: int) -> bool:
        """Advance one durable unit of work and report whether another is pending."""

        run = self.get_run(run_id)
        if run.status in {"completed", "completed_with_exceptions", "failed", "stopped"}:
            return False
        if run.stop_requested:
            run.status = "stopped"
            run.completed_at = datetime.now(UTC)
            self.db.commit()
            return False
        if run.calls_used >= run.authorized_call_limit:
            self._pause(run, "Authorized provider-call limit reached")
            return False
        run.heartbeat_at = datetime.now(UTC)
        try:
            if self._next_mapping_submission(run) is not None:
                self._process_mapping(run)
            elif self._next_transcription_item(run) is not None:
                self._process_transcription(run)
            elif self._next_grading_item(run) is not None:
                self._process_grading(run)
            else:
                self._finish_processing(run)
                return False
        except (LocalScriptPreparationError, VisualTranscriptionError) as exc:
            self.db.rollback()
            current = self.db.get(BulkEvaluationRun, run_id)
            if current is not None:
                self._mark_running_uncertain(
                    current,
                    _UNCERTAIN_PROVIDER_WARNING,
                )
                self._pause(current, sanitize_provider_error(str(exc)))
            return False
        except Exception as exc:
            self.db.rollback()
            current = self.db.get(BulkEvaluationRun, run_id)
            if current is not None:
                self._mark_running_uncertain(
                    current,
                    _UNCERTAIN_PROVIDER_WARNING,
                )
                self._pause(current, sanitize_provider_error(str(exc)))
            return False
        return self.get_run(run_id).status not in {
            "completed",
            "completed_with_exceptions",
            "failed",
            "stopped",
            "paused",
            "review_ready",
        }

    def stop(self, run: BulkEvaluationRun, *, teacher_id: int) -> BulkEvaluationRun:
        if run.status in {"completed", "completed_with_exceptions", "failed", "stopped"}:
            return run
        run.stop_requested = True
        run.status = "stopping"
        self._audit_run(run, "bulk_evaluation_stop_requested", teacher_id)
        self.db.commit()
        return self.get_run(run.id)

    def resume(self, run: BulkEvaluationRun, *, teacher_id: int) -> BulkEvaluationRun:
        active = {"preflighting", "queued", "mapping", "transcribing", "grading", "stopping"}
        stale_before = datetime.now(UTC) - timedelta(
            seconds=self.settings.cohort_dispatch_heartbeat_timeout_seconds
        )
        if run.status in active and run.heartbeat_at and run.heartbeat_at < stale_before:
            running_items = list(
                self.db.scalars(
                    select(BulkEvaluationItem).where(
                        BulkEvaluationItem.run_id == run.id,
                        BulkEvaluationItem.status == "running",
                    )
                ).all()
            )
            for item in running_items:
                item.status = "uncertain"
                item.exception_codes = list(
                    dict.fromkeys([*item.exception_codes, "provider_contract_failure"])
                )
                item.warnings = [
                    *item.warnings,
                    "Worker heartbeat expired during a provider call; explicit item retry required",
                ]
                item.completed_at = datetime.now(UTC)
            run.status = "paused"
            run.error = "Interrupted provider work was quarantined as uncertain"
            self.db.commit()
        if run.status not in {"paused", "stopped"}:
            raise BulkEvaluationError("Only a paused or stopped bulk run can resume")
        if run.calls_used >= run.authorized_call_limit:
            raise BulkEvaluationError(
                "Increase authorization in a new run; call limit is exhausted"
            )
        run.stop_requested = False
        run.error = None
        run.status = self._status_for_pending_stage(run)
        run.completed_at = None
        run.heartbeat_at = datetime.now(UTC)
        self._audit_run(run, "bulk_evaluation_resumed", teacher_id)
        self.db.commit()
        return self.get_run(run.id)

    def resume_item(
        self, run: BulkEvaluationRun, item: BulkEvaluationItem, *, teacher_id: int
    ) -> BulkEvaluationItem:
        if item.run_id != run.id or item.status not in {"exception", "uncertain"}:
            raise BulkEvaluationError("Only an exception or uncertain item can be resumed")
        mapping = self.db.get(AnswerRegionMapping, item.mapping_id) if item.mapping_id else None
        region = self.db.get(AnswerRegion, item.answer_region_id) if item.answer_region_id else None
        if mapping is None or not (mapping.teacher_confirmed or mapping.bulk_policy_verified):
            item.stage = "mapping"
        elif region is None or not (region.manual_answer_text or "").strip():
            item.stage = "transcription"
        else:
            item.stage = "grading"
        item.status = "pending"
        item.exception_codes = []
        item.warnings = []
        item.completed_at = None
        self._audit_run(
            run,
            "bulk_evaluation_item_resumed",
            teacher_id,
            {"item_id": item.id, "stage": item.stage},
        )
        if run.status in {"review_ready", "completed_with_exceptions"}:
            run.status = self._status_for_pending_stage(run)
            run.stage = item.stage
            run.completed_at = None
        self._refresh_counts(run)
        self.db.commit()
        self.db.refresh(item)
        return item

    def review_snapshot_hash(self, run: BulkEvaluationRun) -> str:
        items = list(
            self.db.scalars(
                select(BulkEvaluationItem)
                .where(
                    BulkEvaluationItem.run_id == run.id,
                    BulkEvaluationItem.status == "graded",
                    BulkEvaluationItem.grade_suggestion_id.is_not(None),
                )
                .order_by(BulkEvaluationItem.id)
            ).all()
        )
        payload: list[dict[str, object]] = []
        for item in items:
            suggestion = self.db.get(GradeSuggestion, item.grade_suggestion_id)
            if suggestion is None:
                continue
            payload.append(
                {
                    "item_id": item.id,
                    "suggestion_id": suggestion.id,
                    "answer_region_id": suggestion.answer_region_id,
                    "score": str(suggestion.score),
                    "max_score": str(suggestion.max_score),
                    "confidence": str(suggestion.confidence),
                    "evidence_sha256": item.evidence_snapshot_sha256,
                    "rubric_sha256": item.rubric_snapshot_sha256,
                }
            )
        return _canonical_hash(payload)

    def approve_clean(
        self,
        run: BulkEvaluationRun,
        *,
        suggestion_ids: list[int],
        review_snapshot_sha256: str,
        teacher_id: int,
    ) -> tuple[int, int, str]:
        if run.status not in {"review_ready", "completed_with_exceptions"}:
            raise BulkEvaluationError("Bulk run is not ready for teacher approval")
        current_hash = self.review_snapshot_hash(run)
        if current_hash != review_snapshot_sha256:
            raise BulkEvaluationError("Clean draft set changed; refresh before approving")
        eligible_items = list(
            self.db.scalars(
                select(BulkEvaluationItem).where(
                    BulkEvaluationItem.run_id == run.id,
                    BulkEvaluationItem.status == "graded",
                    BulkEvaluationItem.grade_suggestion_id.is_not(None),
                )
            ).all()
        )
        eligible_ids = {int(item.grade_suggestion_id) for item in eligible_items}
        if set(suggestion_ids) != eligible_ids or len(suggestion_ids) != len(eligible_ids):
            raise BulkEvaluationError(
                "Approval must match the exact clean draft snapshot; flagged items are excluded"
            )
        approved_count = 0
        already_approved = 0
        final_service = FinalGradeService(self.db)
        items_by_suggestion = {
            int(item.grade_suggestion_id): item for item in eligible_items
        }
        for suggestion_id in sorted(eligible_ids):
            item = items_by_suggestion[suggestion_id]
            existing = self.db.scalar(
                select(FinalGrade).where(FinalGrade.answer_region_id == item.answer_region_id)
            )
            final_grade, created = final_service.approve_suggestion(
                suggestion_id,
                teacher_id,
                teacher_comment="Approved from clean bulk-supervised snapshot",
            )
            item = self.db.get(BulkEvaluationItem, item.id)
            if item is not None:
                item.final_grade_id = final_grade.id
                item.status = "approved"
                item.stage = "complete"
                item.completed_at = datetime.now(UTC)
                self.db.commit()
            if created and existing is None:
                approved_count += 1
            else:
                already_approved += 1
        run = self.db.get(BulkEvaluationRun, run.id)
        assert run is not None
        self._refresh_counts(run)
        remaining = self.db.scalar(
            select(BulkEvaluationItem.id)
            .where(
                BulkEvaluationItem.run_id == run.id,
                BulkEvaluationItem.status.not_in(("approved", "exception", "uncertain")),
            )
            .limit(1)
        )
        run.status = (
            "completed"
            if run.exception_count == 0 and remaining is None
            else "completed_with_exceptions"
        )
        run.stage = "complete"
        run.review_snapshot_sha256 = current_hash
        run.completed_at = datetime.now(UTC)
        self._audit_run(
            run,
            "bulk_evaluation_clean_drafts_approved",
            teacher_id,
            {
                "approved_count": approved_count,
                "already_approved_count": already_approved,
                "review_snapshot_sha256": current_hash,
                "suggestion_ids": sorted(eligible_ids),
            },
        )
        self.db.commit()
        return approved_count, already_approved, current_hash

    def build_results_workbook(self, run: BulkEvaluationRun) -> bytes:
        rows = self.db.execute(
            select(BulkEvaluationItem, Submission, Question)
            .join(Submission, Submission.id == BulkEvaluationItem.submission_id)
            .join(Question, Question.id == BulkEvaluationItem.question_id)
            .where(BulkEvaluationItem.run_id == run.id)
            .order_by(Submission.student_identifier, Question.id)
        ).all()
        workbook = Workbook()
        approved_sheet = workbook.active
        approved_sheet.title = "Approved Scores"
        approved_sheet.append(
            [
                "student_identifier",
                "student_name",
                "question",
                "max_marks",
                "approved_score",
                "status",
                "suggestion_id",
                "final_grade_id",
            ]
        )
        totals: dict[int, dict[str, object]] = {}
        expected_questions = len({question.id for _, _, question in rows})
        for item, submission, question in rows:
            final_grade = (
                self.db.get(FinalGrade, item.final_grade_id)
                if item.final_grade_id
                else None
            )
            approved = bool(final_grade and final_grade.approval_status == "approved")
            approved_sheet.append(
                [
                    submission.student_identifier,
                    submission.student_name,
                    question.question_no,
                    question.total_marks,
                    final_grade.final_score if approved else None,
                    "APPROVED" if approved else "INCOMPLETE",
                    item.grade_suggestion_id,
                    item.final_grade_id,
                ]
            )
            total = totals.setdefault(
                submission.id,
                {
                    "identifier": submission.student_identifier,
                    "name": submission.student_name,
                    "score": Decimal("0"),
                    "max": Decimal("0"),
                    "approved": 0,
                },
            )
            if approved and final_grade is not None:
                total["score"] = Decimal(str(total["score"])) + final_grade.final_score
                total["max"] = Decimal(str(total["max"])) + question.total_marks
                total["approved"] = int(total["approved"]) + 1

        totals_sheet = workbook.create_sheet("Student Totals")
        totals_sheet.append(
            [
                "student_identifier",
                "student_name",
                "approved_score",
                "approved_max_marks",
                "approved_questions",
                "expected_questions",
                "status",
            ]
        )
        for value in totals.values():
            complete = int(value["approved"]) == expected_questions
            totals_sheet.append(
                [
                    value["identifier"],
                    value["name"],
                    value["score"],
                    value["max"],
                    value["approved"],
                    expected_questions,
                    "COMPLETE" if complete else "INCOMPLETE",
                ]
            )

        exception_sheet = workbook.create_sheet("Exceptions")
        exception_sheet.append(
            [
                "student_identifier",
                "question",
                "stage",
                "exception_codes",
                "diagnostic_item_id",
            ]
        )
        for item, submission, question in rows:
            if item.status not in {"exception", "uncertain"}:
                continue
            exception_sheet.append(
                [
                    submission.student_identifier,
                    question.question_no,
                    item.stage,
                    ", ".join(item.exception_codes),
                    item.id,
                ]
            )

        summary_sheet = workbook.create_sheet("Run Summary")
        for key, value in (
            ("run_id", run.id),
            ("assessment_id", run.assessment_id),
            ("provider", run.provider),
            ("model", run.model_name),
            ("policy_version", run.policy_version),
            ("status", run.status),
            ("submission_count", run.total_submissions),
            ("page_count", run.total_pages),
            ("provider_calls", run.calls_used),
            ("authorized_call_limit", run.authorized_call_limit),
            ("clean_items", run.clean_item_count),
            ("exceptions", run.exception_count),
            ("approved_items", run.approved_count),
            ("reference_bundle_sha256", run.reference_bundle_sha256),
            ("archive_sha256", run.archive_sha256),
        ):
            summary_sheet.append([key, value])
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _next_mapping_submission(self, run: BulkEvaluationRun) -> int | None:
        return self.db.scalar(
            select(BulkEvaluationItem.submission_id)
            .where(
                BulkEvaluationItem.run_id == run.id,
                BulkEvaluationItem.status == "pending",
                BulkEvaluationItem.stage == "mapping",
            )
            .order_by(BulkEvaluationItem.submission_id)
            .limit(1)
        )

    def _next_transcription_item(self, run: BulkEvaluationRun) -> BulkEvaluationItem | None:
        return self.db.scalar(
            select(BulkEvaluationItem)
            .where(
                BulkEvaluationItem.run_id == run.id,
                BulkEvaluationItem.status == "pending",
                BulkEvaluationItem.stage == "transcription",
            )
            .order_by(BulkEvaluationItem.id)
            .limit(1)
        )

    def _next_grading_item(self, run: BulkEvaluationRun) -> BulkEvaluationItem | None:
        return self.db.scalar(
            select(BulkEvaluationItem)
            .where(
                BulkEvaluationItem.run_id == run.id,
                BulkEvaluationItem.status == "pending",
                BulkEvaluationItem.stage == "grading",
            )
            .order_by(BulkEvaluationItem.id)
            .limit(1)
        )

    def _existing_submission_mappings(self, submission_id: int) -> list[AnswerRegionMapping]:
        """Mappings already on record for a script, with the geometry loaded.

        Segments are eager-loaded because the caller re-derives overlap and the
        mapping snapshot hash from them without a second provider call.
        """
        return list(
            self.db.scalars(
                select(AnswerRegionMapping)
                .options(
                    selectinload(AnswerRegionMapping.answer_region).selectinload(
                        AnswerRegion.segments
                    )
                )
                .where(AnswerRegionMapping.submission_id == submission_id)
                .order_by(AnswerRegionMapping.id)
            ).all()
        )

    def _process_mapping(self, run: BulkEvaluationRun) -> None:
        submission_id = self._next_mapping_submission(run)
        if submission_id is None:
            return
        submission = self.db.scalar(
            select(Submission)
            .options(selectinload(Submission.pages))
            .where(Submission.id == submission_id)
        )
        teacher = self.db.get(User, run.created_by_teacher_id)
        if submission is None or teacher is None:
            raise BulkEvaluationError("Bulk mapping ownership context is unavailable")
        items = list(
            self.db.scalars(
                select(BulkEvaluationItem).where(
                    BulkEvaluationItem.run_id == run.id,
                    BulkEvaluationItem.submission_id == submission.id,
                    BulkEvaluationItem.stage == "mapping",
                    BulkEvaluationItem.status.in_(("pending", "running")),
                )
            ).all()
        )
        if not items:
            return
        run.status = "mapping"
        run.stage = "mapping"

        mappings = self._existing_submission_mappings(submission.id)
        if mappings:
            # Re-entry, not a first pass. A teacher who resumes a flagged item
            # puts it back into this stage while the rest of the script is
            # already mapped, and LocalScriptPreparationService refuses to
            # prepare over its own output ("Draft mappings already exist").
            # That refusal used to escape as a run-level failure, which marked
            # every in-flight item -- on this script and on any other -- with a
            # provider_contract_failure that no provider had produced, burying
            # the real finding underneath it.
            #
            # Re-running the mapper would be pointless anyway: same pages, same
            # locked references, same answer. So settle the resumed items
            # against the mapping already on record, for no provider call, and
            # let each one land on its true finding.
            self._audit_run(
                run,
                "bulk_evaluation_mapping_reused",
                None,
                {
                    "submission_id": submission.id,
                    "item_count": len(items),
                    "mapping_count": len(mappings),
                },
                actor_type="worker",
            )
        else:
            remaining = run.authorized_call_limit - run.calls_used
            if len(submission.pages) > remaining:
                self._pause(run, "Provider-call budget is insufficient for the next script")
                return
            started = datetime.now(UTC)
            for item in items:
                item.status = "running"
                item.started_at = started
            self.db.commit()

            mappings = LocalScriptPreparationService(self.db).prepare(
                submission=submission,
                teacher=teacher,
                expected_model=run.model_name,
                replace_existing=False,
                maximum_ocr_calls=remaining,
            )
            calls = self._latest_mapping_call_count(submission.id)
            self._consume_calls(run, max(len(submission.pages), calls))
        mapping_by_question = {mapping.question_id: mapping for mapping in mappings}
        overlap_questions = self._overlapping_question_ids(mappings)
        for item in items:
            mapping = mapping_by_question.get(item.question_id)
            codes = self._mapping_exception_codes(mapping, overlap_questions)
            item.mapping_id = mapping.id if mapping else None
            item.answer_region_id = mapping.answer_region_id if mapping else None
            item.mapping_confidence = mapping.confidence if mapping else None
            if codes:
                item.status = "exception"
                item.stage = "mapping"
                item.exception_codes = codes
                item.warnings = self._mapping_warnings(mapping)
                item.completed_at = datetime.now(UTC)
                continue
            assert mapping is not None and mapping.answer_region is not None
            mapping.bulk_policy_verified = True
            mapping.bulk_verification_run_id = run.id
            item.status = "pending"
            item.stage = "transcription"
            # A re-settled item can arrive carrying the finding that flagged it
            # last time; it cleared this stage now, so it must not hand a stale
            # code down to transcription.
            item.exception_codes = []
            item.warnings = []
            item.completed_at = None
            item.verification_source = "bulk_policy"
            item.source_snapshot_sha256 = self._mapping_snapshot_hash(mapping)
            item.provider_call_count += 0
        run.heartbeat_at = datetime.now(UTC)
        self._refresh_counts(run)
        self.db.commit()

    def _process_transcription(self, run: BulkEvaluationRun) -> None:
        item = self._next_transcription_item(run)
        if item is None:
            return
        region = self.db.get(AnswerRegion, item.answer_region_id)
        teacher = self.db.get(User, run.created_by_teacher_id)
        if region is None or teacher is None:
            self._quarantine(item, "incomplete_region", "Mapped answer region is unavailable")
            self._refresh_counts(run)
            self.db.commit()
            return
        run.status = "transcribing"
        run.stage = "transcription"
        item.status = "running"
        item.started_at = item.started_at or datetime.now(UTC)
        self.db.commit()
        service = Qwen38VisualTranscriptionService(self.db)

        current_run = (
            self.db.get(AnswerRegionOcrRun, item.transcription_run_id)
            if item.transcription_run_id
            else None
        )
        if current_run is not None and current_run.profile == "qwen38_thinking_repair":
            required = max(1, current_run.call_limit)
            self._require_calls(run, required)
            service.run_thinking_repair(current_run.id)
            current_run = self.db.get(AnswerRegionOcrRun, current_run.id)
            used = max(1, current_run.calls_used if current_run else required)
            self._consume_calls(run, used)
            item.provider_call_count += used
            if current_run is None or current_run.status != "succeeded":
                error_msg = (
                    current_run.error
                    if current_run and current_run.error
                    else "Thinking repair could not resolve visual evidence"
                )
                self._quarantine(item, "ambiguous_symbol", error_msg)
                run.heartbeat_at = datetime.now(UTC)
                self._refresh_counts(run)
                self.db.commit()
                return
            if self._formula_dense(current_run.draft_text or ""):
                if not self._verify_formula_transcription(run, item, region, current_run):
                    self._quarantine(
                        item,
                        "verification_disagreement",
                        "Critical mathematical tokens disagreed across fresh visual reads",
                    )
                    self._refresh_counts(run)
                    self.db.commit()
                    return
            self._accept_or_quarantine_transcription(run, item, region, current_run)
        else:
            self._require_calls(run, 1)
            transcript_run = service.create(region, teacher=teacher, expected_model=run.model_name)
            item.transcription_run_id = transcript_run.id
            self.db.commit()
            service.run(transcript_run.id)
            transcript_run = self.db.get(AnswerRegionOcrRun, transcript_run.id)
            self._consume_calls(run, 1)
            item.provider_call_count += 1
            if transcript_run is None or transcript_run.status != "succeeded":
                raise VisualTranscriptionError("Visual transcription failed its provider contract")
            normalized = transcript_run.normalized_result or {}
            editing = normalized.get("editing_analysis") or {}
            needs_repair = bool(
                normalized.get("requires_thinking_repair")
                or (isinstance(editing, dict) and any(editing.get(key) for key in (
                    "cancellation_detected",
                    "replacement_detected",
                    "uncertain_correction_detected",
                )))
            )
            if needs_repair:
                self._require_calls(run, 3)
                repair = service.create_thinking_repair(
                    region,
                    transcript_run,
                    teacher=teacher,
                    expected_model=run.model_name,
                )
                item.transcription_run_id = repair.id
                item.status = "pending"
                item.warnings = ["thinking_repair_pending"]
                self.db.commit()
                return
            if self._formula_dense(transcript_run.draft_text or ""):
                if not self._verify_formula_transcription(
                    run, item, region, transcript_run
                ):
                    self._quarantine(
                        item,
                        "verification_disagreement",
                        "Critical mathematical tokens disagreed across fresh visual reads",
                    )
                    self._refresh_counts(run)
                    self.db.commit()
                    return
            self._accept_or_quarantine_transcription(run, item, region, transcript_run)
        run.heartbeat_at = datetime.now(UTC)
        self._refresh_counts(run)
        self.db.commit()

    def _accept_or_quarantine_transcription(
        self,
        run: BulkEvaluationRun,
        item: BulkEvaluationItem,
        region: AnswerRegion,
        transcript_run: AnswerRegionOcrRun,
    ) -> None:
        normalized = transcript_run.normalized_result or {}
        confidence = Decimal(str(normalized.get("confidence") or 0))
        item.transcription_confidence = confidence
        if normalized.get("is_blank"):
            self._quarantine(item, "probable_blank", "Probable blank answer needs teacher review")
            return
        editing = normalized.get("editing_analysis") or {}
        unresolved = bool(
            normalized.get("uncertain_glyphs")
            or (isinstance(editing, dict) and editing.get("uncertain_correction_detected"))
            or any(
                marker in (transcript_run.draft_text or "").casefold()
                for marker in ("[illegible", "[unclear", "[visible writing unresolved")
            )
        )
        if unresolved:
            self._quarantine(
                item, "ambiguous_symbol", "Transcription contains unresolved visual evidence"
            )
            return
        if confidence < self.settings.bulk_transcription_auto_pass_min_confidence:
            self._quarantine(
                item, "unreadable_handwriting", "Transcription confidence is below auto-pass"
            )
            return
        transcript_run.status = "confirmed"
        transcript_run.confirmed_text = transcript_run.draft_text
        transcript_run.confirmed_at = datetime.now(UTC)
        transcript_run.confirmed_by_teacher_id = None
        transcript_run.verification_source = "bulk_policy"
        transcript_run.bulk_verification_run_id = run.id
        region.manual_answer_text = transcript_run.draft_text
        region.evidence_status = "complete"
        region.full_answer_confirmed = False
        region.full_answer_verification_source = "bulk_policy"
        region.bulk_verification_run_id = run.id
        region.continuation_check_status = (
            "continuation_confirmed_included"
            if len(region.segments) > 1
            else "checked_no_continuation"
        )
        item.evidence_snapshot_sha256 = hashlib.sha256(
            (transcript_run.draft_text or "").encode("utf-8")
        ).hexdigest()
        item.verification_source = "bulk_policy"
        item.status = "pending"
        item.stage = "grading"
        item.exception_codes = []
        item.warnings = []

    def _verify_formula_transcription(
        self,
        run: BulkEvaluationRun,
        item: BulkEvaluationItem,
        region: AnswerRegion,
        source_run: AnswerRegionOcrRun,
    ) -> bool:
        self._require_calls(run, 1)
        verification = AnswerRegionOcrRun(
            answer_region_id=region.id,
            requested_by_teacher_id=run.created_by_teacher_id,
            request_id=f"qwen38-bulk-verify-{region.id}-{uuid4().hex}",
            status="running",
            profile="qwen38_formula_verification",
            task_kind="visual_transcription_verification",
            reasoning_mode="off",
            prompt_version=source_run.prompt_version,
            source_image_sha256=source_run.source_image_sha256,
            source_image_hashes=source_run.source_image_hashes,
            input_manifest_sha256=source_run.input_manifest_sha256,
            model_asset_sha256=self.settings.local_qwen38_model_sha256 or None,
            mmproj_asset_sha256=self.settings.local_qwen38_mmproj_sha256 or None,
            queued_at=datetime.now(UTC),
            started_at=datetime.now(UTC),
            heartbeat_at=datetime.now(UTC),
            call_limit=1,
            calls_used=1,
            provider="llama_cpp_qwen38",
            model_name=run.model_name,
            warnings=["bulk_formula_verification"],
        )
        self.db.add(verification)
        self.db.commit()
        holder_id = f"bulk_formula_verify:{verification.id}:{uuid4().hex}"
        lease = LocalModelLeaseService(self.db)
        try:
            with lease.hold(
                model_phase="Qwen38",
                holder_kind="visual_transcription",
                holder_id=holder_id,
            ):
                if self.settings.local_ai_phase_switch_enabled:
                    LocalAiPhaseManager(settings=self.settings, db=self.db).switch(
                        "Qwen38", lease_holder_id=holder_id
                    )
                adapter = BrainAdapter.for_provider(self.settings, "llama_cpp_qwen38")
                adapter.verify_available_model()
                images: list[tuple[bytes, str]] = []
                for segment in sorted(region.segments, key=lambda value: value.order_index):
                    images.append(
                        (
                            self.storage.resolve_relative(segment.image_path).read_bytes(),
                            "image/png",
                        )
                    )
                lease.heartbeat(holder_id=holder_id)
                result = adapter.provider.transcribe_images(
                    images=images,
                    label=region.question.question_no,
                )
                lease.heartbeat(holder_id=holder_id)
            verification.status = "succeeded"
            verification.completed_at = datetime.now(UTC)
            verification.heartbeat_at = verification.completed_at
            verification.draft_text = result.draft_text
            verification.output_sha256 = hashlib.sha256(
                result.draft_text.encode("utf-8")
            ).hexdigest()
            verification.normalized_result = {
                "task_kind": "visual_transcription_verification",
                "reasoning_mode": "off",
                "source_run_id": source_run.id,
                "critical_tokens": self._critical_tokens(result.draft_text),
                "confidence": str(result.confidence),
                "prompt_tokens": result.prompt_tokens,
                "completion_tokens": result.completion_tokens,
            }
            verification.latency_ms = result.latency_ms
            self._consume_calls(run, 1)
            item.provider_call_count += 1
            source_tokens = self._critical_tokens(source_run.draft_text or "")
            matched = source_tokens == self._critical_tokens(result.draft_text)
            verification.verification_source = "bulk_policy" if matched else None
            verification.bulk_verification_run_id = run.id if matched else None
            self.db.commit()
            return matched
        except Exception as exc:
            self.db.rollback()
            failed = self.db.get(AnswerRegionOcrRun, verification.id)
            if failed is not None:
                failed.status = "failed"
                failed.error = sanitize_provider_error(str(exc))[:500]
                failed.completed_at = datetime.now(UTC)
                self.db.commit()
            raise

    def _formula_dense(self, text: str) -> bool:
        return bool(re.search(r"(?:\\frac|\\bar|[=+\-×÷*/]|\d+\s*/\s*\d+)", text))

    def _critical_tokens(self, text: str) -> list[str]:
        normalized = text.replace("−", "-").replace("·", "*").casefold()
        return re.findall(
            r"\\(?:frac|bar|overline|sqrt|cap|cup)|\d+(?:\.\d+)?|[=+\-×÷*/<>]",
            normalized,
        )

    def _process_grading(self, run: BulkEvaluationRun) -> None:
        item = self._next_grading_item(run)
        if item is None:
            return
        region = self.db.get(AnswerRegion, item.answer_region_id)
        if region is None:
            self._quarantine(item, "stale_evidence", "Answer evidence disappeared before grading")
            self.db.commit()
            return
        self._require_calls(run, 1)
        run.status = "grading"
        run.stage = "grading"
        item.status = "running"
        item.started_at = item.started_at or datetime.now(UTC)
        self.db.commit()
        adapter = BrainAdapter.for_provider(self.settings, "llama_cpp_qwen38")
        grading = GradingService(self.db, adapter=adapter)
        rubric = self.db.scalar(
            select(Rubric)
            .where(Rubric.question_id == region.question_id, Rubric.is_active.is_(True))
            .order_by(Rubric.version.desc(), Rubric.id.desc())
        )
        if rubric is None:
            self._quarantine(item, "stale_evidence", "Active rubric is unavailable")
            self.db.commit()
            return
        rubric_hash = rubric_snapshot_hash(region.question, rubric)
        job = grading.create_queued_grading_job(region.id)
        item.grading_job_id = job.id
        self.db.commit()
        job, suggestion = grading.run_queued_job(
            job.id,
            marking_policy=run.marking_policy,
            expected_rubric_id=rubric.id,
            expected_rubric_hash=rubric_hash,
        )
        self._consume_calls(run, 1)
        item.provider_call_count += 1
        item.grading_job_id = job.id
        item.grade_suggestion_id = suggestion.id
        item.grading_confidence = suggestion.confidence
        item.rubric_snapshot_sha256 = rubric_hash
        suggestion.raw_response_json = {
            **(suggestion.raw_response_json or {}),
            "bulk_evaluation_run_id": run.id,
            "bulk_policy_version": run.policy_version,
            "image_input_disabled": True,
            "local_provider": True,
            "teacher_review_required": True,
        }
        if (
            suggestion.confidence is None
            or suggestion.confidence < self.settings.bulk_grading_clean_min_confidence
        ):
            self._quarantine(item, "verification_disagreement", "Draft grade confidence is low")
            item.stage = "review"
        else:
            item.status = "graded"
            item.stage = "review"
            item.completed_at = datetime.now(UTC)
        run.heartbeat_at = datetime.now(UTC)
        self._refresh_counts(run)
        self.db.commit()

    def _finish_processing(self, run: BulkEvaluationRun) -> None:
        self._refresh_counts(run)
        run.stage = "review"
        run.status = "review_ready" if run.clean_item_count else "completed_with_exceptions"
        run.review_snapshot_sha256 = self.review_snapshot_hash(run)
        run.completed_at = datetime.now(UTC)
        run.heartbeat_at = run.completed_at
        self._audit_run(
            run,
            "bulk_evaluation_processing_completed",
            None,
            {
                "calls_used": run.calls_used,
                "clean_item_count": run.clean_item_count,
                "exception_count": run.exception_count,
            },
            actor_type="worker",
        )
        self.db.commit()

    def _mapping_exception_codes(
        self, mapping: AnswerRegionMapping | None, overlap_questions: set[int]
    ) -> list[str]:
        if mapping is None or mapping.answer_region is None:
            return ["missing_answer"]
        codes: list[str] = []
        if mapping.question_id in overlap_questions:
            codes.append("cross_question_overlap")
        if (
            mapping.blocker_reason
            and mapping.blocker_reason != "Qwen marked this mapping uncertain"
        ):
            codes.append("incomplete_region")
        warnings = " ".join(self._mapping_warnings(mapping)).casefold()
        if "continuation" in warnings:
            codes.append("possible_continuation")
        if "outside every prepared" in warnings or "not assigned" in warnings:
            codes.append("unassigned_ink")
        if mapping.confidence is None or (
            mapping.confidence < self.settings.bulk_mapping_auto_pass_min_confidence
        ):
            codes.append("incomplete_region")
        return list(dict.fromkeys(codes))

    def _mapping_warnings(self, mapping: AnswerRegionMapping | None) -> list[str]:
        if mapping is None:
            return ["No answer region was mapped"]
        source = mapping.source_reference or {}
        values = [*(source.get("warnings") or []), *(source.get("qwen_warnings") or [])]
        if (
            mapping.blocker_reason
            and mapping.blocker_reason != "Qwen marked this mapping uncertain"
        ):
            values.append(mapping.blocker_reason)
        return list(dict.fromkeys(str(value) for value in values if value))

    def _overlapping_question_ids(self, mappings: list[AnswerRegionMapping]) -> set[int]:
        segments: list[tuple[int, int, float, float, float, float]] = []
        for mapping in mappings:
            if mapping.question_id is None or mapping.answer_region is None:
                continue
            for segment in mapping.answer_region.segments:
                segments.append(
                    (
                        mapping.question_id,
                        segment.submission_page_id,
                        float(segment.x),
                        float(segment.y),
                        float(segment.width),
                        float(segment.height),
                    )
                )
        result: set[int] = set()
        for index, left in enumerate(segments):
            for right in segments[index + 1 :]:
                if left[0] == right[0] or left[1] != right[1]:
                    continue
                overlap_width = max(
                    0.0,
                    min(left[2] + left[4], right[2] + right[4])
                    - max(left[2], right[2]),
                )
                overlap_height = max(
                    0.0,
                    min(left[3] + left[5], right[3] + right[5])
                    - max(left[3], right[3]),
                )
                overlap = overlap_width * overlap_height
                smaller = min(left[4] * left[5], right[4] * right[5])
                if smaller > 0 and overlap / smaller > 0.10:
                    result.update((left[0], right[0]))
        return result

    def _mapping_snapshot_hash(self, mapping: AnswerRegionMapping) -> str:
        region = mapping.answer_region
        return _canonical_hash(
            {
                "mapping_id": mapping.id,
                "question_id": mapping.question_id,
                "confidence": str(mapping.confidence),
                "segments": [
                    {
                        "page_id": segment.submission_page_id,
                        "order": segment.order_index,
                        "x": str(segment.x),
                        "y": str(segment.y),
                        "width": str(segment.width),
                        "height": str(segment.height),
                    }
                    for segment in (region.segments if region else [])
                ],
            }
        )

    def _latest_mapping_call_count(self, submission_id: int) -> int:
        audit = self.db.scalar(
            select(AuditLog)
            .where(
                AuditLog.entity_type == "submission",
                AuditLog.entity_id == submission_id,
                AuditLog.event_type == "submission_script_draft_prepared",
            )
            .order_by(AuditLog.id.desc())
        )
        return int((audit.payload_json or {}).get("visual_mapping_call_count", 0)) if audit else 0

    def _consume_calls(self, run: BulkEvaluationRun, count: int) -> None:
        if count < 0 or run.calls_used + count > run.authorized_call_limit:
            raise BulkEvaluationError("Provider-call budget would be exceeded")
        run.calls_used += count

    def _require_calls(self, run: BulkEvaluationRun, count: int) -> None:
        if run.calls_used + count > run.authorized_call_limit:
            self._pause(run, "Provider-call budget is insufficient for the next task")
            raise BulkEvaluationError("Provider-call budget is insufficient for the next task")

    def _quarantine(self, item: BulkEvaluationItem, code: str, warning: str) -> None:
        item.status = "exception"
        item.exception_codes = list(dict.fromkeys([*(item.exception_codes or []), code]))
        item.warnings = list(dict.fromkeys([*(item.warnings or []), warning]))
        item.completed_at = datetime.now(UTC)

    def _mark_running_uncertain(self, run: BulkEvaluationRun, warning: str) -> None:
        items = list(
            self.db.scalars(
                select(BulkEvaluationItem).where(
                    BulkEvaluationItem.run_id == run.id,
                    BulkEvaluationItem.status == "running",
                )
            ).all()
        )
        now = datetime.now(UTC)
        for item in items:
            item.status = "uncertain"
            item.exception_codes = list(
                dict.fromkeys([*(item.exception_codes or []), "provider_contract_failure"])
            )
            item.warnings = list(dict.fromkeys([*(item.warnings or []), warning]))
            item.completed_at = now
        self._release_orphaned_transcriptions(run, warning, now)
        self._refresh_counts(run)

    def _release_orphaned_transcriptions(
        self, run: BulkEvaluationRun, warning: str, now: datetime
    ) -> None:
        """Close transcription rows this run left marked in flight.

        Reclaiming the item is not enough, and reclaiming only the items that
        are *currently* running is not enough either. A real interrupted run
        left AnswerRegionOcrRun 88 with status "running" and completed_at NULL
        from one day to the next: the pause had already flipped its item to
        "uncertain", so by the time recovery ran there was no running item left
        pointing at it and the ledger kept claiming a provider call was in
        flight for a process that had died. Reconciling by ledger state rather
        than by item state heals that orphan and the fresh-interruption case
        with one rule.
        """
        run_items = list(
            self.db.scalars(
                select(BulkEvaluationItem).where(BulkEvaluationItem.run_id == run.id)
            ).all()
        )
        for item in run_items:
            if not item.transcription_run_id:
                continue
            transcript_run = self.db.get(AnswerRegionOcrRun, item.transcription_run_id)
            # Only in-flight rows are reclaimed. A succeeded transcript is
            # evidence a teacher may already have seen; never restamp it.
            if transcript_run is None or transcript_run.status != "running":
                continue
            transcript_run.status = "failed"
            transcript_run.completed_at = now
            transcript_run.heartbeat_at = now
            transcript_run.error = warning
            transcript_run.warnings = list(
                dict.fromkeys([*(transcript_run.warnings or []), "interrupted_run_reclaimed"])
            )

    def _refresh_counts(self, run: BulkEvaluationRun) -> None:
        items = list(
            self.db.scalars(
                select(BulkEvaluationItem).where(BulkEvaluationItem.run_id == run.id)
            ).all()
        )
        run.processed_items = sum(
            item.status in {"exception", "uncertain", "graded", "approved"}
            for item in items
        )
        run.clean_item_count = sum(item.status in {"graded", "approved"} for item in items)
        run.exception_count = sum(item.status in {"exception", "uncertain"} for item in items)
        run.approved_count = sum(item.status == "approved" for item in items)

    def _status_for_pending_stage(self, run: BulkEvaluationRun) -> str:
        stages = {
            item.stage
            for item in run.items
            if item.status in {"pending", "running"}
        }
        if "mapping" in stages:
            return "mapping"
        if "transcription" in stages:
            return "transcribing"
        if "grading" in stages:
            return "grading"
        return "review_ready"

    def _pause(self, run: BulkEvaluationRun, error: str) -> None:
        run.status = "paused"
        run.error = sanitize_provider_error(error)[:500]
        run.heartbeat_at = datetime.now(UTC)
        self._audit_run(
            run,
            "bulk_evaluation_paused",
            None,
            {"failure_category": "provider_or_integrity_failure"},
            actor_type="worker",
        )
        self.db.commit()

    def _audit_run(
        self,
        run: BulkEvaluationRun,
        event_type: str,
        actor_id: int | None,
        extra: dict[str, object] | None = None,
        *,
        actor_type: str = "teacher",
    ) -> None:
        self.db.add(
            AuditLog(
                actor_type=actor_type,
                actor_id=actor_id,
                event_type=event_type,
                entity_type="bulk_evaluation_run",
                entity_id=run.id,
                payload_json={
                    "status": run.status,
                    "stage": run.stage,
                    "calls_used": run.calls_used,
                    **(extra or {}),
                },
            )
        )

    def _cleanup_submissions(self, submission_ids: list[int]) -> None:
        for submission_id in submission_ids:
            self.storage.delete_submission_files(submission_id, [])
