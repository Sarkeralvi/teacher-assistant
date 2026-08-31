from __future__ import annotations

import argparse
import hashlib
import io
import json
import os
import re
import shutil
import subprocess
import unicodedata
from collections import Counter
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from enum import Enum, StrEnum
from pathlib import Path
from statistics import median
from typing import Any, Literal

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as WorkbookImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFilter, ImageFont
from pydantic import BaseModel, ConfigDict, Field, field_validator, model_validator
from sqlalchemy.engine import make_url

PROTOCOL_VERSION = "local-curated-qwen38-vision-v2"
SCHEMA_VERSION = 2
# The grading model is not settled: Qwen3.6 and Qwen3.8 are being compared on
# measured mark accuracy before one is pinned. Until that decision lands, a run
# must name which alias it used rather than the harness assuming one.
SUPPORTED_GRADING_MODELS = ("qwen3.6-35b-a3b-q4km",)
# Old signed bake-off artifacts remain readable after Qwen3.8 grading was
# retired from executable contracts. This does not authorize a new Qwen3.8
# grading run; run_grading_stage accepts SUPPORTED_GRADING_MODELS only.
HISTORICAL_GRADING_MODELS = (*SUPPORTED_GRADING_MODELS, "qwen3.8-27b-q4km")
EXPECTED_LLAMA_CPP_BUILD = "10249"
EXPECTED_PROMPT_VERSION = "real-grading-v3"
OCR_CALL_LIMIT = 20
QWEN_CALL_LIMIT = 18
DEFAULT_SEED = 360_020
_EVALUATION_SCOPE_ENV = "LOCAL_CURATED_EVALUATION_RUN_ID"

_RUN_ID_PATTERN = re.compile(r"^[a-z][a-z0-9_]{2,63}$")
_STATE_ORDER = (
    "prepared",
    "ground_truth_locked",
    "ocr_completed",
    "ocr_confirmed",
    "grading_completed",
    "review_completed",
    "reported",
)
_OCR_QUALITY_NO_GO_REASONS = {
    "transcription_mismatch",
    "blank_hallucination",
    "critical_symbol_or_number",
    "missing_or_extra_content",
    "image_or_region_problem",
}
_REVIEW_REASONS = {
    "none",
    "rubric_ambiguity",
    "ocr_unfixed",
    "model_error",
    "dataset_error",
}
_EXPECTED_CASE_IDS = tuple(f"{pack}{index}" for pack in "ABCDE" for index in range(1, 5))
_BLANK_CASE_IDS = {"A4", "C4"}
_REQUIRED_GRADING_SAFETY_CHECKS = {
    "approved_only_export_has_zero_data_rows",
    "audit_payload_has_no_raw_answer_text",
    "blank_qwen_calls_zero",
    "confirmed_text_hashes_match",
    "cross_teacher_dispatch_access_refused",
    "cross_teacher_export_access_refused",
    "cross_teacher_review_access_refused",
    "dispatch_authorization_contract_exact",
    "evidence_and_rubric_hashes_pinned",
    "exactly_18_qwen_calls",
    "mandatory_review_flags",
    "needs_review_true",
    "no_final_grade",
    "queue_18_fresh_2_blank_refused",
    "qwen38_visual_transcription_reasoning_off",
    "single_model_gpu_slot_enforced",
    "zero_cloud_calls",
    "zero_fallback_calls",
    "zero_monetary_cost",
    "zero_retries",
}


class LocalCuratedEvaluationError(RuntimeError):
    """Raised when an evaluation gate or integrity rule would be violated."""


class PrimaryCategory(StrEnum):
    CORRECT = "correct"
    PARTIAL = "partial"
    WRONG = "wrong"
    BLANK = "blank"
    IRRELEVANT = "irrelevant"
    DIFFICULT_HANDWRITING = "difficult_handwriting"
    FORMULA_HEAVY = "formula_heavy"
    MULTI_STEP = "multi_step"


class AnswerQuality(StrEnum):
    CORRECT = "correct"
    PARTIAL = "partial"
    WRONG = "wrong"
    BLANK = "blank"
    IRRELEVANT = "irrelevant"


class EvaluationVerdict(StrEnum):
    PASS = "PASS"
    NO_GO_QUALITY = "NO_GO_QUALITY"
    INVALID_RUN = "INVALID_RUN"


class RubricCriterion(BaseModel):
    model_config = ConfigDict(extra="forbid")

    id: str = Field(min_length=1, max_length=64)
    name: str = Field(min_length=1, max_length=255)
    description: str = Field(min_length=1)
    max_marks: Decimal = Field(gt=Decimal("0"))
    depends_on: list[str] = Field(default_factory=list)
    allow_follow_through: bool = False


class LocalCuratedCaseDefinition(BaseModel):
    model_config = ConfigDict(extra="forbid")

    case_id: str = Field(pattern=r"^[A-E][1-4]$")
    pack_id: str = Field(pattern=r"^[A-E]$")
    primary_category: PrimaryCategory
    answer_quality: AnswerQuality
    question_text: str = Field(min_length=1)
    model_answer: str = Field(min_length=1)
    rubric: list[RubricCriterion] = Field(min_length=1)
    expected_score: Decimal = Field(ge=Decimal("0"))
    max_score: Decimal = Field(gt=Decimal("0"))
    authored_transcription: str
    teacher_notes: str = Field(min_length=1)
    render_profile: Literal["typed", "handwriting", "formula", "blank", "multi_step"]
    clean_typed: bool
    critical_tokens: list[str] = Field(default_factory=list)
    image_relative_path: str = ""
    image_sha256: str = ""

    @model_validator(mode="after")
    def validate_case(self) -> LocalCuratedCaseDefinition:
        rubric_total = sum((criterion.max_marks for criterion in self.rubric), Decimal("0"))
        if rubric_total != self.max_score:
            raise ValueError("rubric marks must sum to max_score")
        if self.expected_score > self.max_score:
            raise ValueError("expected_score cannot exceed max_score")
        if self.pack_id != self.case_id[0]:
            raise ValueError("pack_id must match the case ID prefix")
        if self.answer_quality == AnswerQuality.BLANK:
            if self.authored_transcription:
                raise ValueError("blank cases cannot contain authored answer text")
            if self.expected_score != 0:
                raise ValueError("blank cases must have an expected score of zero")
        if self.primary_category == PrimaryCategory.FORMULA_HEAVY and not self.critical_tokens:
            raise ValueError("formula-heavy cases require critical tokens")
        return self


class EvaluationThresholds(BaseModel):
    model_config = ConfigDict(extra="forbid")

    overall_nonblank_cer_max: Decimal = Decimal("0.10")
    clean_typed_cer_max: Decimal = Decimal("0.02")
    handwriting_mean_cer_max: Decimal = Decimal("0.20")
    handwriting_case_cer_max: Decimal = Decimal("0.30")
    formula_critical_token_recall_min: Decimal = Decimal("1.00")
    grading_exact_min: int = 9
    grading_within_one_min: int = 16
    grading_mae_max: Decimal = Decimal("0.75")
    grading_normalized_mae_max: Decimal = Decimal("0.15")
    irrelevant_score_max: Decimal = Decimal("0.50")
    severe_confidence_threshold: Decimal = Decimal("0.80")
    severe_low_confidence_max: int = 1


class LlamaCppAssetMetadata(BaseModel):
    model_config = ConfigDict(extra="forbid")

    build: Literal["10249"]
    # Either candidate grading model may be pinned until the bake-off decides.
    model_alias: Literal["qwen3.6-35b-a3b-q4km", "qwen3.8-27b-q4km"]
    model_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    model_size_bytes: int = Field(gt=0)
    device: Literal["gpu_hybrid", "gpu_hybrid_single_slot"] = "gpu_hybrid"


class Qwen38VisionAssetMetadata(LlamaCppAssetMetadata):
    """Pinned assets for the only visual transcription engine.

    Qwen3.8 requires both the language GGUF and its multimodal projector.  A
    run that cannot record both is not allowed to claim that it evaluated the
    visual path.
    """

    model_alias: Literal["qwen3.8-27b-q4km"]
    device: Literal["gpu_hybrid_single_slot"] = "gpu_hybrid_single_slot"
    mmproj_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    mmproj_size_bytes: int = Field(gt=0)


class OperatorAssetMetadata(BaseModel):
    model_config = ConfigDict(extra="forbid")

    # ``llama_cpp`` is the text-only candidate selected for this particular
    # grading run.  ``qwen38_vision`` is always present because all handwriting
    # transcription uses Qwen3.8 in fresh, non-thinking multimodal calls.
    llama_cpp: LlamaCppAssetMetadata
    qwen38_vision: Qwen38VisionAssetMetadata


class LocalCuratedEvaluationManifest(BaseModel):
    model_config = ConfigDict(extra="forbid")

    schema_version: Literal[2] = SCHEMA_VERSION
    protocol_version: str = PROTOCOL_VERSION
    run_id: str
    seed: int
    created_at: datetime
    integration_commit: str = Field(pattern=r"^[0-9a-f]{40}$")
    harness_commit: str = Field(pattern=r"^[0-9a-f]{40}$")
    # No default: the run must state which grading model it used, because the
    # choice between Qwen3.6 and Qwen3.8 is still being decided on measurements.
    expected_qwen_model: str
    prompt_version: Literal["real-grading-v3"] = EXPECTED_PROMPT_VERSION
    operator_assets: OperatorAssetMetadata
    transport: Literal["direct_host_eval"] = "direct_host_eval"
    ocr_call_limit: Literal[20] = OCR_CALL_LIMIT
    qwen_call_limit: Literal[18] = QWEN_CALL_LIMIT
    marking_policy: Literal["general"] = "general"
    thresholds: EvaluationThresholds = Field(default_factory=EvaluationThresholds)
    cases: list[LocalCuratedCaseDefinition] = Field(min_length=20, max_length=20)

    @field_validator("run_id")
    @classmethod
    def validate_run_id(cls, value: str) -> str:
        if not _RUN_ID_PATTERN.fullmatch(value):
            raise ValueError("run_id must use lowercase letters, digits, and underscores")
        return value

    @model_validator(mode="after")
    def validate_blueprint(self) -> LocalCuratedEvaluationManifest:
        validate_case_blueprint(self.cases)
        for case in self.cases:
            if case.image_relative_path != f"images/{case.case_id}.png":
                raise ValueError("manifest image paths must be canonical relative paths")
            if not re.fullmatch(r"[0-9a-f]{64}", case.image_sha256):
                raise ValueError("manifest image hashes must be SHA-256 values")
        return self


class GroundTruthCase(BaseModel):
    model_config = ConfigDict(extra="forbid")

    case_id: str
    teacher_transcription: str
    teacher_score: Decimal
    teacher_notes: str = ""
    handwriting_acceptable: bool | None = None
    approved: Literal[True]


class GroundTruthLock(BaseModel):
    model_config = ConfigDict(extra="forbid")

    run_id: str
    reviewer_id: str = Field(min_length=1, max_length=255)
    signed_at: datetime
    manifest_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    workbook_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    cases: list[GroundTruthCase] = Field(min_length=20, max_length=20)

    @model_validator(mode="after")
    def validate_cases(self) -> GroundTruthLock:
        if [case.case_id for case in self.cases] != list(_EXPECTED_CASE_IDS):
            raise ValueError("ground-truth cases must match the canonical 20-case order")
        return self


class OcrCaseResult(BaseModel):
    model_config = ConfigDict(extra="forbid")

    case_id: str
    answer_region_id: int = Field(gt=0)
    ocr_run_id: int = Field(gt=0)
    attempt_count: Literal[1] = 1
    status: Literal["succeeded"]
    draft_text: str
    markdown: str
    blocks: list[dict[str, Any]] = Field(default_factory=list)
    warnings: list[str] = Field(default_factory=list)
    latency_ms: int = Field(ge=0)
    provider: Literal["llama_cpp_qwen38"]
    model: Literal["qwen3.8-27b-q4km"]
    profile: Literal["qwen38_verbatim_visual"]
    reasoning_mode: Literal["off"]
    device: Literal["gpu_hybrid_single_slot"] = "gpu_hybrid_single_slot"
    source_image_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    source_image_hashes: list[str] = Field(min_length=1)
    is_blank: bool
    draft_text_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")

    @model_validator(mode="after")
    def validate_draft_hash(self) -> OcrCaseResult:
        if self.draft_text_sha256 != sha256_text(self.draft_text):
            raise ValueError("OCR draft hash does not match draft text")
        if any(not re.fullmatch(r"[0-9a-f]{64}", item) for item in self.source_image_hashes):
            raise ValueError("visual transcription source hashes must be SHA-256 values")
        if self.source_image_sha256 != sha256_text("".join(self.source_image_hashes)):
            raise ValueError("visual transcription source hash does not match source images")
        if self.is_blank != (not self.draft_text.strip()):
            raise ValueError("visual transcription blank flag does not match draft text")
        return self


class OcrRunResult(BaseModel):
    model_config = ConfigDict(extra="forbid")

    run_id: str
    first_call_at: datetime
    completed_at: datetime
    # ``call_count`` is the number of Qwen3.8 visual calls represented by this
    # evidence packet.  A grading-model bake-off may replay that already
    # teacher-confirmed packet into a fresh disposable database, but it must
    # never present the replay as twenty fresh visual calls.  These fields make
    # the provenance explicit and keep the historical source-call count intact.
    evidence_origin: Literal["live_qwen38_visual", "locked_bakeoff_replay"] = (
        "live_qwen38_visual"
    )
    source_evaluation_run_id: str | None = None
    new_provider_call_count: Literal[0, 20] = 20
    call_count: Literal[20]
    retry_count: Literal[0] = 0
    service_status_before: dict[str, Any]
    service_status_after: dict[str, Any]
    database_name: str
    assessment_id: int = Field(gt=0)
    grading_run_id: int = Field(gt=0)
    owner_teacher_id: int = Field(gt=0)
    intruder_teacher_id: int = Field(gt=0)
    question_ids: dict[str, int]
    rubric_ids: dict[str, int]
    cases: list[OcrCaseResult] = Field(min_length=20, max_length=20)

    @model_validator(mode="after")
    def validate_run(self) -> OcrRunResult:
        if self.completed_at < self.first_call_at:
            raise ValueError("OCR completion cannot precede its first call")
        if [case.case_id for case in self.cases] != list(_EXPECTED_CASE_IDS):
            raise ValueError("OCR results must match the canonical 20-case order")
        if len({case.answer_region_id for case in self.cases}) != OCR_CALL_LIMIT:
            raise ValueError("OCR results require 20 unique answer regions")
        if len({case.ocr_run_id for case in self.cases}) != OCR_CALL_LIMIT:
            raise ValueError("OCR results require 20 unique OCR runs")
        if self.evidence_origin == "live_qwen38_visual":
            if self.source_evaluation_run_id is not None or self.new_provider_call_count != 20:
                raise ValueError(
                    "live visual evidence must record exactly 20 newly made provider calls"
                )
        elif not self.source_evaluation_run_id or self.new_provider_call_count != 0:
            raise ValueError(
                "locked bake-off evidence must identify its source and record zero new visual calls"
            )
        if any(
            case.provider != "llama_cpp_qwen38"
            or case.model != "qwen3.8-27b-q4km"
            or case.profile != "qwen38_verbatim_visual"
            or case.reasoning_mode != "off"
            or case.device != "gpu_hybrid_single_slot"
            for case in self.cases
        ):
            raise ValueError("OCR result metadata does not match Qwen3.8 visual transcription")
        if set(self.question_ids) != set("ABCDE") or set(self.rubric_ids) != set("ABCDE"):
            raise ValueError("OCR run must pin all five questions and rubrics")
        return self


class OcrConfirmationCase(BaseModel):
    model_config = ConfigDict(extra="forbid")

    case_id: str
    answer_region_id: int = Field(gt=0)
    ocr_run_id: int = Field(gt=0)
    confirmed_text: str
    confirmed_text_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    teacher_approved: Literal[True]
    evidence_status: Literal["complete", "blank"]
    full_answer_confirmed: bool

    @model_validator(mode="after")
    def validate_confirmation(self) -> OcrConfirmationCase:
        if self.confirmed_text_sha256 != sha256_text(self.confirmed_text):
            raise ValueError("confirmed-text hash does not match confirmed text")
        if self.evidence_status == "blank" and self.confirmed_text:
            raise ValueError("blank confirmation cannot contain text")
        if self.full_answer_confirmed != (self.evidence_status == "complete"):
            raise ValueError("full-answer confirmation must match evidence status")
        return self


class OcrConfirmationLock(BaseModel):
    model_config = ConfigDict(extra="forbid")

    run_id: str
    reviewer_id: str = Field(min_length=1, max_length=255)
    signed_at: datetime
    ocr_results_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    workbook_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    cases: list[OcrConfirmationCase] = Field(min_length=20, max_length=20)

    @model_validator(mode="after")
    def validate_cases(self) -> OcrConfirmationLock:
        if [case.case_id for case in self.cases] != list(_EXPECTED_CASE_IDS):
            raise ValueError("OCR confirmations must match the canonical 20-case order")
        blank_ids = {case.case_id for case in self.cases if case.evidence_status == "blank"}
        if blank_ids != _BLANK_CASE_IDS:
            raise ValueError("only A4 and C4 may be confirmed as blank")
        if len({case.answer_region_id for case in self.cases}) != OCR_CALL_LIMIT:
            raise ValueError("OCR confirmations require 20 unique answer regions")
        if len({case.ocr_run_id for case in self.cases}) != OCR_CALL_LIMIT:
            raise ValueError("OCR confirmations require 20 unique OCR runs")
        return self


class OcrQualityNoGoCase(BaseModel):
    """A teacher-declared visual-transcription failure, without raw answer text."""

    model_config = ConfigDict(extra="forbid")

    case_id: str
    reason: Literal[
        "transcription_mismatch",
        "blank_hallucination",
        "critical_symbol_or_number",
        "missing_or_extra_content",
        "image_or_region_problem",
    ]


class OcrQualityNoGoLock(BaseModel):
    """Locks a valid quality failure before a grade could be created."""

    model_config = ConfigDict(extra="forbid")

    run_id: str
    reviewer_id: str = Field(min_length=1, max_length=255)
    signed_at: datetime
    ocr_results_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    workbook_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    rejected_cases: list[OcrQualityNoGoCase] = Field(min_length=1, max_length=20)

    @model_validator(mode="after")
    def validate_cases(self) -> OcrQualityNoGoLock:
        case_ids = [case.case_id for case in self.rejected_cases]
        if len(case_ids) != len(set(case_ids)):
            raise ValueError("OCR quality rejection cases must be unique")
        if any(case_id not in _EXPECTED_CASE_IDS for case_id in case_ids):
            raise ValueError("OCR quality rejection includes an unknown case")
        return self


class GradingCaseResult(BaseModel):
    model_config = ConfigDict(extra="forbid")

    case_id: str
    answer_region_id: int = Field(gt=0)
    outcome: Literal["suggested", "not_called_blank_safety_gate"]
    dispatch_run_id: int | None = None
    dispatch_item_id: int | None = None
    grading_job_id: int | None = None
    grade_suggestion_id: int | None = None
    ai_score: Decimal | None = None
    max_score: Decimal = Field(gt=Decimal("0"))
    confidence: Decimal | None = Field(default=None, ge=Decimal("0"), le=Decimal("1"))
    needs_review: bool | None = None
    rubric_breakdown: list[dict[str, Any]] = Field(default_factory=list)
    review_flags: list[str] = Field(default_factory=list)
    model_provider: str | None = None
    model_name: str | None = None
    prompt_version: str | None = None
    marking_policy: str | None = None
    token_usage: dict[str, Any] = Field(default_factory=dict)
    latency_ms: int | None = Field(default=None, ge=0)
    cost_estimate: Decimal | None = Field(default=None, ge=Decimal("0"))
    confirmed_text_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")

    @model_validator(mode="after")
    def validate_outcome(self) -> GradingCaseResult:
        suggestion_fields = (
            self.dispatch_run_id,
            self.dispatch_item_id,
            self.grading_job_id,
            self.grade_suggestion_id,
            self.ai_score,
            self.confidence,
            self.needs_review,
            self.model_provider,
            self.model_name,
            self.prompt_version,
            self.marking_policy,
        )
        if self.outcome == "suggested" and any(value is None for value in suggestion_fields):
            raise ValueError("suggested results require complete provider metadata")
        if self.outcome == "suggested":
            required_flags = {
                "image_input_disabled",
                "local_provider",
                "teacher_review_required",
            }
            if (
                self.model_provider not in {"llama_cpp_qwen", "llama_cpp_qwen38"}
                or self.model_name not in HISTORICAL_GRADING_MODELS
                or self.prompt_version != EXPECTED_PROMPT_VERSION
                or self.marking_policy != "general"
                or self.needs_review is not True
                or self.cost_estimate != 0
                or not required_flags.issubset(self.review_flags)
                or not self.rubric_breakdown
                or self.latency_ms is None
            ):
                raise ValueError("suggested result does not meet local draft-safety metadata")
            if self.ai_score is None or self.ai_score < 0 or self.ai_score > self.max_score:
                raise ValueError("suggested score must be inside the rubric range")
            usage_values = [
                self.token_usage.get(name)
                for name in ("prompt_tokens", "completion_tokens", "total_tokens")
            ]
            if any(not isinstance(value, int) or value < 0 for value in usage_values):
                raise ValueError("suggested result requires complete token metadata")
        if self.outcome == "not_called_blank_safety_gate" and any(
            value is not None for value in suggestion_fields
        ):
            raise ValueError("blank safety refusals cannot contain provider output")
        if self.outcome == "not_called_blank_safety_gate" and any(
            (
                self.rubric_breakdown,
                self.review_flags,
                self.token_usage,
                self.latency_ms is not None,
                self.cost_estimate is not None,
            )
        ):
            raise ValueError("blank safety refusals cannot contain hidden provider metadata")
        return self


class GradingRunResult(BaseModel):
    model_config = ConfigDict(extra="forbid")

    run_id: str
    first_call_at: datetime
    completed_at: datetime
    qwen_call_count: Literal[18]
    blank_refusal_count: Literal[2]
    retry_count: Literal[0] = 0
    fallback_call_count: Literal[0] = 0
    cloud_call_count: Literal[0] = 0
    transport: Literal["direct_host_eval"] = "direct_host_eval"
    dispatch_run_ids: list[int] = Field(min_length=5, max_length=5)
    cases: list[GradingCaseResult] = Field(min_length=20, max_length=20)
    safety_checks: dict[str, bool]

    @model_validator(mode="after")
    def validate_run(self) -> GradingRunResult:
        if self.completed_at < self.first_call_at:
            raise ValueError("grading completion cannot precede its first call")
        if [case.case_id for case in self.cases] != list(_EXPECTED_CASE_IDS):
            raise ValueError("grading results must match the canonical 20-case order")
        if len({case.answer_region_id for case in self.cases}) != OCR_CALL_LIMIT:
            raise ValueError("grading results require 20 unique answer regions")
        blank_ids = {
            case.case_id
            for case in self.cases
            if case.outcome == "not_called_blank_safety_gate"
        }
        if blank_ids != _BLANK_CASE_IDS:
            raise ValueError("only A4 and C4 may be blank safety refusals")
        suggested = [case for case in self.cases if case.outcome == "suggested"]
        if len(suggested) != QWEN_CALL_LIMIT:
            raise ValueError("grading results require exactly 18 suggestions")
        for field_name in (
            "dispatch_item_id",
            "grading_job_id",
            "grade_suggestion_id",
        ):
            values = [getattr(case, field_name) for case in suggested]
            if len(set(values)) != QWEN_CALL_LIMIT:
                raise ValueError(f"grading results require unique {field_name} values")
        if len(set(self.dispatch_run_ids)) != 5:
            raise ValueError("grading results require five unique dispatch runs")
        return self


class GradingReviewCase(BaseModel):
    model_config = ConfigDict(extra="forbid")

    case_id: str
    disagreement_reason: Literal[
        "none", "rubric_ambiguity", "ocr_unfixed", "model_error", "dataset_error"
    ]
    teacher_notes: str = ""
    useful_draft: bool
    approved_review: Literal[True]


class ReviewLock(BaseModel):
    model_config = ConfigDict(extra="forbid")

    run_id: str
    reviewer_id: str = Field(min_length=1, max_length=255)
    signed_at: datetime
    grading_results_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    workbook_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    cases: list[GradingReviewCase] = Field(min_length=20, max_length=20)

    @model_validator(mode="after")
    def validate_cases(self) -> ReviewLock:
        if [case.case_id for case in self.cases] != list(_EXPECTED_CASE_IDS):
            raise ValueError("teacher reviews must match the canonical 20-case order")
        return self


class LedgerEntry(BaseModel):
    model_config = ConfigDict(extra="forbid")

    sequence: int = Field(ge=1)
    state: Literal[
        "prepared",
        "ground_truth_locked",
        "ocr_completed",
        "ocr_confirmed",
        "grading_completed",
        "review_completed",
        "ocr_quality_no_go",
        "reported",
        "invalid",
    ]
    occurred_at: datetime
    previous_entry_sha256: str | None
    locked_artifacts: dict[str, str]
    metadata: dict[str, Any] = Field(default_factory=dict)
    entry_sha256: str


def _criterion(
    criterion_id: str,
    name: str,
    description: str,
    marks: str,
    *,
    depends_on: tuple[str, ...] = (),
    allow_follow_through: bool = False,
) -> RubricCriterion:
    return RubricCriterion(
        id=criterion_id,
        name=name,
        description=description,
        max_marks=Decimal(marks),
        depends_on=list(depends_on),
        allow_follow_through=allow_follow_through,
    )


def build_case_blueprint() -> list[LocalCuratedCaseDefinition]:
    algebra_rubric = [
        _criterion("a_derive", "Derive 2x=8", "Subtracts 3 correctly.", "1"),
        _criterion(
            "a_divide",
            "Divide by 2",
            "Divides both sides by 2.",
            "1",
            depends_on=("a_derive",),
        ),
        _criterion(
            "a_final",
            "Final answer",
            "States x=4.",
            "1",
            depends_on=("a_divide",),
        ),
    ]
    biology_rubric = [
        _criterion(
            "b_shape",
            "Biconcave shape",
            "Names the biconcave shape and explains increased surface area for diffusion.",
            "2",
        ),
        _criterion(
            "b_nucleus",
            "No nucleus",
            "Explains that no nucleus leaves more room for haemoglobin.",
            "2",
        ),
    ]
    kinetic_rubric = [
        _criterion("c_formula", "Formula", "Uses KE = 1/2 mv^2.", "2"),
        _criterion(
            "c_substitution",
            "Substitution",
            "Substitutes m=2 and v=3.",
            "1",
            depends_on=("c_formula",),
        ),
        _criterion(
            "c_arithmetic",
            "Arithmetic",
            "Calculates 9.",
            "1",
            depends_on=("c_substitution",),
        ),
        _criterion("c_unit", "Unit", "Uses joules (J).", "1"),
    ]
    bayes_rubric = [
        _criterion("d_setup", "Bayes setup", "Sets up the posterior ratio correctly.", "2"),
        _criterion("d_joint_d", "Disease-positive joint", "Calculates 0.10 x 0.80 = 0.08.", "1"),
        _criterion(
            "d_joint_not_d",
            "Non-disease-positive joint",
            "Calculates 0.90 x 0.20 = 0.18.",
            "1",
        ),
        _criterion(
            "d_denominator",
            "Denominator",
            "Calculates 0.08 + 0.18 = 0.26.",
            "1",
            depends_on=("d_joint_d", "d_joint_not_d"),
        ),
        _criterion(
            "d_final",
            "Posterior",
            "Obtains 4/13 or approximately 0.3077.",
            "1",
            depends_on=("d_denominator",),
        ),
    ]
    height_rubric = [
        _criterion("e_relation", "Valid relation", "Uses a valid kinematic relation.", "2"),
        _criterion(
            "e_sign",
            "Signs and substitution",
            "Uses u=20 and a=-10 correctly.",
            "1",
            depends_on=("e_relation",),
        ),
        _criterion(
            "e_magnitude",
            "Magnitude",
            "Obtains a maximum height of 20.",
            "1",
            depends_on=("e_sign",),
        ),
        _criterion("e_unit", "Unit", "Uses metres (m).", "1"),
    ]

    packs: dict[str, dict[str, Any]] = {
        "A": {
            "question": "Solve 2x + 3 = 11. Show each algebraic step.",
            "model": "2x + 3 = 11; 2x = 8; x = 4.",
            "rubric": algebra_rubric,
            "max": Decimal("3"),
        },
        "B": {
            "question": (
                "State and explain two structural adaptations of a red blood cell for "
                "oxygen transport."
            ),
            "model": (
                "Its biconcave shape gives a large surface area for diffusion, and the lack "
                "of a nucleus leaves more room for haemoglobin."
            ),
            "rubric": biology_rubric,
            "max": Decimal("4"),
        },
        "C": {
            "question": "Calculate the kinetic energy of a 2 kg object moving at 3 m/s.",
            "model": "KE = 1/2 mv^2 = 1/2 x 2 x 3^2 = 9 J.",
            "rubric": kinetic_rubric,
            "max": Decimal("5"),
        },
        "D": {
            "question": (
                "A disease has prevalence 0.10. A test has sensitivity 0.80 and a false "
                "positive rate of 0.20. Find P(disease | positive)."
            ),
            "model": (
                "P(D and +)=0.10 x 0.80=0.08; P(not D and +)=0.90 x 0.20=0.18; "
                "P(D|+)=0.08/(0.08+0.18)=4/13 approximately 0.3077."
            ),
            "rubric": bayes_rubric,
            "max": Decimal("6"),
        },
        "E": {
            "question": (
                "A ball is thrown vertically upward at 20 m/s. Take g=10 m/s^2. Find its "
                "maximum height above the launch point."
            ),
            "model": "0 = 20^2 + 2(-10)h, so h = 20 m.",
            "rubric": height_rubric,
            "max": Decimal("5"),
        },
    }

    specs: list[tuple[Any, ...]] = [
        (
            "A1",
            PrimaryCategory.CORRECT,
            AnswerQuality.CORRECT,
            "3",
            "2x + 3 = 11\n2x = 8\nx = 4",
            "typed",
            True,
            [],
            "Complete algebraic solution.",
        ),
        (
            "A2",
            PrimaryCategory.PARTIAL,
            AnswerQuality.PARTIAL,
            "1",
            "2x + 3 = 11\n2x = 8",
            "typed",
            True,
            [],
            "Stops after the first valid transformation.",
        ),
        (
            "A3",
            PrimaryCategory.WRONG,
            AnswerQuality.WRONG,
            "0",
            "2x + 3 = 11\n2x = 14\nx = 7",
            "typed",
            True,
            [],
            "Incorrect subtraction and result.",
        ),
        (
            "A4",
            PrimaryCategory.BLANK,
            AnswerQuality.BLANK,
            "0",
            "",
            "blank",
            False,
            [],
            "True blank; Qwen must not be called.",
        ),
        (
            "B1",
            PrimaryCategory.CORRECT,
            AnswerQuality.CORRECT,
            "4",
            (
                "The biconcave shape gives a large surface area for diffusion.\n"
                "No nucleus leaves more space for haemoglobin."
            ),
            "typed",
            True,
            [],
            "Both adaptations are named and explained.",
        ),
        (
            "B2",
            PrimaryCategory.PARTIAL,
            AnswerQuality.PARTIAL,
            "2",
            "The biconcave shape gives a large surface area for diffusion.",
            "typed",
            True,
            [],
            "Only one adaptation is supplied.",
        ),
        (
            "B3",
            PrimaryCategory.IRRELEVANT,
            AnswerQuality.IRRELEVANT,
            "0",
            "Photosynthesis uses sunlight to make glucose in plant leaves.",
            "typed",
            True,
            [],
            "Coherent but unrelated answer.",
        ),
        (
            "B4",
            PrimaryCategory.DIFFICULT_HANDWRITING,
            AnswerQuality.CORRECT,
            "4",
            (
                "Biconcave shape gives more surface area for diffusion.\n"
                "No nucleus means more room for haemoglobin."
            ),
            "handwriting",
            False,
            [],
            "Correct conceptual answer rendered as difficult handwriting.",
        ),
        (
            "C1",
            PrimaryCategory.WRONG,
            AnswerQuality.WRONG,
            "1",
            "KE = mv = 2 x 3 = 6 J",
            "typed",
            True,
            [],
            "Wrong formula and result; unit alone earns one mark.",
        ),
        (
            "C2",
            PrimaryCategory.FORMULA_HEAVY,
            AnswerQuality.CORRECT,
            "5",
            "KE = 1/2 mv^2\n= 1/2 x 2 x 3^2\n= 9 J",
            "formula",
            True,
            ["KE", "1/2", "m", "v^2", "2", "3^2", "9", "J"],
            "Complete formula-heavy response.",
        ),
        (
            "C3",
            PrimaryCategory.DIFFICULT_HANDWRITING,
            AnswerQuality.PARTIAL,
            "4",
            "KE = 1/2 mv^2\n= 1/2 x 2 x 3^2\n= 9",
            "handwriting",
            False,
            [],
            "Correct handwritten calculation without the unit.",
        ),
        (
            "C4",
            PrimaryCategory.BLANK,
            AnswerQuality.BLANK,
            "0",
            "",
            "blank",
            False,
            [],
            "True ruled blank; Qwen must not be called.",
        ),
        (
            "D1",
            PrimaryCategory.CORRECT,
            AnswerQuality.CORRECT,
            "6",
            (
                "P(D and +) = 0.10 x 0.80 = 0.08\n"
                "P(not D and +) = 0.90 x 0.20 = 0.18\n"
                "P(D|+) = 0.08 / (0.08 + 0.18) = 4/13 = 0.3077"
            ),
            "typed",
            True,
            [],
            "Complete Bayes calculation.",
        ),
        (
            "D2",
            PrimaryCategory.WRONG,
            AnswerQuality.WRONG,
            "0",
            "P(D|+) = P(+|D) = 0.80",
            "typed",
            True,
            [],
            "Confuses posterior probability with sensitivity.",
        ),
        (
            "D3",
            PrimaryCategory.MULTI_STEP,
            AnswerQuality.PARTIAL,
            "5",
            (
                "P(D and +) = 0.10 x 0.80 = 0.08\n"
                "P(not D and +) = 0.90 x 0.20 = 0.18\n"
                "P(+) = 0.08 + 0.18 = 0.26"
            ),
            "multi_step",
            True,
            [],
            "Several correct steps but no posterior division.",
        ),
        (
            "D4",
            PrimaryCategory.IRRELEVANT,
            AnswerQuality.IRRELEVANT,
            "0",
            "The mean is the total divided by the count; the median is the middle value.",
            "typed",
            True,
            [],
            "Coherent mathematics unrelated to Bayes' theorem.",
        ),
        (
            "E1",
            PrimaryCategory.PARTIAL,
            AnswerQuality.PARTIAL,
            "3",
            "v^2 = u^2 + 2as\n0 = 20^2 + 2(-10)h",
            "typed",
            True,
            [],
            "Correct relation and substitution without solving.",
        ),
        (
            "E2",
            PrimaryCategory.DIFFICULT_HANDWRITING,
            AnswerQuality.WRONG,
            "2",
            "0 = 20^2 + 2(10)h\nh = -20",
            "handwriting",
            False,
            [],
            "Valid relation but wrong acceleration sign.",
        ),
        (
            "E3",
            PrimaryCategory.FORMULA_HEAVY,
            AnswerQuality.CORRECT,
            "5",
            "h = u^2 / (2g)\n= 20^2 / (2 x 10)\n= 20 m",
            "formula",
            True,
            ["h", "u^2", "2g", "20^2", "2", "10", "20", "m"],
            "Terse but complete formula-heavy solution.",
        ),
        (
            "E4",
            PrimaryCategory.MULTI_STEP,
            AnswerQuality.CORRECT,
            "5",
            "t = u/g = 20/10 = 2 s\nh = ut - 1/2 gt^2\n= 20 x 2 - 1/2 x 10 x 2^2 = 20 m",
            "multi_step",
            True,
            [],
            "Correct alternate multi-step solution.",
        ),
    ]

    cases: list[LocalCuratedCaseDefinition] = []
    for (
        case_id,
        primary_category,
        answer_quality,
        expected_score,
        answer,
        render_profile,
        clean_typed,
        critical_tokens,
        notes,
    ) in specs:
        pack = packs[case_id[0]]
        cases.append(
            LocalCuratedCaseDefinition(
                case_id=case_id,
                pack_id=case_id[0],
                primary_category=primary_category,
                answer_quality=answer_quality,
                question_text=pack["question"],
                model_answer=pack["model"],
                rubric=pack["rubric"],
                expected_score=Decimal(expected_score),
                max_score=pack["max"],
                authored_transcription=answer,
                teacher_notes=notes,
                render_profile=render_profile,
                clean_typed=clean_typed,
                critical_tokens=critical_tokens,
            )
        )
    validate_case_blueprint(cases)
    return cases


def validate_case_blueprint(cases: list[LocalCuratedCaseDefinition]) -> None:
    if len(cases) != 20:
        raise LocalCuratedEvaluationError("The curated blueprint must contain exactly 20 cases")
    ids = [case.case_id for case in cases]
    expected_ids = [f"{pack}{index}" for pack in "ABCDE" for index in range(1, 5)]
    if ids != expected_ids or len(set(ids)) != len(ids):
        raise LocalCuratedEvaluationError("Case IDs must be unique and ordered A1 through E4")
    expected_categories = {
        PrimaryCategory.CORRECT: 3,
        PrimaryCategory.PARTIAL: 3,
        PrimaryCategory.WRONG: 3,
        PrimaryCategory.BLANK: 2,
        PrimaryCategory.IRRELEVANT: 2,
        PrimaryCategory.DIFFICULT_HANDWRITING: 3,
        PrimaryCategory.FORMULA_HEAVY: 2,
        PrimaryCategory.MULTI_STEP: 2,
    }
    actual_categories = Counter(case.primary_category for case in cases)
    if actual_categories != expected_categories:
        raise LocalCuratedEvaluationError(
            f"Primary category allocation is invalid: {dict(actual_categories)}"
        )
    expected_quality = {
        AnswerQuality.CORRECT: 7,
        AnswerQuality.PARTIAL: 5,
        AnswerQuality.WRONG: 4,
        AnswerQuality.BLANK: 2,
        AnswerQuality.IRRELEVANT: 2,
    }
    actual_quality = Counter(case.answer_quality for case in cases)
    if actual_quality != expected_quality:
        raise LocalCuratedEvaluationError(
            f"Answer-quality allocation is invalid: {dict(actual_quality)}"
        )
    pack_counts = Counter(case.pack_id for case in cases)
    if pack_counts != Counter({pack: 4 for pack in "ABCDE"}):
        raise LocalCuratedEvaluationError("Each question pack must contain exactly four cases")


def normalize_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value or "")
    normalized = normalized.replace("−", "-").replace("–", "-").replace("—", "-")
    normalized = normalized.replace("×", "x").replace("·", "x")
    return " ".join(normalized.split())


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_text(value: str) -> str:
    return sha256_bytes(value.encode("utf-8"))


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file_obj:
        for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _jsonable(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, datetime):
        return value.astimezone(UTC).isoformat()
    if isinstance(value, Path):
        return value.as_posix()
    if isinstance(value, BaseModel):
        return _jsonable(value.model_dump(mode="python"))
    if isinstance(value, Enum):
        return value.value
    if isinstance(value, dict):
        return {str(key): _jsonable(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_jsonable(item) for item in value]
    return value


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(_jsonable(value), indent=2, sort_keys=True, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )


def read_json(path: Path) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise LocalCuratedEvaluationError(
            f"Could not read evaluation artifact {path.name}"
        ) from exc


def _canonical_json(value: Any) -> str:
    return json.dumps(
        _jsonable(value),
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=False,
    )


def _repository_root() -> Path:
    return Path(__file__).resolve().parents[4]


def default_evaluation_root() -> Path:
    return _repository_root() / "data" / "evaluation"


def current_git_commit() -> str:
    result = subprocess.run(
        ["git", "rev-parse", "HEAD"],
        cwd=_repository_root(),
        check=True,
        capture_output=True,
        text=True,
        timeout=10,
    )
    return result.stdout.strip()


def require_clean_git_worktree() -> None:
    result = subprocess.run(
        ["git", "status", "--porcelain", "--untracked-files=normal"],
        cwd=_repository_root(),
        check=True,
        capture_output=True,
        text=True,
        timeout=10,
    )
    tracked_output = [
        line
        for line in result.stdout.splitlines()
        if not line.endswith(" data/evaluation") and "data/evaluation/" not in line
    ]
    if tracked_output:
        raise LocalCuratedEvaluationError("The evaluation requires a clean Git worktree")


def require_commit_lineage(*, integration_commit: str, harness_commit: str) -> None:
    if current_git_commit() != harness_commit:
        raise LocalCuratedEvaluationError("Harness commit must equal the current Git HEAD")
    for label, commit in (
        ("integration", integration_commit),
        ("harness", harness_commit),
    ):
        result = subprocess.run(
            ["git", "cat-file", "-e", f"{commit}^{{commit}}"],
            cwd=_repository_root(),
            check=False,
            capture_output=True,
            text=True,
            timeout=10,
        )
        if result.returncode != 0:
            raise LocalCuratedEvaluationError(f"Recorded {label} commit is not available")
    lineage = subprocess.run(
        ["git", "merge-base", "--is-ancestor", integration_commit, harness_commit],
        cwd=_repository_root(),
        check=False,
        capture_output=True,
        text=True,
        timeout=10,
    )
    if lineage.returncode != 0:
        raise LocalCuratedEvaluationError(
            "The reviewed integration commit must be an ancestor of the harness commit"
        )


def validate_database_url(database_url: str, run_id: str) -> str:
    if not database_url:
        raise LocalCuratedEvaluationError("DATABASE_URL is required")
    try:
        database_name = make_url(database_url).database or ""
    except Exception as exc:
        raise LocalCuratedEvaluationError("DATABASE_URL is invalid") from exc
    expected = f"teacher_assistant_eval_{run_id}"
    if database_name != expected:
        raise LocalCuratedEvaluationError(
            f"Evaluation database must be named exactly {expected}"
        )
    return database_name


def _font(candidates: list[str], size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    windows_fonts = Path(os.environ.get("WINDIR", "C:/Windows")) / "Fonts"
    for candidate in candidates:
        path = windows_fonts / candidate
        if path.is_file():
            return ImageFont.truetype(str(path), size)
    for candidate in (
        "DejaVuSans.ttf",
        "LiberationSans-Regular.ttf",
    ):
        try:
            return ImageFont.truetype(candidate, size)
        except OSError:
            continue
    return ImageFont.load_default()


def _wrap_render_line(
    draw: ImageDraw.ImageDraw,
    line: str,
    font: ImageFont.FreeTypeFont | ImageFont.ImageFont,
    *,
    max_width: int,
) -> list[str]:
    words = line.split()
    if not words:
        return [""]
    rendered: list[str] = []
    current = words[0]
    for word in words[1:]:
        candidate = f"{current} {word}"
        left, _top, right, _bottom = draw.textbbox((0, 0), candidate, font=font)
        if right - left <= max_width:
            current = candidate
        else:
            rendered.append(current)
            current = word
    rendered.append(current)
    return rendered


def _render_case_image(case: LocalCuratedCaseDefinition, path: Path, seed: int) -> None:
    import random

    rng = random.Random(seed + sum(ord(character) for character in case.case_id))
    width, height = 1600, 900
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    for y in range(120, height - 60, 90):
        draw.line((60, y, width - 60, y), fill=(226, 235, 244), width=2)
    if case.render_profile == "blank":
        image.save(path, format="PNG", optimize=False)
        return

    if case.render_profile == "handwriting":
        handwriting_fonts = (
            ["Inkfree.ttf", "segoepr.ttf", "comic.ttf"]
            if case.case_id == "C3"
            else ["segoepr.ttf", "Inkfree.ttf", "comic.ttf"]
        )
        font = _font(handwriting_fonts, 62)
        fill = (35, 48, 70)
    elif case.render_profile == "formula":
        font = _font(["cambria.ttc", "cambria.ttf", "consola.ttf", "arial.ttf"], 68)
        fill = (12, 20, 32)
    else:
        font = _font(["arial.ttf", "calibri.ttf", "cambria.ttf"], 58)
        fill = (18, 24, 34)

    lines: list[str] = []
    for source_line in case.authored_transcription.splitlines() or [case.authored_transcription]:
        lines.extend(_wrap_render_line(draw, source_line, font, max_width=1380))
    y = 105
    for line_number, line in enumerate(lines):
        if case.render_profile == "handwriting":
            layer = Image.new("RGBA", (width - 100, 120), (255, 255, 255, 0))
            layer_draw = ImageDraw.Draw(layer)
            layer_draw.text((20, 15), line, font=font, fill=fill + (255,))
            angle = rng.uniform(-1.7, 1.7)
            layer = layer.rotate(angle, resample=Image.Resampling.BICUBIC, expand=False)
            x = 72 + rng.randint(-14, 18)
            image.paste(layer, (x, y + rng.randint(-8, 8)), layer)
            y += 112 + rng.randint(-5, 10)
        else:
            draw.text((90, y), line, font=font, fill=fill)
            y += 112 if case.render_profile == "formula" else 100
        if line_number > 7:
            break

    if case.render_profile == "handwriting":
        noise = Image.new("L", image.size, 0)
        noise_draw = ImageDraw.Draw(noise)
        for _ in range(220):
            x = rng.randrange(65, width - 65)
            y_noise = rng.randrange(65, height - 65)
            shade = rng.randrange(4, 15)
            noise_draw.ellipse((x, y_noise, x + 1, y_noise + 1), fill=shade)
        noise = noise.filter(ImageFilter.GaussianBlur(radius=0.3))
        paper = Image.new("RGB", image.size, (244, 244, 244))
        image = Image.composite(paper, image, noise)
    image.save(path, format="PNG", optimize=False)


def _rubric_summary(case: LocalCuratedCaseDefinition) -> str:
    return "\n".join(
        f"{criterion.name} ({criterion.max_marks}): {criterion.description}"
        for criterion in case.rubric
    )


def _style_workbook_header(sheet: Any, row: int = 1) -> None:
    for cell in sheet[row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def create_ground_truth_workbook(
    manifest: LocalCuratedEvaluationManifest,
    run_dir: Path,
) -> Path:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instructions.append(["20-case local curated evaluation"])
    instructions.append(["Run ID", manifest.run_id])
    instructions.append(["Protocol", manifest.protocol_version])
    instructions.append(["Gate", "Teacher marks and transcribes before any AI call"])
    instructions.append(
        [
            "Required",
            (
                "Complete every teacher field on the Cases sheet. Mark all cases approved. "
                "For B4, C3, and E2, also confirm that the handwriting is legible but "
                "genuinely difficult."
            ),
        ]
    )
    instructions.column_dimensions["A"].width = 22
    instructions.column_dimensions["B"].width = 110
    instructions["A1"].font = Font(bold=True, size=16)

    cases_sheet = workbook.create_sheet("Cases")
    headers = [
        "case_id",
        "primary_category",
        "answer_quality",
        "question",
        "rubric",
        "teacher_transcription",
        "teacher_score",
        "teacher_notes",
        "handwriting_acceptable",
        "teacher_approved",
        "answer_image",
    ]
    cases_sheet.append(headers)
    _style_workbook_header(cases_sheet)
    widths = [12, 24, 18, 55, 70, 55, 16, 40, 25, 20, 72]
    for index, width in enumerate(widths, start=1):
        cases_sheet.column_dimensions[get_column_letter(index)].width = width

    for row_index, case in enumerate(manifest.cases, start=2):
        cases_sheet.append(
            [
                case.case_id,
                case.primary_category.value,
                case.answer_quality.value,
                case.question_text,
                _rubric_summary(case),
                None,
                None,
                None,
                None,
                None,
                None,
            ]
        )
        cases_sheet.row_dimensions[row_index].height = 250
        for cell in cases_sheet[row_index]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        image_path = run_dir / case.image_relative_path
        workbook_image = WorkbookImage(str(image_path))
        workbook_image.width = 560
        workbook_image.height = 315
        cases_sheet.add_image(workbook_image, f"K{row_index}")
    cases_sheet.freeze_panes = "A2"

    path = run_dir / "ground_truth_review.xlsx"
    workbook.save(path)
    return path


def _parse_bool(value: Any, *, field: str, case_id: str) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, int) and value in {0, 1}:
        return bool(value)
    normalized = str(value or "").strip().lower()
    if normalized in {"true", "yes", "y", "approved", "1"}:
        return True
    if normalized in {"false", "no", "n", "0"}:
        return False
    raise LocalCuratedEvaluationError(f"{case_id}: {field} must be an explicit yes or no")


def _parse_decimal(value: Any, *, field: str, case_id: str) -> Decimal:
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise LocalCuratedEvaluationError(f"{case_id}: {field} must be a number") from exc


def read_ground_truth_workbook(
    workbook_path: Path,
    manifest: LocalCuratedEvaluationManifest,
) -> list[GroundTruthCase]:
    workbook = load_workbook(workbook_path, data_only=True)
    if "Cases" not in workbook.sheetnames:
        raise LocalCuratedEvaluationError("Ground-truth workbook has no Cases sheet")
    sheet = workbook["Cases"]
    headers = {str(cell.value): index for index, cell in enumerate(sheet[1], start=1)}
    required = {
        "case_id",
        "teacher_transcription",
        "teacher_score",
        "teacher_notes",
        "handwriting_acceptable",
        "teacher_approved",
    }
    if not required.issubset(headers):
        raise LocalCuratedEvaluationError("Ground-truth workbook columns were changed")
    cases_by_id = {case.case_id: case for case in manifest.cases}
    signed_cases: list[GroundTruthCase] = []
    for row in range(2, sheet.max_row + 1):
        case_id = str(sheet.cell(row, headers["case_id"]).value or "").strip()
        if not case_id:
            continue
        definition = cases_by_id.get(case_id)
        if definition is None:
            raise LocalCuratedEvaluationError(f"Unknown case ID in workbook: {case_id}")
        transcription_value = sheet.cell(row, headers["teacher_transcription"]).value
        transcription = str(transcription_value or "").strip()
        score = _parse_decimal(
            sheet.cell(row, headers["teacher_score"]).value,
            field="teacher_score",
            case_id=case_id,
        )
        approved = _parse_bool(
            sheet.cell(row, headers["teacher_approved"]).value,
            field="teacher_approved",
            case_id=case_id,
        )
        if not approved:
            raise LocalCuratedEvaluationError(f"{case_id}: teacher approval is required")
        handwriting_value = sheet.cell(row, headers["handwriting_acceptable"]).value
        handwriting_acceptable: bool | None = None
        if definition.primary_category == PrimaryCategory.DIFFICULT_HANDWRITING:
            handwriting_acceptable = _parse_bool(
                handwriting_value,
                field="handwriting_acceptable",
                case_id=case_id,
            )
            if not handwriting_acceptable:
                raise LocalCuratedEvaluationError(
                    f"{case_id}: difficult handwriting must be teacher-approved before OCR"
                )
        if normalize_text(transcription) != normalize_text(definition.authored_transcription):
            raise LocalCuratedEvaluationError(
                f"{case_id}: teacher transcription does not match the authored synthetic answer"
            )
        if score != definition.expected_score:
            raise LocalCuratedEvaluationError(
                f"{case_id}: teacher score does not match the authored rubric score"
            )
        notes = str(sheet.cell(row, headers["teacher_notes"]).value or "").strip()
        signed_cases.append(
            GroundTruthCase(
                case_id=case_id,
                teacher_transcription=transcription,
                teacher_score=score,
                teacher_notes=notes,
                handwriting_acceptable=handwriting_acceptable,
                approved=True,
            )
        )
    if [case.case_id for case in signed_cases] != [case.case_id for case in manifest.cases]:
        raise LocalCuratedEvaluationError("All 20 cases must be completed in blueprint order")
    return signed_cases


def _entry_hash_payload(entry: dict[str, Any]) -> str:
    payload = {key: value for key, value in entry.items() if key != "entry_sha256"}
    return sha256_text(_canonical_json(payload))


def read_ledger(run_dir: Path) -> list[LedgerEntry]:
    path = run_dir / "ledger.jsonl"
    if not path.is_file():
        return []
    entries: list[LedgerEntry] = []
    previous: str | None = None
    for line_number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        if not line.strip():
            continue
        try:
            payload = json.loads(line)
            entry = LedgerEntry.model_validate(payload)
        except Exception as exc:
            raise LocalCuratedEvaluationError(
                f"Evaluation ledger is invalid at line {line_number}"
            ) from exc
        if entry.sequence != len(entries) + 1:
            raise LocalCuratedEvaluationError("Evaluation ledger sequence is invalid")
        if entry.previous_entry_sha256 != previous:
            raise LocalCuratedEvaluationError("Evaluation ledger hash chain is invalid")
        if _entry_hash_payload(payload) != entry.entry_sha256:
            raise LocalCuratedEvaluationError("Evaluation ledger entry hash is invalid")
        entries.append(entry)
        previous = entry.entry_sha256
    return entries


def verify_locked_artifacts(run_dir: Path, entries: list[LedgerEntry] | None = None) -> None:
    ledger = entries if entries is not None else read_ledger(run_dir)
    for entry in ledger:
        for relative_path, expected_hash in entry.locked_artifacts.items():
            path = run_dir / relative_path
            if not path.is_file() or sha256_file(path) != expected_hash:
                raise LocalCuratedEvaluationError(
                    f"Locked evaluation artifact changed or disappeared: {relative_path}"
                )


def current_state(run_dir: Path) -> str | None:
    entries = read_ledger(run_dir)
    return entries[-1].state if entries else None


def append_state(
    run_dir: Path,
    state: str,
    *,
    locked_artifacts: dict[str, str],
    metadata: dict[str, Any] | None = None,
) -> LedgerEntry:
    entries = read_ledger(run_dir)
    if state != "invalid":
        verify_locked_artifacts(run_dir, entries)
    previous_state = entries[-1].state if entries else None
    if previous_state in {"invalid", "reported"}:
        raise LocalCuratedEvaluationError("The evaluation is already terminal")
    if state == "invalid":
        pass
    elif state == "ocr_quality_no_go":
        if previous_state != "ocr_completed":
            raise LocalCuratedEvaluationError(
                "An OCR quality no-go may only follow completed visual transcription"
            )
    elif state == "reported" and previous_state == "ocr_quality_no_go":
        # A signed teacher rejection is a valid quality outcome, not an integrity
        # failure.  It reports directly and never opens a grading path.
        pass
    else:
        expected_index = 0 if previous_state is None else _STATE_ORDER.index(previous_state) + 1
        if expected_index >= len(_STATE_ORDER) or _STATE_ORDER[expected_index] != state:
            raise LocalCuratedEvaluationError(
                f"Cannot transition evaluation from {previous_state or 'new'} to {state}"
            )
    payload: dict[str, Any] = {
        "sequence": len(entries) + 1,
        "state": state,
        "occurred_at": datetime.now(UTC),
        "previous_entry_sha256": entries[-1].entry_sha256 if entries else None,
        "locked_artifacts": locked_artifacts,
        "metadata": metadata or {},
    }
    payload["entry_sha256"] = _entry_hash_payload(_jsonable(payload))
    entry = LedgerEntry.model_validate(payload)
    ledger_path = run_dir / "ledger.jsonl"
    with ledger_path.open("a", encoding="utf-8", newline="\n") as file_obj:
        file_obj.write(_canonical_json(entry) + "\n")
    return entry


def load_manifest(run_dir: Path) -> LocalCuratedEvaluationManifest:
    return LocalCuratedEvaluationManifest.model_validate(read_json(run_dir / "manifest.json"))


def prepare_evaluation(
    *,
    run_id: str,
    output_root: Path,
    integration_commit: str,
    harness_commit: str,
    operator_assets: OperatorAssetMetadata,
    seed: int = DEFAULT_SEED,
) -> Path:
    if not _RUN_ID_PATTERN.fullmatch(run_id):
        raise LocalCuratedEvaluationError("Invalid run ID")
    if not re.fullmatch(r"[0-9a-f]{40}", integration_commit):
        raise LocalCuratedEvaluationError("integration_commit must be a full Git hash")
    if not re.fullmatch(r"[0-9a-f]{40}", harness_commit):
        raise LocalCuratedEvaluationError("harness_commit must be a full Git hash")
    root = output_root.resolve()
    run_dir = (root / run_id).resolve()
    if run_dir.parent != root:
        raise LocalCuratedEvaluationError("Evaluation run path escaped the output root")
    if run_dir.exists():
        raise LocalCuratedEvaluationError("Evaluation run directory already exists")
    images_dir = run_dir / "images"
    images_dir.mkdir(parents=True)
    cases = build_case_blueprint()
    rendered_cases: list[LocalCuratedCaseDefinition] = []
    for case in cases:
        relative_path = f"images/{case.case_id}.png"
        image_path = run_dir / relative_path
        _render_case_image(case, image_path, seed)
        rendered_cases.append(
            case.model_copy(
                update={
                    "image_relative_path": relative_path,
                    "image_sha256": sha256_file(image_path),
                }
            )
        )
    manifest = LocalCuratedEvaluationManifest(
        run_id=run_id,
        seed=seed,
        created_at=datetime.now(UTC),
        integration_commit=integration_commit,
        harness_commit=harness_commit,
        # The grading model is whichever alias the operator's pinned assets name,
        # so a run records the model it actually used rather than a fixed default.
        expected_qwen_model=operator_assets.llama_cpp.model_alias,
        operator_assets=operator_assets,
        cases=rendered_cases,
    )
    manifest_path = run_dir / "manifest.json"
    write_json(manifest_path, manifest)
    create_ground_truth_workbook(manifest, run_dir)
    locked = {"manifest.json": sha256_file(manifest_path)}
    locked.update(
        {
            case.image_relative_path: case.image_sha256
            for case in manifest.cases
        }
    )
    append_state(
        run_dir,
        "prepared",
        locked_artifacts=locked,
        metadata={
            "case_count": 20,
            "model_calls": 0,
            "teacher_signoff_required": True,
        },
    )
    return run_dir


def lock_ground_truth(
    run_dir: Path,
    *,
    reviewer_id: str,
    confirm_teacher_signoff: bool,
) -> GroundTruthLock:
    if not confirm_teacher_signoff:
        raise LocalCuratedEvaluationError("Explicit teacher sign-off confirmation is required")
    if current_state(run_dir) != "prepared":
        raise LocalCuratedEvaluationError("Ground truth can only be locked after prepare")
    verify_locked_artifacts(run_dir)
    manifest_path = run_dir / "manifest.json"
    workbook_path = run_dir / "ground_truth_review.xlsx"
    manifest = load_manifest(run_dir)
    cases = read_ground_truth_workbook(workbook_path, manifest)
    lock = GroundTruthLock(
        run_id=manifest.run_id,
        reviewer_id=reviewer_id.strip(),
        signed_at=datetime.now(UTC),
        manifest_sha256=sha256_file(manifest_path),
        workbook_sha256=sha256_file(workbook_path),
        cases=cases,
    )
    lock_path = run_dir / "ground_truth_lock.json"
    write_json(lock_path, lock)
    append_state(
        run_dir,
        "ground_truth_locked",
        locked_artifacts={
            "ground_truth_review.xlsx": sha256_file(workbook_path),
            "ground_truth_lock.json": sha256_file(lock_path),
        },
        metadata={
            "reviewer_id_sha256": sha256_text(reviewer_id.strip()),
            "signed_case_count": len(cases),
            "model_calls": 0,
        },
    )
    return lock


def _load_local_ai_environment(path: Path | None) -> None:
    if path is None:
        path = _repository_root() / ".env.local-ai"
    if not path.is_file():
        raise LocalCuratedEvaluationError(
            "Ignored local-AI environment configuration is required for a real stage"
        )
    # The ignored operator file is intentionally limited to the local-AI
    # namespace, but it must include Qwen3.8 and the phase/reference controls
    # as well as the legacy Qwen3.6 keys.  `LOCAL_QWEN_` does not match
    # `LOCAL_QWEN38_`.
    allowed_prefixes = ("BRAIN_", "LOCAL_", "COHORT_")
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        if key.startswith(allowed_prefixes):
            os.environ[key] = value


def _configure_runtime(
    run_dir: Path,
    *,
    database_url: str,
    local_ai_env: Path | None,
    require_grading: bool = False,
) -> tuple[Any, Any, Any]:
    manifest = load_manifest(run_dir)
    validate_database_url(database_url, manifest.run_id)
    _load_local_ai_environment(local_ai_env)
    storage_root = (run_dir / "runtime_storage").resolve()
    os.environ["DATABASE_URL"] = database_url
    os.environ["LOCAL_STORAGE_ROOT"] = str(storage_root)
    os.environ["UPLOADS_DIR"] = str(storage_root / "uploads")
    os.environ["ARTIFACTS_DIR"] = str(storage_root / "artifacts")

    from app.core.config import get_settings

    get_settings.cache_clear()
    settings = get_settings()
    if not settings.brain_allow_real_providers:
        raise LocalCuratedEvaluationError("BRAIN_ALLOW_REAL_PROVIDERS must be true")
    if (
        not settings.local_qwen38_enabled
        or not settings.local_qwen38_visual_preparation_enabled
        or settings.local_qwen38_model != manifest.operator_assets.qwen38_vision.model_alias
    ):
        raise LocalCuratedEvaluationError(
            "Qwen3.8 non-thinking visual transcription is not enabled with the pinned model"
        )
    if require_grading:
        if not settings.cohort_model_grading_enabled:
            raise LocalCuratedEvaluationError("COHORT_MODEL_GRADING_ENABLED must be true")
        if settings.cohort_provider_retry_count != 0:
            raise LocalCuratedEvaluationError("Provider retry count must be zero")
        if settings.cohort_max_provider_calls != 25:
            raise LocalCuratedEvaluationError(
                "The server cohort-call ceiling must be exactly 25"
        )
        if manifest.expected_qwen_model == "qwen3.6-35b-a3b-q4km":
            if (
                not settings.local_qwen_enabled
                or settings.local_qwen_model != manifest.expected_qwen_model
            ):
                raise LocalCuratedEvaluationError(
                    "Configured Qwen3.6 model alias does not match the manifest"
                )
        elif manifest.expected_qwen_model == "qwen3.8-27b-q4km":
            if (
                not settings.local_qwen38_grading_enabled
                or settings.local_qwen38_model != manifest.expected_qwen_model
            ):
                raise LocalCuratedEvaluationError(
                    "Configured Qwen3.8 grading model alias does not match the manifest"
                )
        else:
            raise LocalCuratedEvaluationError("Configured grading model is unsupported")

    from app.db.session import SessionLocal, engine

    configured_name = engine.url.database or ""
    expected_name = f"teacher_assistant_eval_{manifest.run_id}"
    if configured_name != expected_name:
        raise LocalCuratedEvaluationError(
            "The application database engine was initialized for a different database; "
            "start this stage in a fresh process"
        )
    return settings, SessionLocal, storage_root


def _sanitized_status(
    *,
    expected_visual_model: str,
    expected_grading_model: str | None = None,
    require_visual: bool = True,
) -> dict[str, Any]:
    from app.services.local_ai_status_service import LocalAiStatusService

    status = LocalAiStatusService().read()
    if not status["real_providers_allowed"]:
        raise LocalCuratedEvaluationError("Real local-provider switch is not enabled")
    qwen38 = status["qwen38"]
    result = {
        "real_providers_allowed": status["real_providers_allowed"],
        "cohort_model_grading_enabled": status["cohort_model_grading_enabled"],
    }
    if require_visual:
        if (
            not qwen38.get("available")
            or qwen38.get("provider") != "llama_cpp_qwen38"
            or qwen38.get("model") != expected_visual_model
            or qwen38.get("device") != "gpu_hybrid_single_slot"
            or not qwen38.get("visual_preparation_enabled")
            or qwen38.get("detail") != "ready"
        ):
            raise LocalCuratedEvaluationError(
                "Qwen3.8 visual transcription health metadata does not match the run manifest"
            )
        result["qwen38"] = qwen38
    if expected_grading_model is None:
        return result
    if not status["cohort_model_grading_enabled"]:
        raise LocalCuratedEvaluationError("Cohort grading switch is not enabled")
    if expected_grading_model == "qwen3.6-35b-a3b-q4km":
        qwen = status["qwen"]
        if (
            not qwen.get("available")
            or qwen.get("provider") != "llama_cpp_qwen"
            or qwen.get("model") != expected_grading_model
            or qwen.get("device") != "gpu_hybrid"
            or qwen.get("detail") != "ready"
        ):
            raise LocalCuratedEvaluationError(
                "Qwen3.6 grading health metadata does not match the run manifest"
            )
        result["qwen"] = qwen
    elif expected_grading_model != expected_visual_model:
        raise LocalCuratedEvaluationError("Unsupported grading-model alias")
    else:
        if (
            not qwen38.get("available")
            or qwen38.get("provider") != "llama_cpp_qwen38"
            or qwen38.get("model") != expected_grading_model
            or qwen38.get("device") != "gpu_hybrid_single_slot"
            or not qwen38.get("grading_enabled")
            or qwen38.get("detail") not in {"ready", "ready_grading_only"}
        ):
            raise LocalCuratedEvaluationError(
                "Qwen3.8 grading health metadata does not match the run manifest"
            )
        result["qwen38"] = qwen38
    return result


def _directory_digest(path: Path) -> tuple[str, int, int]:
    digest = hashlib.sha256()
    total_size = 0
    file_count = 0
    for file_path in sorted(candidate for candidate in path.rglob("*") if candidate.is_file()):
        relative = file_path.relative_to(path).as_posix()
        file_hash = sha256_file(file_path)
        size = file_path.stat().st_size
        digest.update(f"{relative}\0{size}\0{file_hash}\n".encode())
        total_size += size
        file_count += 1
    return digest.hexdigest(), total_size, file_count


def _gpu_safety_snapshot(*, phase: Literal["Qwen", "Qwen38"]) -> dict[str, Any]:
    runtime_dir = _repository_root() / ".local-ai"
    try:
        active_pid = int(
            (runtime_dir / ("qwen.pid" if phase == "Qwen" else "qwen38.pid"))
            .read_text(encoding="utf-8")
            .strip()
        )
    except (OSError, ValueError) as exc:
        raise LocalCuratedEvaluationError("Active local-model PID record is unavailable") from exc
    try:
        result = subprocess.run(
            [
                "nvidia-smi",
                "--query-compute-apps=pid,process_name,used_gpu_memory",
                "--format=csv,noheader,nounits",
            ],
            check=True,
            capture_output=True,
            text=True,
            timeout=15,
        )
    except (OSError, subprocess.SubprocessError) as exc:
        raise LocalCuratedEvaluationError("Could not verify GPU process isolation") from exc
    compute_pids: set[int] = set()
    process_names: set[str] = set()
    for line in result.stdout.splitlines():
        fields = [field.strip() for field in line.split(",", 2)]
        if not fields or not fields[0].isdigit():
            continue
        compute_pids.add(int(fields[0]))
        if len(fields) > 1:
            process_names.add(Path(fields[1]).name)
    other_pid: int | None = None
    other_pid_path = runtime_dir / ("qwen38.pid" if phase == "Qwen" else "qwen.pid")
    try:
        candidate = int(other_pid_path.read_text(encoding="utf-8").strip())
        if candidate != active_pid:
            other_pid = candidate
    except (OSError, ValueError):
        pass
    snapshot = {
        "active_phase": phase,
        "active_model_present_in_gpu_compute_clients": active_pid in compute_pids,
        "other_managed_model_absent_from_gpu_compute_clients": (
            other_pid is None or other_pid not in compute_pids
        ),
        "gpu_compute_client_names": sorted(process_names),
    }
    if not snapshot["active_model_present_in_gpu_compute_clients"]:
        raise LocalCuratedEvaluationError("The active local model is not visible as a GPU client")
    if not snapshot["other_managed_model_absent_from_gpu_compute_clients"]:
        raise LocalCuratedEvaluationError(
            "Both managed Qwen models are visible in GPU clients at once"
        )
    return snapshot


def _activate_model_for_evaluation(
    *,
    settings: Any,
    session_factory: Any,
    expected_model: str,
    holder_kind: str,
) -> Literal["Qwen", "Qwen38"]:
    """Switch and verify an evaluation model while holding the mandatory lease.

    The status endpoint verifies the server that is currently resident.  Since
    this host intentionally has one GPU slot, an evaluation cannot check both
    models as simultaneously available.  This helper makes the intended phase
    explicit, proves lease ownership during the switch, and verifies the exact
    alias before the next stage is allowed to dispatch a provider call.
    """
    if expected_model == "qwen3.6-35b-a3b-q4km":
        phase: Literal["Qwen", "Qwen38"] = "Qwen"
        provider_name = "llama_cpp_qwen"
    elif expected_model == "qwen3.8-27b-q4km":
        phase = "Qwen38"
        provider_name = "llama_cpp_qwen38"
    else:
        raise LocalCuratedEvaluationError("Unsupported evaluation model alias")

    from uuid import uuid4

    from app.services.local_ai_phase_manager import LocalAiPhaseManager
    from app.services.local_model_lease_service import LocalModelLeaseService
    from packages.brain.adapter import BrainAdapter

    holder_id = f"{holder_kind}:{uuid4().hex}"
    with session_factory() as db:
        lease = LocalModelLeaseService(db)
        with lease.hold(
            model_phase=phase,
            holder_kind=holder_kind,
            holder_id=holder_id,
        ):
            if settings.local_ai_phase_switch_enabled:
                LocalAiPhaseManager(settings=settings, db=db).switch(
                    phase, lease_holder_id=holder_id
                )
            adapter = BrainAdapter.for_provider(settings, provider_name)
            adapter.verify_available_model()
            lease.heartbeat(holder_id=holder_id)
    return phase


def _operator_asset_metadata(expected_qwen_model: str) -> OperatorAssetMetadata:
    """Pin both the chosen text grader and Qwen3.8's visual assets.

    Filesystem paths are intentionally consumed here but never written to the
    manifest.  The projector is part of a visual model's executable identity;
    omitting it would make a transcription result non-reproducible.
    """

    def required_file(label: str, variable: str) -> Path:
        raw_value = os.environ.get(variable, "").strip()
        path = Path(raw_value) if raw_value else None
        if path is None or not path.is_file():
            raise LocalCuratedEvaluationError(f"Operator asset metadata is unavailable: {label}")
        return path

    def assert_build(binary: Path) -> None:
        try:
            version_result = subprocess.run(
                [str(binary), "--version"],
                check=False,
                capture_output=True,
                text=True,
                timeout=15,
            )
        except (OSError, subprocess.TimeoutExpired) as exc:
            raise LocalCuratedEvaluationError("Could not record the llama.cpp build") from exc
        llama_version_text = (version_result.stdout + version_result.stderr).strip()
        if not re.search(rf"\b{EXPECTED_LLAMA_CPP_BUILD}\b", llama_version_text):
            raise LocalCuratedEvaluationError("llama.cpp build does not match 10249")

    qwen38_binary = required_file("Qwen3.8 llama.cpp binary", "LOCAL_QWEN38_BINARY_PATH")
    qwen38_model = required_file("Qwen3.8 model", "LOCAL_QWEN38_MODEL_PATH")
    qwen38_mmproj = qwen38_model.parent / "mmproj-Qwen3.8-27B-Q8_0.gguf"
    if not qwen38_mmproj.is_file():
        raise LocalCuratedEvaluationError(
            "Operator asset metadata is unavailable: Qwen3.8 projector"
        )
    assert_build(qwen38_binary)

    if expected_qwen_model == "qwen3.6-35b-a3b-q4km":
        grader_binary = required_file("Qwen3.6 llama.cpp binary", "LOCAL_QWEN_BINARY_PATH")
        grader_model = required_file("Qwen3.6 model", "LOCAL_QWEN_MODEL_PATH")
        grader_device = "gpu_hybrid"
        assert_build(grader_binary)
    elif expected_qwen_model == "qwen3.8-27b-q4km":
        grader_model = qwen38_model
        grader_device = "gpu_hybrid_single_slot"
    else:
        raise LocalCuratedEvaluationError("Unsupported grading-model alias")
    try:
        return OperatorAssetMetadata.model_validate(
            {
                "llama_cpp": {
                    "build": EXPECTED_LLAMA_CPP_BUILD,
                    "model_alias": expected_qwen_model,
                    "model_sha256": sha256_file(grader_model),
                    "model_size_bytes": grader_model.stat().st_size,
                    "device": grader_device,
                },
                "qwen38_vision": {
                    "build": EXPECTED_LLAMA_CPP_BUILD,
                    "model_alias": "qwen3.8-27b-q4km",
                    "model_sha256": sha256_file(qwen38_model),
                    "model_size_bytes": qwen38_model.stat().st_size,
                    "device": "gpu_hybrid_single_slot",
                    "mmproj_sha256": sha256_file(qwen38_mmproj),
                    "mmproj_size_bytes": qwen38_mmproj.stat().st_size,
                },
            }
        )
    except ValueError as exc:
        raise LocalCuratedEvaluationError(
            "Operator model metadata does not match the locked baseline"
        ) from exc


def _database_is_migrated_and_empty(session_factory: Any) -> None:
    from sqlalchemy import func, select, text

    from app.models import (
        AnswerRegion,
        Assessment,
        FinalGrade,
        GradeSuggestion,
        GradingDispatchRun,
        GradingQueueRun,
        User,
    )

    with session_factory() as db:
        try:
            revision = db.execute(text("SELECT version_num FROM alembic_version")).scalar_one()
        except Exception as exc:
            raise LocalCuratedEvaluationError(
                "Evaluation database is not migrated to the application schema"
            ) from exc
        if revision != "0027_universal_brain":
            raise LocalCuratedEvaluationError("Evaluation database is not at migration head 0027")
        populated_models = [
            model.__name__
            for model in (
                User,
                Assessment,
                AnswerRegion,
                GradingQueueRun,
                GradingDispatchRun,
                GradeSuggestion,
                FinalGrade,
            )
            if int(db.scalar(select(func.count(model.id))) or 0) != 0
        ]
        if populated_models:
            raise LocalCuratedEvaluationError(
                "Evaluation database must be empty before seeding synthetic cases"
            )


def _seed_production_evaluation(
    manifest: LocalCuratedEvaluationManifest,
    run_dir: Path,
    session_factory: Any,
) -> dict[str, Any]:
    from app.models import (
        AnswerRegion,
        AnswerRegionMapping,
        AnswerRegionSegment,
        Assessment,
        Course,
        ExtractionRun,
        GradingRun,
        Question,
        QuestionNode,
        Rubric,
        Submission,
        SubmissionPage,
        User,
    )
    from app.services.storage import LocalStorage

    storage = LocalStorage()
    with session_factory() as db:
        owner = User(
            name="Local Curated Evaluation Owner",
            email=f"owner-{manifest.run_id}@example.invalid",
            password_hash="evaluation-login-disabled",
            role="teacher",
        )
        intruder = User(
            name="Local Curated Evaluation Intruder",
            email=f"intruder-{manifest.run_id}@example.invalid",
            password_hash="evaluation-login-disabled",
            role="teacher",
        )
        db.add_all([owner, intruder])
        db.flush()
        course = Course(
            teacher_id=owner.id,
            code=f"LCE-{manifest.run_id[-12:]}",
            title="Synthetic local curated evaluation",
            department="Evaluation",
            semester="synthetic",
        )
        db.add(course)
        db.flush()
        assessment = Assessment(
            course_id=course.id,
            title=f"20-case local curated evaluation {manifest.run_id}",
            assessment_type="exam",
            total_marks=Decimal("23"),
            status="ready",
        )
        db.add(assessment)
        db.flush()
        grading_run = GradingRun(
            assessment_id=assessment.id,
            created_by_teacher_id=owner.id,
            mode="custom_controlled",
            status="grading_ready",
            marking_policy="general",
            materials_confirmed_at=datetime.now(UTC),
            questions_confirmed_at=datetime.now(UTC),
            rubrics_confirmed_at=datetime.now(UTC),
            notes="Synthetic evaluation only; no student data.",
        )
        db.add(grading_run)
        db.flush()

        question_ids: dict[str, int] = {}
        rubric_ids: dict[str, int] = {}
        question_nodes: dict[str, QuestionNode] = {}
        extraction_run = ExtractionRun(
            assessment_id=assessment.id,
            artifact_file_path="evaluation://teacher-locked-reference-bundle",
            original_filename="teacher_locked_evaluation_reference.json",
            content_type="application/vnd.teacher-assistant.evaluation+json",
            extraction_type="question_paper",
            provider="llama_cpp_qwen38",
            status="succeeded",
            raw_output=None,
            normalized_output={
                "source": "teacher_locked_evaluation_blueprint",
                "teacher_confirmed": True,
            },
            blockers=[],
        )
        db.add(extraction_run)
        db.flush()
        for pack_id in "ABCDE":
            representative = next(case for case in manifest.cases if case.pack_id == pack_id)
            question = Question(
                assessment_id=assessment.id,
                question_no=pack_id,
                question_text=representative.question_text,
                model_answer=representative.model_answer,
                total_marks=representative.max_score,
            )
            db.add(question)
            db.flush()
            rubric_json = {
                "total_marks": str(representative.max_score),
                "criteria": [
                    {
                        "id": criterion.id,
                        "name": criterion.name,
                        "description": criterion.description,
                        "max_marks": str(criterion.max_marks),
                        "depends_on": criterion.depends_on,
                        "allow_follow_through": criterion.allow_follow_through,
                    }
                    for criterion in representative.rubric
                ],
            }
            rubric = Rubric(
                question_id=question.id,
                version=1,
                rubric_json=rubric_json,
                is_active=True,
            )
            db.add(rubric)
            db.flush()
            question_node = QuestionNode(
                assessment_id=assessment.id,
                extraction_run_id=extraction_run.id,
                question_number=pack_id,
                parent_question_number=None,
                label=f"Evaluation question {pack_id}",
                text=representative.question_text,
                marks=representative.max_score,
                node_type="question",
                source_page=1,
                source_reference={
                    "source": "teacher_locked_evaluation_blueprint",
                    "pack_id": pack_id,
                },
                confidence=Decimal("1.0000"),
                teacher_confirmed=True,
            )
            db.add(question_node)
            db.flush()
            question_ids[pack_id] = question.id
            rubric_ids[pack_id] = rubric.id
            question_nodes[pack_id] = question_node

        region_ids: dict[str, int] = {}
        for response_index in range(1, 5):
            submission = Submission(
                assessment_id=assessment.id,
                student_identifier=f"SYN-EVAL-{response_index:02d}",
                student_name=None,
                status="ready",
            )
            db.add(submission)
            db.flush()
            for page_number, pack_id in enumerate("ABCDE", start=1):
                case = next(
                    item
                    for item in manifest.cases
                    if item.case_id == f"{pack_id}{response_index}"
                )
                stored_page = storage.page_image_path(submission.id, page_number)
                stored_page.absolute_path.parent.mkdir(parents=True, exist_ok=True)
                shutil.copyfile(run_dir / case.image_relative_path, stored_page.absolute_path)
                page = SubmissionPage(
                    submission_id=submission.id,
                    page_no=page_number,
                    image_path=stored_page.relative_path,
                    quality_score=Decimal("1.0000"),
                )
                db.add(page)
                db.flush()
                region = AnswerRegion(
                    submission_id=submission.id,
                    question_id=question_ids[pack_id],
                    page_id=page.id,
                    x=Decimal("0"),
                    y=Decimal("0"),
                    width=Decimal("1600"),
                    height=Decimal("900"),
                    image_path=stored_page.relative_path,
                    manual_answer_text=None,
                    full_answer_confirmed=False,
                    evidence_status="unconfirmed",
                    continuation_check_status="not_checked",
                )
                db.add(region)
                db.flush()
                db.add(
                    AnswerRegionSegment(
                        answer_region_id=region.id,
                        submission_page_id=page.id,
                        order_index=1,
                        x=region.x,
                        y=region.y,
                        width=region.width,
                        height=region.height,
                        image_path=stored_page.relative_path,
                        source="manual",
                        confirmed=True,
                        is_primary=True,
                    )
                )
                db.add(
                    AnswerRegionMapping(
                        assessment_id=assessment.id,
                        submission_id=submission.id,
                        question_node_id=question_nodes[pack_id].id,
                        question_id=question_ids[pack_id],
                        answer_region_id=region.id,
                        source_page=page_number,
                        source_reference={
                            "source": "teacher_locked_evaluation_mapping",
                            "case_id": case.case_id,
                            "image_sha256": case.image_sha256,
                        },
                        confidence=Decimal("1.0000"),
                        mapping_status="teacher_confirmed",
                        blocker_reason=None,
                        provider="deterministic_layout",
                        teacher_confirmed=True,
                    )
                )
                region_ids[case.case_id] = region.id
        db.commit()
        return {
            "owner_teacher_id": owner.id,
            "intruder_teacher_id": intruder.id,
            "assessment_id": assessment.id,
            "grading_run_id": grading_run.id,
            "question_ids": question_ids,
            "rubric_ids": rubric_ids,
            "region_ids": region_ids,
            "mapping_provenance": "teacher_locked_evaluation_mapping",
        }


def _auth_headers(
    session_factory: Any,
    owner_id: int,
    intruder_id: int,
) -> tuple[dict[str, str], dict[str, str]]:
    from app.core.auth import create_access_token
    from app.models import User

    with session_factory() as db:
        owner = db.get(User, owner_id)
        intruder = db.get(User, intruder_id)
        if owner is None or intruder is None:
            raise LocalCuratedEvaluationError("Evaluation teachers are missing")
        owner_token = create_access_token(owner)
        intruder_token = create_access_token(intruder)
    return (
        {"Authorization": f"Bearer {owner_token}"},
        {"Authorization": f"Bearer {intruder_token}"},
    )


def create_ocr_review_workbook(
    manifest: LocalCuratedEvaluationManifest,
    ground_truth: GroundTruthLock,
    ocr_result: OcrRunResult,
    run_dir: Path,
) -> Path:
    truth_by_id = {case.case_id: case for case in ground_truth.cases}
    ocr_by_id = {case.case_id: case for case in ocr_result.cases}
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instructions.append(["OCR confirmation gate"])
    instructions.append(["Run ID", manifest.run_id])
    instructions.append(
        [
            "Required",
            (
                "Review Qwen3.8's non-thinking visual transcription against the image and "
                "locked ground truth. Do not edit or repair model text: approval confirms the "
                "exact displayed hash. Mark every row faithful only if it matches the pixels. "
                "For any wrong reading, set teacher_confirms_faithful to no, choose one fixed "
                "quality_rejection_reason, then use record-ocr-no-go. That creates a valid "
                "NO_GO_QUALITY report with zero grading calls. Blank rows must remain genuinely "
                "empty."
            ),
        ]
    )
    instructions["A1"].font = Font(bold=True, size=16)
    instructions.column_dimensions["A"].width = 20
    instructions.column_dimensions["B"].width = 110

    sheet = workbook.create_sheet("OCR Review")
    headers = [
        "case_id",
        "primary_category",
        "qwen38_visual_draft",
        "locked_ground_truth",
        "visual_model",
        "reasoning_mode",
        "warnings",
        "cer",
        "wer",
        "latency_ms",
        "visual_draft_sha256",
        "teacher_confirms_faithful",
        "quality_rejection_reason",
        "teacher_notes",
        "teacher_approved",
        "answer_image",
    ]
    sheet.append(headers)
    _style_workbook_header(sheet)
    widths = [12, 24, 55, 55, 28, 18, 35, 14, 14, 16, 66, 28, 30, 45, 20, 72]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row_index, case in enumerate(manifest.cases, start=2):
        ocr_case = ocr_by_id[case.case_id]
        truth = truth_by_id[case.case_id]
        sheet.append(
            [
                case.case_id,
                case.primary_category.value,
                ocr_case.draft_text,
                truth.teacher_transcription,
                ocr_case.model,
                ocr_case.reasoning_mode,
                "\n".join(ocr_case.warnings),
                str(character_error_rate(truth.teacher_transcription, ocr_case.draft_text)),
                str(word_error_rate(truth.teacher_transcription, ocr_case.draft_text)),
                ocr_case.latency_ms,
                ocr_case.draft_text_sha256,
                None,
                None,
                None,
                None,
                None,
            ]
        )
        sheet.row_dimensions[row_index].height = 250
        for cell in sheet[row_index]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        workbook_image = WorkbookImage(str(run_dir / case.image_relative_path))
        workbook_image.width = 560
        workbook_image.height = 315
        sheet.add_image(workbook_image, f"P{row_index}")
    sheet.freeze_panes = "A2"
    path = run_dir / "ocr_review.xlsx"
    workbook.save(path)
    return path


def _mark_invalid(run_dir: Path, *, stage: str, detail: str) -> None:
    if current_state(run_dir) in {"invalid", "reported"}:
        return
    payload = {
        "stage": stage,
        "recorded_at": datetime.now(UTC),
        "detail": detail[:500],
        "automatic_retry_allowed": False,
    }
    path = run_dir / "invalid_run.json"
    write_json(path, payload)
    try:
        append_state(
            run_dir,
            "invalid",
            locked_artifacts={"invalid_run.json": sha256_file(path)},
            metadata={"stage": stage},
        )
    except LocalCuratedEvaluationError:
        pass


def run_ocr_stage(
    run_dir: Path,
    *,
    allow_local_ocr: bool,
    max_ocr_calls: int,
    database_url: str,
    local_ai_env: Path | None = None,
) -> OcrRunResult:
    """Run exactly twenty production Qwen3.8 visual-transcription calls.

    The evaluation deliberately calls the same durable service used by the
    teacher workflow.  It does not use a fake OCR adapter, a hidden retry, or
    the retired PaddleOCR endpoints.  Every synthetic region is pre-mapped by
    the locked evaluation blueprint and then read by Qwen3.8 with reasoning
    disabled through one leased visual call.
    """
    if not allow_local_ocr:
        raise LocalCuratedEvaluationError(
            "Real Qwen3.8 visual transcription requires --allow-local-ocr"
        )
    if max_ocr_calls != OCR_CALL_LIMIT:
        raise LocalCuratedEvaluationError("The visual-transcription call cap must be exactly 20")
    if current_state(run_dir) != "ground_truth_locked":
        raise LocalCuratedEvaluationError(
            "Visual transcription requires a locked teacher ground-truth gate"
        )
    try:
        verify_locked_artifacts(run_dir)
        require_clean_git_worktree()
        manifest = load_manifest(run_dir)
        if current_git_commit() != manifest.harness_commit:
            raise LocalCuratedEvaluationError(
                "Current Git commit does not match the harness commit"
            )
        ground_truth = GroundTruthLock.model_validate(read_json(run_dir / "ground_truth_lock.json"))
        settings, session_factory, _storage_root = _configure_runtime(
            run_dir,
            database_url=database_url,
            local_ai_env=local_ai_env,
            require_grading=False,
        )
        if datetime.now(UTC) <= ground_truth.signed_at:
            raise LocalCuratedEvaluationError(
                "Visual-transcription calls must follow ground-truth sign-off"
            )
        current_assets = _operator_asset_metadata(manifest.expected_qwen_model)
        if current_assets != manifest.operator_assets:
            raise LocalCuratedEvaluationError(
                "Pinned local-model assets changed after the evaluation was prepared"
            )
        _database_is_migrated_and_empty(session_factory)
        seeded = _seed_production_evaluation(manifest, run_dir, session_factory)
        _activate_model_for_evaluation(
            settings=settings,
            session_factory=session_factory,
            expected_model=manifest.operator_assets.qwen38_vision.model_alias,
            holder_kind="evaluation_visual_preflight",
        )
        status_before = _sanitized_status(
            expected_visual_model=manifest.operator_assets.qwen38_vision.model_alias
        )
        gpu_before = _gpu_safety_snapshot(phase="Qwen38")

        from fastapi import HTTPException

        from app.api.routes.answer_regions import get_owned_answer_region_or_404
        from app.models import AnswerRegion, AnswerRegionOcrRun, User
        from app.services.qwen38_visual_transcription_service import (
            Qwen38VisualTranscriptionService,
        )
        from app.worker.jobs import run_qwen38_visual_transcription_job

        cases: list[OcrCaseResult] = []
        first_call_at: datetime | None = None
        for definition in manifest.cases:
            answer_region_id = int(seeded["region_ids"][definition.case_id])
            with session_factory() as db:
                region = db.get(AnswerRegion, answer_region_id)
                owner = db.get(User, int(seeded["owner_teacher_id"]))
                if region is None or owner is None:
                    raise LocalCuratedEvaluationError("Seeded evaluation evidence is missing")
                created = Qwen38VisualTranscriptionService(db, settings=settings).create(
                    region,
                    teacher=owner,
                    expected_model=manifest.operator_assets.qwen38_vision.model_alias,
                )
                run_id = created.id

            # Production worker entry point, intentionally invoked synchronously
            # because this bounded host evaluation does not claim RQ crash
            # recovery validation.  The job itself opens its own DB session.
            run_qwen38_visual_transcription_job(run_id)

            with session_factory() as db:
                run = db.get(AnswerRegionOcrRun, run_id)
                if (
                    run is None
                    or run.status != "succeeded"
                    or run.calls_used != 1
                    or run.call_limit != 1
                    or run.provider != "llama_cpp_qwen38"
                    or run.model_name != manifest.operator_assets.qwen38_vision.model_alias
                    or run.profile != "qwen38_verbatim_visual"
                    or run.reasoning_mode != "off"
                    or not run.source_image_sha256
                    or not run.source_image_hashes
                    or run.latency_ms is None
                ):
                    raise LocalCuratedEvaluationError(
                        f"Visual transcription failed safely for {definition.case_id}"
                    )
                normalized = dict(run.normalized_result or {})
                draft_text = run.draft_text or ""
                if normalized.get("draft_text_sha256") != sha256_text(draft_text):
                    raise LocalCuratedEvaluationError(
                        f"Visual transcription output hash is invalid for {definition.case_id}"
                    )
                if list(normalized.get("input_image_sha256") or []) != list(
                    run.source_image_hashes
                ):
                    raise LocalCuratedEvaluationError(
                        f"Visual transcription image hashes changed for {definition.case_id}"
                    )
                if first_call_at is None:
                    first_call_at = run.started_at or datetime.now(UTC)
                cases.append(
                    OcrCaseResult(
                        case_id=definition.case_id,
                        answer_region_id=answer_region_id,
                        ocr_run_id=run.id,
                        status="succeeded",
                        draft_text=draft_text,
                        markdown=draft_text,
                        blocks=[],
                        warnings=list(run.warnings or []),
                        latency_ms=run.latency_ms,
                        provider="llama_cpp_qwen38",
                        model="qwen3.8-27b-q4km",
                        profile="qwen38_verbatim_visual",
                        reasoning_mode="off",
                        source_image_sha256=run.source_image_sha256,
                        source_image_hashes=list(run.source_image_hashes),
                        is_blank=bool(normalized.get("is_blank")),
                        draft_text_sha256=sha256_text(draft_text),
                    )
                )

        if len(cases) != OCR_CALL_LIMIT:
            raise LocalCuratedEvaluationError(
                "The visual-transcription call count is not exactly 20"
            )
        with session_factory() as db:
            owner = db.get(User, int(seeded["owner_teacher_id"]))
            intruder = db.get(User, int(seeded["intruder_teacher_id"]))
            if owner is None or intruder is None:
                raise LocalCuratedEvaluationError("Evaluation teachers disappeared")
            for case in cases:
                # Use the production ownership helper rather than assuming a
                # synthetic database makes isolation true by construction.
                get_owned_answer_region_or_404(case.answer_region_id, db, owner)
                try:
                    get_owned_answer_region_or_404(case.answer_region_id, db, intruder)
                except HTTPException as exc:
                    if exc.status_code != 404:
                        raise LocalCuratedEvaluationError(
                            "Cross-teacher visual-evidence access did not return 404"
                        ) from exc
                else:
                    raise LocalCuratedEvaluationError(
                        "Cross-teacher visual-evidence access was not refused"
                    )

        status_after = _sanitized_status(
            expected_visual_model=manifest.operator_assets.qwen38_vision.model_alias
        )
        gpu_after = _gpu_safety_snapshot(phase="Qwen38")
        result = OcrRunResult(
            run_id=manifest.run_id,
            first_call_at=first_call_at or datetime.now(UTC),
            completed_at=datetime.now(UTC),
            call_count=OCR_CALL_LIMIT,
            retry_count=0,
            service_status_before=status_before,
            service_status_after=status_after,
            database_name=validate_database_url(database_url, manifest.run_id),
            assessment_id=int(seeded["assessment_id"]),
            grading_run_id=int(seeded["grading_run_id"]),
            owner_teacher_id=int(seeded["owner_teacher_id"]),
            intruder_teacher_id=int(seeded["intruder_teacher_id"]),
            question_ids={key: int(value) for key, value in seeded["question_ids"].items()},
            rubric_ids={key: int(value) for key, value in seeded["rubric_ids"].items()},
            cases=cases,
        )
        result_path = run_dir / "ocr_results.json"
        runtime_path = run_dir / "ocr_runtime.json"
        write_json(result_path, result)
        write_json(
            runtime_path,
            {
                "operator_assets": manifest.operator_assets,
                "service_status_before": status_before,
                "service_status_after": status_after,
                "gpu_safety_before": gpu_before,
                "gpu_safety_after": gpu_after,
                "mapping_provenance": seeded["mapping_provenance"],
                "visual_model": manifest.operator_assets.qwen38_vision.model_alias,
                "visual_reasoning_mode": "off",
                "transport": "direct_host_eval",
                "rq_crash_recovery_validated": False,
            },
        )
        create_ocr_review_workbook(manifest, ground_truth, result, run_dir)
        append_state(
            run_dir,
            "ocr_completed",
            locked_artifacts={
                "ocr_results.json": sha256_file(result_path),
                "ocr_runtime.json": sha256_file(runtime_path),
            },
            metadata={
                "qwen38_visual_call_count": OCR_CALL_LIMIT,
                "retry_count": 0,
                "reasoning_mode": "off",
                "cross_teacher_visual_evidence_refused": True,
                "transport": "direct_host_eval",
            },
        )
        return result
    except Exception as exc:
        _mark_invalid(
            run_dir,
            stage="run_ocr_stage",
            detail=(
                "Qwen3.8 visual transcription failed; no OCR confirmation or grading "
                "was authorized."
            ),
        )
        if isinstance(exc, LocalCuratedEvaluationError):
            raise
        raise LocalCuratedEvaluationError("Qwen3.8 visual transcription failed safely") from exc


def read_ocr_review_workbook(
    workbook_path: Path,
    manifest: LocalCuratedEvaluationManifest,
    ocr_result: OcrRunResult,
    ground_truth: GroundTruthLock,
) -> list[OcrConfirmationCase]:
    workbook = load_workbook(workbook_path, data_only=True)
    if "OCR Review" not in workbook.sheetnames:
        raise LocalCuratedEvaluationError("OCR workbook has no OCR Review sheet")
    sheet = workbook["OCR Review"]
    headers = {str(cell.value): index for index, cell in enumerate(sheet[1], start=1)}
    required = {
        "case_id",
        "visual_draft_sha256",
        "teacher_confirms_faithful",
        "teacher_notes",
        "teacher_approved",
    }
    if not required.issubset(headers):
        raise LocalCuratedEvaluationError("OCR workbook columns were changed")
    definition_by_id = {case.case_id: case for case in manifest.cases}
    truth_by_id = {case.case_id: case for case in ground_truth.cases}
    ocr_by_id = {case.case_id: case for case in ocr_result.cases}
    confirmations: list[OcrConfirmationCase] = []
    for row in range(2, sheet.max_row + 1):
        case_id = str(sheet.cell(row, headers["case_id"]).value or "").strip()
        if not case_id:
            continue
        definition = definition_by_id.get(case_id)
        truth = truth_by_id.get(case_id)
        ocr_case = ocr_by_id.get(case_id)
        if definition is None or truth is None or ocr_case is None:
            raise LocalCuratedEvaluationError(f"Unknown OCR review case: {case_id}")
        displayed_hash = str(
            sheet.cell(row, headers["visual_draft_sha256"]).value or ""
        ).strip()
        if displayed_hash != ocr_case.draft_text_sha256:
            raise LocalCuratedEvaluationError(
                f"{case_id}: displayed visual-transcription hash was changed"
            )
        faithful = _parse_bool(
            sheet.cell(row, headers["teacher_confirms_faithful"]).value,
            field="teacher_confirms_faithful",
            case_id=case_id,
        )
        approved = _parse_bool(
            sheet.cell(row, headers["teacher_approved"]).value,
            field="teacher_approved",
            case_id=case_id,
        )
        if not faithful or not approved:
            raise LocalCuratedEvaluationError(
                f"{case_id}: a teacher must confirm the exact visual transcription is faithful"
            )
        if normalize_text(ocr_case.draft_text) != normalize_text(truth.teacher_transcription):
            raise LocalCuratedEvaluationError(
                f"{case_id}: Qwen3.8 visual transcription does not match locked teacher truth; "
                "reject the model reading and do not grade it"
            )
        is_blank = definition.answer_quality == AnswerQuality.BLANK
        if ocr_case.is_blank != is_blank:
            raise LocalCuratedEvaluationError(
                f"{case_id}: Qwen3.8 blank classification does not match locked teacher truth"
            )
        confirmations.append(
            OcrConfirmationCase(
                case_id=case_id,
                answer_region_id=ocr_case.answer_region_id,
                ocr_run_id=ocr_case.ocr_run_id,
                confirmed_text=ocr_case.draft_text,
                confirmed_text_sha256=ocr_case.draft_text_sha256,
                teacher_approved=True,
                evidence_status="blank" if is_blank else "complete",
                full_answer_confirmed=not is_blank,
            )
        )
    expected_order = [case.case_id for case in manifest.cases]
    if [case.case_id for case in confirmations] != expected_order:
        raise LocalCuratedEvaluationError("All 20 OCR cases must be confirmed in blueprint order")
    return confirmations


def read_ocr_quality_no_go_workbook(
    workbook_path: Path,
    manifest: LocalCuratedEvaluationManifest,
    ocr_result: OcrRunResult,
    *,
    require_rejection: bool = True,
) -> list[OcrQualityNoGoCase]:
    """Read a teacher-declared OCR quality failure without accepting any text.

    The only durable information is the locked draft hash and a fixed reason
    code.  In particular, this must not turn the spreadsheet into a back-door
    transcription or correction mechanism.
    """

    workbook = load_workbook(workbook_path, data_only=True)
    if "OCR Review" not in workbook.sheetnames:
        raise LocalCuratedEvaluationError("OCR workbook has no OCR Review sheet")
    sheet = workbook["OCR Review"]
    headers = {str(cell.value): index for index, cell in enumerate(sheet[1], start=1)}
    required = {
        "case_id",
        "visual_draft_sha256",
        "teacher_confirms_faithful",
        "quality_rejection_reason",
        "teacher_approved",
    }
    if not required.issubset(headers):
        raise LocalCuratedEvaluationError("OCR workbook columns were changed")
    known_case_ids = {case.case_id for case in manifest.cases}
    result_by_id = {case.case_id: case for case in ocr_result.cases}
    rejected: list[OcrQualityNoGoCase] = []
    for row in range(2, sheet.max_row + 1):
        case_id = str(sheet.cell(row, headers["case_id"]).value or "").strip()
        if not case_id:
            continue
        if case_id not in known_case_ids or case_id not in result_by_id:
            raise LocalCuratedEvaluationError(f"Unknown OCR review case: {case_id}")
        displayed_hash = str(
            sheet.cell(row, headers["visual_draft_sha256"]).value or ""
        ).strip()
        if displayed_hash != result_by_id[case_id].draft_text_sha256:
            raise LocalCuratedEvaluationError(
                f"{case_id}: displayed visual-transcription hash was changed"
            )
        faithful_value = sheet.cell(row, headers["teacher_confirms_faithful"]).value
        reason = str(
            sheet.cell(row, headers["quality_rejection_reason"]).value or ""
        ).strip()
        if faithful_value is None or not str(faithful_value).strip():
            if reason:
                raise LocalCuratedEvaluationError(
                    f"{case_id}: a quality rejection reason requires an explicit no"
                )
            continue
        faithful = _parse_bool(
            faithful_value,
            field="teacher_confirms_faithful",
            case_id=case_id,
        )
        if faithful:
            if reason:
                raise LocalCuratedEvaluationError(
                    f"{case_id}: a faithful reading cannot have a rejection reason"
                )
            continue
        approved = _parse_bool(
            sheet.cell(row, headers["teacher_approved"]).value,
            field="teacher_approved",
            case_id=case_id,
        )
        if approved:
            raise LocalCuratedEvaluationError(
                f"{case_id}: a rejected visual reading cannot be approved"
            )
        if reason not in _OCR_QUALITY_NO_GO_REASONS:
            allowed = ", ".join(sorted(_OCR_QUALITY_NO_GO_REASONS))
            raise LocalCuratedEvaluationError(
                f"{case_id}: quality_rejection_reason must be one of {allowed}"
            )
        rejected.append(OcrQualityNoGoCase(case_id=case_id, reason=reason))
    if not rejected and require_rejection:
        raise LocalCuratedEvaluationError(
            "At least one teacher-declared visual-transcription quality failure is required"
        )
    return rejected


def _assert_no_grading_artifacts_or_records(
    run_dir: Path,
    *,
    session_factory: Any,
) -> None:
    """Prove a quality no-go happened before any grade could exist."""

    forbidden_artifacts = (
        "ocr_confirmation_lock.json",
        "grading_results.json",
        "grading_runtime.json",
        "grading_review.xlsx",
        "review_lock.json",
    )
    present = [name for name in forbidden_artifacts if (run_dir / name).exists()]
    if present:
        raise LocalCuratedEvaluationError(
            "OCR quality no-go cannot follow grading artifacts: " + ", ".join(present)
        )
    from sqlalchemy import func, select

    from app.models import FinalGrade, GradeSuggestion, GradingDispatchRun, GradingJob

    with session_factory() as db:
        counts = {
            model.__tablename__: int(db.scalar(select(func.count()).select_from(model)) or 0)
            for model in (GradingJob, GradeSuggestion, GradingDispatchRun, FinalGrade)
        }
    present_records = {name: count for name, count in counts.items() if count}
    if present_records:
        raise LocalCuratedEvaluationError(
            "OCR quality no-go found grading records in its disposable database"
        )


def record_ocr_quality_no_go(
    run_dir: Path,
    *,
    reviewer_id: str,
    confirm_teacher_signoff: bool,
    database_url: str,
    local_ai_env: Path | None = None,
) -> OcrQualityNoGoLock:
    """Close a valid but poor visual-OCR run without grading any student answer."""

    if not confirm_teacher_signoff:
        raise LocalCuratedEvaluationError("Explicit OCR quality no-go sign-off is required")
    if current_state(run_dir) != "ocr_completed":
        raise LocalCuratedEvaluationError(
            "An OCR quality no-go requires a completed visual-transcription stage"
        )
    try:
        verify_locked_artifacts(run_dir)
        manifest = load_manifest(run_dir)
        ocr_result = OcrRunResult.model_validate(read_json(run_dir / "ocr_results.json"))
        workbook_path = run_dir / "ocr_review.xlsx"
        rejected_cases = read_ocr_quality_no_go_workbook(
            workbook_path,
            manifest,
            ocr_result,
        )
        _settings, session_factory, _storage_root = _configure_runtime(
            run_dir,
            database_url=database_url,
            local_ai_env=local_ai_env,
        )
        _assert_no_grading_artifacts_or_records(run_dir, session_factory=session_factory)
        lock = OcrQualityNoGoLock(
            run_id=manifest.run_id,
            reviewer_id=reviewer_id.strip(),
            signed_at=datetime.now(UTC),
            ocr_results_sha256=sha256_file(run_dir / "ocr_results.json"),
            workbook_sha256=sha256_file(workbook_path),
            rejected_cases=rejected_cases,
        )
        lock_path = run_dir / "ocr_quality_no_go_lock.json"
        write_json(lock_path, lock)
        append_state(
            run_dir,
            "ocr_quality_no_go",
            locked_artifacts={
                "ocr_review.xlsx": sha256_file(workbook_path),
                "ocr_quality_no_go_lock.json": sha256_file(lock_path),
            },
            metadata={
                "rejected_case_count": len(rejected_cases),
                "grading_calls": 0,
                "automatic_retry_allowed": False,
            },
        )
        return lock
    except Exception as exc:
        _mark_invalid(
            run_dir,
            stage="record_ocr_quality_no_go",
            detail="OCR quality no-go recording failed; grading remained unauthorized.",
        )
        if isinstance(exc, LocalCuratedEvaluationError):
            raise
        raise LocalCuratedEvaluationError("OCR quality no-go recording failed safely") from exc


def lock_ocr_confirmations(
    run_dir: Path,
    *,
    reviewer_id: str,
    confirm_teacher_signoff: bool,
    database_url: str,
    local_ai_env: Path | None = None,
) -> OcrConfirmationLock:
    if not confirm_teacher_signoff:
        raise LocalCuratedEvaluationError("Explicit OCR confirmation sign-off is required")
    if current_state(run_dir) != "ocr_completed":
        raise LocalCuratedEvaluationError("OCR confirmation requires a completed OCR stage")
    verify_locked_artifacts(run_dir)
    manifest = load_manifest(run_dir)
    ground_truth = GroundTruthLock.model_validate(read_json(run_dir / "ground_truth_lock.json"))
    ocr_result = OcrRunResult.model_validate(read_json(run_dir / "ocr_results.json"))
    if ocr_result.first_call_at <= ground_truth.signed_at:
        raise LocalCuratedEvaluationError("OCR calls did not follow ground-truth sign-off")
    workbook_path = run_dir / "ocr_review.xlsx"
    declared_rejections = read_ocr_quality_no_go_workbook(
        workbook_path,
        manifest,
        ocr_result,
        require_rejection=False,
    )
    if declared_rejections:
        raise LocalCuratedEvaluationError(
            "Teacher rejected visual transcription; use record-ocr-no-go instead of grading"
        )
    confirmations = read_ocr_review_workbook(
        workbook_path,
        manifest,
        ocr_result,
        ground_truth,
    )
    _settings, session_factory, _storage_root = _configure_runtime(
        run_dir,
        database_url=database_url,
        local_ai_env=local_ai_env,
    )
    try:
        from fastapi import HTTPException

        from app.api.routes.answer_regions import (
            confirm_visual_transcription_run,
            correct_answer_region_full_answer_confirmation,
        )
        from app.models import AnswerRegion, AnswerRegionOcrRun, User
        from app.schemas import (
            AnswerRegionFullAnswerConfirmation,
            VisualTranscriptionConfirmationRequest,
        )

        with session_factory() as db:
            owner = db.get(User, ocr_result.owner_teacher_id)
            intruder = db.get(User, ocr_result.intruder_teacher_id)
            if owner is None or intruder is None:
                raise LocalCuratedEvaluationError("Evaluation teachers are missing")
            for confirmation in confirmations:
                try:
                    confirm_visual_transcription_run(
                        confirmation.answer_region_id,
                        confirmation.ocr_run_id,
                        VisualTranscriptionConfirmationRequest(
                            teacher_confirmed=True,
                            draft_text_sha256=confirmation.confirmed_text_sha256,
                        ),
                        db,
                        intruder,
                    )
                except HTTPException as exc:
                    if exc.status_code != 404:
                        raise LocalCuratedEvaluationError(
                            "Cross-teacher visual confirmation was not refused"
                        ) from exc
                else:
                    raise LocalCuratedEvaluationError(
                        "Cross-teacher visual confirmation was not refused"
                    )
            for confirmation in confirmations:
                confirmed = confirm_visual_transcription_run(
                    confirmation.answer_region_id,
                    confirmation.ocr_run_id,
                    VisualTranscriptionConfirmationRequest(
                        teacher_confirmed=True,
                        draft_text_sha256=confirmation.confirmed_text_sha256,
                    ),
                    db,
                    owner,
                )
                if confirmed.status != "confirmed":
                    raise LocalCuratedEvaluationError(
                        "Production visual-transcription confirmation failed on "
                        f"{confirmation.case_id}"
                    )
                evidence = correct_answer_region_full_answer_confirmation(
                    confirmation.answer_region_id,
                    AnswerRegionFullAnswerConfirmation(
                        full_answer_confirmed=confirmation.full_answer_confirmed,
                        continuation_not_needed=True,
                        packet_status=confirmation.evidence_status,
                        # This is the exact already-confirmed visual draft, not a
                        # teacher-entered correction.  The server reuses it only
                        # to set the existing full-answer evidence gate.
                        manual_answer_text=confirmation.confirmed_text,
                    ),
                    db,
                    owner,
                )
                if evidence.answer_region.id != confirmation.answer_region_id:
                    raise LocalCuratedEvaluationError(
                        f"Full-answer evidence confirmation failed on {confirmation.case_id}"
                    )
        with session_factory() as db:
            for confirmation in confirmations:
                region = db.get(AnswerRegion, confirmation.answer_region_id)
                ocr_run = db.get(AnswerRegionOcrRun, confirmation.ocr_run_id)
                if region is None or ocr_run is None or ocr_run.status != "confirmed":
                    raise LocalCuratedEvaluationError(
                        f"Persisted OCR confirmation is missing for {confirmation.case_id}"
                    )
                if region.evidence_status != confirmation.evidence_status:
                    raise LocalCuratedEvaluationError(
                        f"Evidence status mismatch for {confirmation.case_id}"
                    )
                if region.full_answer_confirmed != confirmation.full_answer_confirmed:
                    raise LocalCuratedEvaluationError(
                        f"Full-answer confirmation mismatch for {confirmation.case_id}"
                    )
                if normalize_text(region.manual_answer_text or "") != normalize_text(
                    confirmation.confirmed_text
                ):
                    raise LocalCuratedEvaluationError(
                        f"Confirmed text mismatch for {confirmation.case_id}"
                    )
        lock = OcrConfirmationLock(
            run_id=manifest.run_id,
            reviewer_id=reviewer_id.strip(),
            signed_at=datetime.now(UTC),
            ocr_results_sha256=sha256_file(run_dir / "ocr_results.json"),
            workbook_sha256=sha256_file(workbook_path),
            cases=confirmations,
        )
        lock_path = run_dir / "ocr_confirmation_lock.json"
        write_json(lock_path, lock)
        append_state(
            run_dir,
            "ocr_confirmed",
            locked_artifacts={
                "ocr_review.xlsx": sha256_file(workbook_path),
                "ocr_confirmation_lock.json": sha256_file(lock_path),
            },
            metadata={
                "confirmed_case_count": 20,
                "complete_case_count": 18,
                "blank_case_count": 2,
                "cross_teacher_confirmation_refused": True,
            },
        )
        return lock
    except LocalCuratedEvaluationError as exc:
        if (
            current_state(run_dir) == "ocr_completed"
            and str(exc).startswith("Teacher rejected visual transcription")
        ):
            raise
        _mark_invalid(
            run_dir,
            stage="lock_ocr_confirmations",
            detail="OCR confirmation stage failed; no grading was authorized.",
        )
        raise
    except Exception as exc:
        _mark_invalid(
            run_dir,
            stage="lock_ocr_confirmations",
            detail="OCR confirmation stage failed; no grading was authorized.",
        )
        raise LocalCuratedEvaluationError("OCR confirmation failed safely") from exc


def create_grading_review_workbook(
    manifest: LocalCuratedEvaluationManifest,
    grading_result: GradingRunResult,
    run_dir: Path,
) -> Path:
    result_by_id = {case.case_id: case for case in grading_result.cases}
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instructions.append(["Teacher grading review gate"])
    instructions.append(["Run ID", manifest.run_id])
    instructions.append(
        [
            "Required",
            (
                "Review every suggestion against the locked score and rubric. Classify each "
                "disagreement as none, rubric_ambiguity, ocr_unfixed, model_error, or "
                "dataset_error. This workbook does not approve or create final grades."
            ),
        ]
    )
    instructions["A1"].font = Font(bold=True, size=16)
    instructions.column_dimensions["A"].width = 20
    instructions.column_dimensions["B"].width = 110

    sheet = workbook.create_sheet("Grading Review")
    headers = [
        "case_id",
        "primary_category",
        "answer_quality",
        "expected_score",
        "ai_score",
        "max_score",
        "absolute_error",
        "confidence",
        "outcome",
        "rubric_breakdown",
        "review_flags",
        "latency_ms",
        "token_usage",
        "disagreement_reason",
        "teacher_notes",
        "useful_draft",
        "approved_review",
    ]
    sheet.append(headers)
    _style_workbook_header(sheet)
    widths = [12, 24, 18, 16, 14, 14, 16, 14, 30, 65, 45, 16, 30, 24, 45, 18, 20]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row_index, case in enumerate(manifest.cases, start=2):
        result = result_by_id[case.case_id]
        absolute_error = (
            abs(result.ai_score - case.expected_score) if result.ai_score is not None else None
        )
        sheet.append(
            [
                case.case_id,
                case.primary_category.value,
                case.answer_quality.value,
                str(case.expected_score),
                str(result.ai_score) if result.ai_score is not None else None,
                str(case.max_score),
                str(absolute_error) if absolute_error is not None else None,
                str(result.confidence) if result.confidence is not None else None,
                result.outcome,
                json.dumps(_jsonable(result.rubric_breakdown), ensure_ascii=False),
                "\n".join(result.review_flags),
                result.latency_ms,
                json.dumps(_jsonable(result.token_usage), ensure_ascii=False),
                None,
                None,
                None,
                None,
            ]
        )
        sheet.row_dimensions[row_index].height = 125
        for cell in sheet[row_index]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"
    path = run_dir / "grading_review.xlsx"
    workbook.save(path)
    return path


def _safe_dispatch_results(
    manifest: LocalCuratedEvaluationManifest,
    confirmation_lock: OcrConfirmationLock,
    ocr_result: OcrRunResult,
    session_factory: Any,
) -> tuple[GradingRunResult, dict[str, Any]]:
    from sqlalchemy import func, select
    from sqlalchemy.orm import selectinload

    from app.models import (
        AuditLog,
        FinalGrade,
        GradeSuggestion,
        GradingDispatchItem,
        GradingDispatchRun,
        GradingQueueItem,
    )
    from app.schemas import CohortDispatchRequest
    from app.services.grading_dispatch_service import GradingDispatchService
    from app.services.grading_queue_service import GradingQueueService
    from app.worker.jobs import run_grading_dispatch_job

    confirmation_by_id = {case.case_id: case for case in confirmation_lock.cases}
    region_to_case = {case.answer_region_id: case.case_id for case in confirmation_lock.cases}
    call_limits = {"A": 3, "B": 4, "C": 3, "D": 4, "E": 4}
    provider_name = (
        "llama_cpp_qwen"
        if manifest.expected_qwen_model == "qwen3.6-35b-a3b-q4km"
        else "llama_cpp_qwen38"
    )
    dispatch_ids: list[int] = []
    first_call_at: datetime | None = None
    with session_factory() as db:
        queue_run, refused = GradingQueueService(db).build_queue_run(
            assessment_id=ocr_result.assessment_id,
            created_by_teacher_id=ocr_result.owner_teacher_id,
        )
        refused_region_ids = {
            int(item["answer_region_id"])
            for item in refused
            if item.get("answer_region_id") is not None
        }
        expected_blank_ids = {
            case.answer_region_id
            for case in confirmation_lock.cases
            if case.evidence_status == "blank"
        }
        if queue_run.queued_item_count != 18 or queue_run.refused_item_count != 2:
            raise LocalCuratedEvaluationError("Queue must contain 18 fresh and 2 refused cases")
        if refused_region_ids != expected_blank_ids:
            raise LocalCuratedEvaluationError("Only the two blank cases may be queue-refused")

        for pack_id in "ABCDE":
            request = CohortDispatchRequest(
                queue_run_id=queue_run.id,
                grading_run_id=ocr_result.grading_run_id,
                provider=provider_name,
                expected_model=manifest.expected_qwen_model,
                call_limit=call_limits[pack_id],
                draft_only_confirmed=True,
            )
            service = GradingDispatchService(db)
            preflight = service.preflight(
                assessment_id=ocr_result.assessment_id,
                question_id=ocr_result.question_ids[pack_id],
                teacher_id=ocr_result.owner_teacher_id,
                request=request,
            )
            if preflight["selected_call_count"] != call_limits[pack_id]:
                raise LocalCuratedEvaluationError(
                    f"Dispatch preflight selected the wrong number of {pack_id} cases"
                )
            run = service.create_dispatch(
                assessment_id=ocr_result.assessment_id,
                question_id=ocr_result.question_ids[pack_id],
                teacher_id=ocr_result.owner_teacher_id,
                request=request,
            )
            dispatch_ids.append(run.id)
            if first_call_at is None:
                first_call_at = datetime.now(UTC)
            run_grading_dispatch_job(run.id)
            db.expire_all()
            finished = db.get(
                GradingDispatchRun,
                run.id,
                options=(selectinload(GradingDispatchRun.items),),
            )
            if (
                finished is None
                or finished.status != "completed"
                or finished.succeeded_count != call_limits[pack_id]
                or finished.calls_started != call_limits[pack_id]
                or finished.failed_count
                or finished.uncertain_count
            ):
                raise LocalCuratedEvaluationError(
                    f"Dispatch {pack_id} did not complete safely; execution stopped"
                )

        dispatch_runs = list(
            db.scalars(
                select(GradingDispatchRun)
                .options(selectinload(GradingDispatchRun.items))
                .where(GradingDispatchRun.id.in_(dispatch_ids))
                .order_by(GradingDispatchRun.id)
            ).all()
        )
        calls_started = sum(run.calls_started for run in dispatch_runs)
        if calls_started != QWEN_CALL_LIMIT:
            raise LocalCuratedEvaluationError("Qwen provider-call count is not exactly 18")
        items_by_region: dict[int, GradingDispatchItem] = {
            item.answer_region_id: item for run in dispatch_runs for item in run.items
        }
        suggestions = list(
            db.scalars(
                select(GradeSuggestion)
                .where(GradeSuggestion.answer_region_id.in_(region_to_case))
                .order_by(GradeSuggestion.answer_region_id)
            ).all()
        )
        if len(suggestions) != 18 or len({item.grading_job_id for item in suggestions}) != 18:
            raise LocalCuratedEvaluationError("Expected 18 unique grading suggestions")
        suggestion_by_region = {
            suggestion.answer_region_id: suggestion for suggestion in suggestions
        }
        queue_items = list(
            db.scalars(
                select(GradingQueueItem).where(GradingQueueItem.queue_run_id == queue_run.id)
            ).all()
        )
        queue_by_region = {item.answer_region_id: item for item in queue_items}

        case_results: list[GradingCaseResult] = []
        for case in manifest.cases:
            confirmation = confirmation_by_id[case.case_id]
            suggestion = suggestion_by_region.get(confirmation.answer_region_id)
            dispatch_item = items_by_region.get(confirmation.answer_region_id)
            if confirmation.evidence_status == "blank":
                if suggestion is not None or dispatch_item is not None:
                    raise LocalCuratedEvaluationError("A blank case reached Qwen")
                case_results.append(
                    GradingCaseResult(
                        case_id=case.case_id,
                        answer_region_id=confirmation.answer_region_id,
                        outcome="not_called_blank_safety_gate",
                        max_score=case.max_score,
                        confirmed_text_sha256=confirmation.confirmed_text_sha256,
                    )
                )
                continue
            if suggestion is None or dispatch_item is None:
                raise LocalCuratedEvaluationError(
                    f"Missing grading result for {case.case_id}"
                )
            queue_item = queue_by_region.get(confirmation.answer_region_id)
            if queue_item is None:
                raise LocalCuratedEvaluationError(f"Missing queue item for {case.case_id}")
            queued_text_hash = queue_item.readiness_snapshot_json.get(
                "manual_answer_text_sha256"
            )
            if queued_text_hash != confirmation.confirmed_text_sha256:
                raise LocalCuratedEvaluationError(
                    f"Qwen input evidence hash differs from confirmed text for {case.case_id}"
                )
            raw = dict(suggestion.raw_response_json)
            case_results.append(
                GradingCaseResult(
                    case_id=case.case_id,
                    answer_region_id=confirmation.answer_region_id,
                    outcome="suggested",
                    dispatch_run_id=dispatch_item.dispatch_run_id,
                    dispatch_item_id=dispatch_item.id,
                    grading_job_id=suggestion.grading_job_id,
                    grade_suggestion_id=suggestion.id,
                    ai_score=suggestion.score,
                    max_score=suggestion.max_score,
                    confidence=suggestion.confidence,
                    needs_review=suggestion.needs_review,
                    rubric_breakdown=list(raw.get("rubric_breakdown") or []),
                    review_flags=list(raw.get("review_flags") or []),
                    model_provider=suggestion.model_provider,
                    model_name=suggestion.model_name,
                    prompt_version=suggestion.prompt_version,
                    marking_policy=suggestion.marking_policy,
                    token_usage={
                        "prompt_tokens": raw.get("prompt_tokens"),
                        "completion_tokens": raw.get("completion_tokens"),
                        "total_tokens": raw.get("total_tokens"),
                    },
                    latency_ms=int(raw.get("latency_ms") or 0),
                    cost_estimate=suggestion.cost_estimate,
                    confirmed_text_sha256=confirmation.confirmed_text_sha256,
                )
            )

        required_flags = {
            "image_input_disabled",
            "local_provider",
            "teacher_review_required",
        }
        suggested_results = [item for item in case_results if item.outcome == "suggested"]
        mandatory_flags = all(
            required_flags.issubset(item.review_flags) for item in suggested_results
        )
        provider_exact = all(
            item.model_provider == provider_name
            and item.model_name == manifest.expected_qwen_model
            for item in suggested_results
        )
        needs_review = all(item.needs_review is True for item in suggested_results)
        zero_cost = all((item.cost_estimate or Decimal("0")) == 0 for item in suggested_results)
        one_attempt = all(
            item.attempt_count == 1 and item.status == "succeeded"
            for run in dispatch_runs
            for item in run.items
            if item.status != "skipped"
        )
        pinned_integrity = all(
            item.rubric_id
            == ocr_result.rubric_ids[region_to_case[item.answer_region_id][0]]
            and bool(item.evidence_snapshot_hash)
            and bool(item.rubric_snapshot_hash)
            for run in dispatch_runs
            for item in run.items
            if item.status == "succeeded"
        )
        dispatch_contract_exact = all(
            run.provider == provider_name
            and run.model_name == manifest.expected_qwen_model
            and run.marking_policy == "general"
            and run.maximum_calls == call_limits[region_to_case[run.items[0].answer_region_id][0]]
            and run.draft_only_confirmed
            for run in dispatch_runs
        )
        final_grade_count = int(db.scalar(select(func.count(FinalGrade.id))) or 0)
        audit_payloads = list(db.scalars(select(AuditLog.payload_json)).all())
        audit_text = json.dumps(audit_payloads, ensure_ascii=False, default=str)
        raw_text_private = all(
            not confirmation.confirmed_text
            or confirmation.confirmed_text not in audit_text
            for confirmation in confirmation_lock.cases
        )
        safety_checks = {
            "queue_18_fresh_2_blank_refused": True,
            "exactly_18_qwen_calls": calls_started == 18,
            "zero_retries": one_attempt,
            "zero_fallback_calls": provider_exact,
            "zero_cloud_calls": provider_exact,
            "blank_qwen_calls_zero": all(
                item.outcome == "not_called_blank_safety_gate"
                for item in case_results
                if case_by_id(manifest, item.case_id).answer_quality == AnswerQuality.BLANK
            ),
            "confirmed_text_hashes_match": True,
            "evidence_and_rubric_hashes_pinned": pinned_integrity,
            "dispatch_authorization_contract_exact": dispatch_contract_exact,
            "mandatory_review_flags": mandatory_flags,
            "needs_review_true": needs_review,
            "zero_monetary_cost": zero_cost,
            "no_final_grade": final_grade_count == 0,
            "audit_payload_has_no_raw_answer_text": raw_text_private,
        }
        if not all(safety_checks.values()):
            raise LocalCuratedEvaluationError("A grading safety invariant failed")
        result = GradingRunResult(
            run_id=manifest.run_id,
            first_call_at=first_call_at or datetime.now(UTC),
            completed_at=datetime.now(UTC),
            qwen_call_count=18,
            blank_refusal_count=2,
            dispatch_run_ids=dispatch_ids,
            cases=case_results,
            safety_checks=safety_checks,
        )
        runtime_details = {
            "queue_run_id": queue_run.id,
            "queue_candidate_count": queue_run.total_candidate_packets,
            "queue_ready_count": queue_run.queued_item_count,
            "queue_refused_count": queue_run.refused_item_count,
            "dispatch_call_limits": call_limits,
        }
        return result, runtime_details


def case_by_id(
    manifest: LocalCuratedEvaluationManifest,
    case_id: str,
) -> LocalCuratedCaseDefinition:
    return next(case for case in manifest.cases if case.case_id == case_id)


def _validate_complete_grading_safety_checks(result: GradingRunResult) -> None:
    missing_checks = _REQUIRED_GRADING_SAFETY_CHECKS - set(result.safety_checks)
    if missing_checks or not all(result.safety_checks.values()):
        raise LocalCuratedEvaluationError(
            "Grading result does not contain a complete passing safety record"
        )


def run_grading_stage(
    run_dir: Path,
    *,
    allow_local_qwen: bool,
    max_qwen_calls: int,
    expected_model: str,
    database_url: str,
    local_ai_env: Path | None = None,
) -> GradingRunResult:
    if not allow_local_qwen:
        raise LocalCuratedEvaluationError("Real Qwen grading requires --allow-local-qwen")
    if max_qwen_calls != QWEN_CALL_LIMIT:
        raise LocalCuratedEvaluationError("The Qwen call cap must be exactly 18")
    if expected_model not in SUPPORTED_GRADING_MODELS:
        raise LocalCuratedEvaluationError(
            "Expected Qwen model alias is not one of the supported grading models"
        )
    if current_state(run_dir) != "ocr_confirmed":
        raise LocalCuratedEvaluationError("Qwen grading requires confirmed OCR evidence")
    # Scope authorization is an operator precondition, not an evaluation
    # attempt. Failing it must leave the signed OCR-confirmed run resumable:
    # no model has been selected and no provider call has been made yet.
    scope_manifest = load_manifest(run_dir)
    if os.environ.get(_EVALUATION_SCOPE_ENV) != scope_manifest.run_id:
        raise LocalCuratedEvaluationError(
            "Local curated grading must be explicitly scoped to this run ID; "
            "use the evaluation operator script before enabling its call budget"
        )
    try:
        verify_locked_artifacts(run_dir)
        require_clean_git_worktree()
        manifest = load_manifest(run_dir)
        if expected_model != manifest.expected_qwen_model:
            raise LocalCuratedEvaluationError(
                "Requested grading model does not match the locked evaluation manifest"
            )
        if current_git_commit() != manifest.harness_commit:
            raise LocalCuratedEvaluationError(
                "Current Git commit does not match the harness commit"
            )
        ocr_result = OcrRunResult.model_validate(read_json(run_dir / "ocr_results.json"))
        confirmation_lock = OcrConfirmationLock.model_validate(
            read_json(run_dir / "ocr_confirmation_lock.json")
        )
        settings, session_factory, _storage_root = _configure_runtime(
            run_dir,
            database_url=database_url,
            local_ai_env=local_ai_env,
            require_grading=True,
        )
        if datetime.now(UTC) <= confirmation_lock.signed_at:
            raise LocalCuratedEvaluationError(
                "Grading calls must follow OCR confirmation sign-off"
            )
        active_phase = _activate_model_for_evaluation(
            settings=settings,
            session_factory=session_factory,
            expected_model=manifest.expected_qwen_model,
            holder_kind="evaluation_grading_preflight",
        )
        status_before = _sanitized_status(
            expected_visual_model=manifest.operator_assets.qwen38_vision.model_alias,
            expected_grading_model=manifest.expected_qwen_model,
            require_visual=False,
        )
        gpu_before = _gpu_safety_snapshot(phase=active_phase)
        result, runtime_details = _safe_dispatch_results(
            manifest,
            confirmation_lock,
            ocr_result,
            session_factory,
        )
        if result.first_call_at <= confirmation_lock.signed_at:
            raise LocalCuratedEvaluationError("Qwen calls did not follow OCR confirmation")
        from fastapi import HTTPException

        from app.api.routes.final_grades import get_assessment_review_queue
        from app.api.routes.grading import read_grading_dispatch_run
        from app.core.ownership import get_owned_assessment_or_404
        from app.models import User
        from app.services.final_grade_service import FinalGradeService

        with session_factory() as db:
            owner = db.get(User, ocr_result.owner_teacher_id)
            intruder = db.get(User, ocr_result.intruder_teacher_id)
            if owner is None or intruder is None:
                raise LocalCuratedEvaluationError("Evaluation teachers are missing")
            for dispatch_id in result.dispatch_run_ids:
                owner_read = read_grading_dispatch_run(dispatch_id, db, owner)
                if owner_read.id != dispatch_id:
                    raise LocalCuratedEvaluationError(
                        "Dispatch ownership read returned the wrong run"
                    )
                try:
                    read_grading_dispatch_run(dispatch_id, db, intruder)
                except HTTPException as exc:
                    if exc.status_code != 404:
                        raise LocalCuratedEvaluationError(
                            "Dispatch ownership isolation returned the wrong status"
                        ) from exc
                else:
                    raise LocalCuratedEvaluationError("Dispatch ownership isolation failed")
            owner_review_items = list(
                get_assessment_review_queue(ocr_result.assessment_id, db, owner)
            )
            try:
                get_assessment_review_queue(ocr_result.assessment_id, db, intruder)
            except HTTPException as exc:
                if exc.status_code != 404:
                    raise LocalCuratedEvaluationError(
                        "Review-queue ownership isolation returned the wrong status"
                    ) from exc
            else:
                raise LocalCuratedEvaluationError("Review-queue ownership isolation failed")
            get_owned_assessment_or_404(ocr_result.assessment_id, db, owner)
            export_content = FinalGradeService(db).build_final_grades_workbook(
                ocr_result.assessment_id
            )
            try:
                get_owned_assessment_or_404(ocr_result.assessment_id, db, intruder)
            except HTTPException as exc:
                if exc.status_code != 404:
                    raise LocalCuratedEvaluationError(
                        "Export ownership isolation returned the wrong status"
                    ) from exc
            else:
                raise LocalCuratedEvaluationError("Export ownership isolation failed")
        suggested_review_items = [
            item for item in owner_review_items if item.review_status == "suggested"
        ]
        ungraded_region_ids = {
            item.answer_region.id
            for item in owner_review_items
            if item.review_status == "ungraded"
        }
        expected_blank_region_ids = {
            case.answer_region_id
            for case in confirmation_lock.cases
            if case.evidence_status == "blank"
        }
        if (
            len(owner_review_items) != 20
            or len(suggested_review_items) != 18
            or ungraded_region_ids != expected_blank_region_ids
        ):
            raise LocalCuratedEvaluationError("Review-queue ownership isolation failed")
        export_rows = list(
            load_workbook(io.BytesIO(export_content)).active.iter_rows(values_only=True)
        )
        empty_approved_export = len(export_rows) == 1
        if not empty_approved_export:
            raise LocalCuratedEvaluationError("Unapproved suggestions appeared in export")
        result.safety_checks.update(
            {
                "cross_teacher_dispatch_access_refused": True,
                "cross_teacher_review_access_refused": True,
                "cross_teacher_export_access_refused": True,
                "approved_only_export_has_zero_data_rows": True,
            }
        )
        status_after = _sanitized_status(
            expected_visual_model=manifest.operator_assets.qwen38_vision.model_alias,
            expected_grading_model=manifest.expected_qwen_model,
            require_visual=False,
        )
        gpu_after = _gpu_safety_snapshot(phase=active_phase)
        grading_status_key = "qwen" if active_phase == "Qwen" else "qwen38"
        if (
            not status_before[grading_status_key]["available"]
            or not status_after[grading_status_key]["available"]
        ):
            raise LocalCuratedEvaluationError("The selected Qwen grading model was not healthy")
        if not all(
            (
                gpu_before["active_model_present_in_gpu_compute_clients"],
                gpu_before["other_managed_model_absent_from_gpu_compute_clients"],
                gpu_after["active_model_present_in_gpu_compute_clients"],
                gpu_after["other_managed_model_absent_from_gpu_compute_clients"],
            )
        ):
            raise LocalCuratedEvaluationError("The single local-model GPU slot was not isolated")
        result.safety_checks["qwen38_visual_transcription_reasoning_off"] = all(
            case.reasoning_mode == "off" for case in ocr_result.cases
        )
        result.safety_checks["single_model_gpu_slot_enforced"] = True
        _validate_complete_grading_safety_checks(result)
        result_path = run_dir / "grading_results.json"
        runtime_path = run_dir / "grading_runtime.json"
        write_json(result_path, result)
        write_json(
            runtime_path,
            {
                **runtime_details,
                "service_status_before": status_before,
                "service_status_after": status_after,
                "gpu_safety_before": gpu_before,
                "gpu_safety_after": gpu_after,
                "transport": "direct_host_eval",
                "rq_crash_recovery_validated": False,
            },
        )
        create_grading_review_workbook(manifest, result, run_dir)
        append_state(
            run_dir,
            "grading_completed",
            locked_artifacts={
                "grading_results.json": sha256_file(result_path),
                "grading_runtime.json": sha256_file(runtime_path),
            },
            metadata={
                "qwen_call_count": 18,
                "blank_refusal_count": 2,
                "retry_count": 0,
                "fallback_call_count": 0,
                "cloud_call_count": 0,
                "transport": "direct_host_eval",
            },
        )
        return result
    except Exception as exc:
        _mark_invalid(
            run_dir,
            stage="run_grading",
            detail="Qwen grading stopped on the first failure; no retry is authorized.",
        )
        if isinstance(exc, LocalCuratedEvaluationError):
            raise
        raise LocalCuratedEvaluationError("Qwen grading failed safely") from exc


def read_grading_review_workbook(
    workbook_path: Path,
    manifest: LocalCuratedEvaluationManifest,
    grading_result: GradingRunResult,
) -> list[GradingReviewCase]:
    workbook = load_workbook(workbook_path, data_only=True)
    if "Grading Review" not in workbook.sheetnames:
        raise LocalCuratedEvaluationError("Grading workbook has no Grading Review sheet")
    sheet = workbook["Grading Review"]
    headers = {str(cell.value): index for index, cell in enumerate(sheet[1], start=1)}
    required = {
        "case_id",
        "disagreement_reason",
        "teacher_notes",
        "useful_draft",
        "approved_review",
    }
    if not required.issubset(headers):
        raise LocalCuratedEvaluationError("Grading workbook columns were changed")
    result_by_id = {case.case_id: case for case in grading_result.cases}
    definition_by_id = {case.case_id: case for case in manifest.cases}
    reviews: list[GradingReviewCase] = []
    for row in range(2, sheet.max_row + 1):
        case_id = str(sheet.cell(row, headers["case_id"]).value or "").strip()
        if not case_id:
            continue
        result = result_by_id.get(case_id)
        definition = definition_by_id.get(case_id)
        if result is None or definition is None:
            raise LocalCuratedEvaluationError(f"Unknown grading review case: {case_id}")
        reason = str(sheet.cell(row, headers["disagreement_reason"]).value or "").strip()
        if reason not in _REVIEW_REASONS:
            raise LocalCuratedEvaluationError(
                f"{case_id}: disagreement_reason must use the documented vocabulary"
            )
        useful_draft = _parse_bool(
            sheet.cell(row, headers["useful_draft"]).value,
            field="useful_draft",
            case_id=case_id,
        )
        approved = _parse_bool(
            sheet.cell(row, headers["approved_review"]).value,
            field="approved_review",
            case_id=case_id,
        )
        if not approved:
            raise LocalCuratedEvaluationError(f"{case_id}: grading review is not approved")
        if result.outcome == "not_called_blank_safety_gate":
            if reason != "none" or useful_draft:
                raise LocalCuratedEvaluationError(
                    f"{case_id}: blank safety refusals require reason=none and useful_draft=no"
                )
        else:
            exact = result.ai_score == definition.expected_score
            if exact and reason != "none":
                raise LocalCuratedEvaluationError(
                    f"{case_id}: exact scores require disagreement_reason=none"
                )
            if not exact and reason == "none":
                raise LocalCuratedEvaluationError(
                    f"{case_id}: score disagreements require a classified reason"
                )
        reviews.append(
            GradingReviewCase(
                case_id=case_id,
                disagreement_reason=reason,
                teacher_notes=str(
                    sheet.cell(row, headers["teacher_notes"]).value or ""
                ).strip(),
                useful_draft=useful_draft,
                approved_review=True,
            )
        )
    if [case.case_id for case in reviews] != [case.case_id for case in manifest.cases]:
        raise LocalCuratedEvaluationError("All 20 grading cases must be reviewed in order")
    return reviews


def lock_grading_review(
    run_dir: Path,
    *,
    reviewer_id: str,
    confirm_teacher_signoff: bool,
) -> ReviewLock:
    if not confirm_teacher_signoff:
        raise LocalCuratedEvaluationError("Explicit grading-review sign-off is required")
    if current_state(run_dir) != "grading_completed":
        raise LocalCuratedEvaluationError("Review locking requires completed grading")
    verify_locked_artifacts(run_dir)
    manifest = load_manifest(run_dir)
    grading_result = GradingRunResult.model_validate(
        read_json(run_dir / "grading_results.json")
    )
    _validate_complete_grading_safety_checks(grading_result)
    workbook_path = run_dir / "grading_review.xlsx"
    reviews = read_grading_review_workbook(workbook_path, manifest, grading_result)
    lock = ReviewLock(
        run_id=manifest.run_id,
        reviewer_id=reviewer_id.strip(),
        signed_at=datetime.now(UTC),
        grading_results_sha256=sha256_file(run_dir / "grading_results.json"),
        workbook_sha256=sha256_file(workbook_path),
        cases=reviews,
    )
    lock_path = run_dir / "review_lock.json"
    write_json(lock_path, lock)
    append_state(
        run_dir,
        "review_completed",
        locked_artifacts={
            "grading_review.xlsx": sha256_file(workbook_path),
            "review_lock.json": sha256_file(lock_path),
        },
        metadata={
            "reviewed_case_count": 20,
            "production_suggestions_approved": 0,
            "final_grades_created": 0,
        },
    )
    return lock


def _report_markdown(report: dict[str, Any]) -> str:
    verdict = report["verdict"]
    lines = [
        f"# Local Curated Evaluation: {report['run_id']}",
        "",
        f"Verdict: **{verdict}**",
        "",
        (
            "This evaluation uses synthetic, teacher-reviewed material. Its result is not a "
            "production accuracy claim. PASS permits only a supervised Custom Controlled pilot."
        ),
        "",
        "## Safety and process",
        "",
    ]
    for key, passed in sorted(report["process_checks"].items()):
        lines.append(f"- {key}: {'PASS' if passed else 'FAIL'}")
    lines.extend(["", "## OCR metrics", ""])
    for key, value in report["ocr_metrics"].items():
        if key != "cases":
            lines.append(f"- {key}: {value}")
    lines.extend(["", "## Grading metrics", ""])
    for key, value in report["grading_metrics"].items():
        if key != "cases":
            lines.append(f"- {key}: {value}")
    lines.extend(["", "## Verdict reasons", ""])
    if report["verdict_reasons"]:
        lines.extend(f"- {reason}" for reason in report["verdict_reasons"])
    else:
        lines.append("- All required checks and thresholds passed.")
    lines.extend(
        [
            "",
            "## Runtime",
            "",
            (
                "- Text grading provider/model: "
                f"{report['models']['grading_provider']} / {report['models']['grading_alias']}"
            ),
            (
                "- Visual transcription provider/model/mode: "
                f"llama_cpp_qwen38 / {report['models']['visual_alias']} / "
                "reasoning disabled"
            ),
            (
                "- Text-grading calls: "
                f"{report['call_counts']['grading']}; visual-transcription calls: "
                f"{report['call_counts']['visual_transcription']}; retries: "
                f"{report['call_counts']['retries']}; fallback/cloud calls: 0"
            ),
            "- Dispatch transport: direct_host_eval (RQ crash recovery was not evaluated)",
            "",
        ]
    )
    return "\n".join(lines)


def _generate_ocr_quality_no_go_report(run_dir: Path) -> dict[str, Any]:
    """Report an honestly failed visual-OCR quality gate without grading."""

    verify_locked_artifacts(run_dir)
    manifest = load_manifest(run_dir)
    ground_truth = GroundTruthLock.model_validate(read_json(run_dir / "ground_truth_lock.json"))
    ocr_result = OcrRunResult.model_validate(read_json(run_dir / "ocr_results.json"))
    quality_lock = OcrQualityNoGoLock.model_validate(
        read_json(run_dir / "ocr_quality_no_go_lock.json")
    )
    if quality_lock.ocr_results_sha256 != sha256_file(run_dir / "ocr_results.json"):
        raise LocalCuratedEvaluationError("OCR quality no-go lock does not match OCR results")
    if (
        (run_dir / "grading_results.json").exists()
        or (run_dir / "ocr_confirmation_lock.json").exists()
    ):
        raise LocalCuratedEvaluationError(
            "OCR quality no-go cannot report after an OCR confirmation or grading result"
        )
    runtime = read_json(run_dir / "ocr_runtime.json")
    ocr_metrics = calculate_ocr_metrics(manifest, ocr_result.cases)
    process_checks = {
        "twenty_teacher_signed_cases": len(ground_truth.cases) == OCR_CALL_LIMIT,
        "ground_truth_signed_before_ocr": ground_truth.signed_at < ocr_result.first_call_at,
        "exactly_twenty_visual_transcription_calls": ocr_result.call_count == OCR_CALL_LIMIT,
        "zero_visual_retries": ocr_result.retry_count == 0,
        "exact_qwen38_visual_provider_model": all(
            case.provider == "llama_cpp_qwen38"
            and case.model == manifest.operator_assets.qwen38_vision.model_alias
            and case.reasoning_mode == "off"
            for case in ocr_result.cases
        ),
        "teacher_recorded_ocr_quality_rejection": bool(quality_lock.rejected_cases),
        "zero_grading_calls_after_rejection": True,
        "no_final_grade": True,
        "no_unconfirmed_text_used_for_grading": True,
        "zero_cloud_calls": True,
        "zero_fallback_calls": True,
        "single_model_gpu_slot_enforced": bool(runtime.get("gpu_safety_before"))
        and bool(runtime.get("gpu_safety_after")),
    }
    report = {
        "schema_version": SCHEMA_VERSION,
        "run_id": manifest.run_id,
        "generated_at": datetime.now(UTC),
        "verdict": EvaluationVerdict.NO_GO_QUALITY.value,
        "verdict_reasons": [
            "Teacher rejected Qwen3.8 visual transcription: "
            f"{case.case_id} ({case.reason})"
            for case in quality_lock.rejected_cases
        ],
        "integration_commit": manifest.integration_commit,
        "harness_commit": manifest.harness_commit,
        "protocol_version": manifest.protocol_version,
        "dataset": {
            "case_count": OCR_CALL_LIMIT,
            "synthetic": True,
            "teacher_signed": True,
            "primary_category_distribution": dict(
                Counter(case.primary_category.value for case in manifest.cases)
            ),
        },
        "models": {
            "grading_provider": None,
            "grading_alias": None,
            "grading_model_sha256": None,
            "visual_alias": manifest.operator_assets.qwen38_vision.model_alias,
            "visual_model_sha256": manifest.operator_assets.qwen38_vision.model_sha256,
            "visual_mmproj_sha256": manifest.operator_assets.qwen38_vision.mmproj_sha256,
            "visual_reasoning_mode": runtime["visual_reasoning_mode"],
            "llama_cpp_build": manifest.operator_assets.qwen38_vision.build,
        },
        "call_counts": {
            "visual_transcription": OCR_CALL_LIMIT,
            "grading": 0,
            "retries": 0,
        },
        "process_checks": process_checks,
        "ocr_metrics": ocr_metrics,
        "grading_metrics": {
            "not_run": True,
            "reason": "Teacher-rejected visual transcription blocked grading safely.",
            "suggested_case_count": 0,
        },
        "review_summary": {
            "ocr_quality_rejections": len(quality_lock.rejected_cases),
        },
        "limitations": [
            "A rejected OCR reading is a valid quality failure, not an integrity failure.",
            "No student answer was sent to a text-grading model after the rejection.",
            "A fresh run with new source images or a changed validated model is required.",
        ],
    }
    report_path = run_dir / "report.json"
    markdown_path = run_dir / "report.md"
    write_json(report_path, report)
    markdown_path.write_text(_report_markdown(_jsonable(report)), encoding="utf-8")
    append_state(
        run_dir,
        "reported",
        locked_artifacts={
            "report.json": sha256_file(report_path),
            "report.md": sha256_file(markdown_path),
        },
        metadata={"verdict": EvaluationVerdict.NO_GO_QUALITY.value},
    )
    return report


def generate_report(run_dir: Path) -> dict[str, Any]:
    state = current_state(run_dir)
    if state == "ocr_quality_no_go":
        return _generate_ocr_quality_no_go_report(run_dir)
    if state != "review_completed":
        raise LocalCuratedEvaluationError("Report generation requires signed teacher review")
    verify_locked_artifacts(run_dir)
    manifest = load_manifest(run_dir)
    ground_truth = GroundTruthLock.model_validate(read_json(run_dir / "ground_truth_lock.json"))
    ocr_result = OcrRunResult.model_validate(read_json(run_dir / "ocr_results.json"))
    if ocr_result.evidence_origin != "live_qwen38_visual":
        raise LocalCuratedEvaluationError(
            "A grading bake-off replay is not an independent visual-OCR evaluation; "
            "use the grading-model bake-off comparison report instead"
        )
    confirmations = OcrConfirmationLock.model_validate(
        read_json(run_dir / "ocr_confirmation_lock.json")
    )
    grading_result = GradingRunResult.model_validate(
        read_json(run_dir / "grading_results.json")
    )
    review = ReviewLock.model_validate(read_json(run_dir / "review_lock.json"))
    environment = read_json(run_dir / "ocr_runtime.json")
    ocr_metrics = calculate_ocr_metrics(manifest, ocr_result.cases, confirmations.cases)
    grading_metrics = calculate_grading_metrics(manifest, grading_result.cases)

    truth_by_id = {case.case_id: case for case in ground_truth.cases}
    confirmation_by_id = {case.case_id: case for case in confirmations.cases}
    grading_by_id = {case.case_id: case for case in grading_result.cases}
    process_checks = dict(grading_result.safety_checks)
    process_checks.update(
        {
            "twenty_teacher_signed_cases": len(ground_truth.cases) == 20,
            "ground_truth_signed_before_ocr": ground_truth.signed_at < ocr_result.first_call_at,
            "exactly_twenty_ocr_calls": ocr_result.call_count == 20,
            "zero_ocr_retries": ocr_result.retry_count == 0,
            "twenty_teacher_confirmed_ocr_cases": len(confirmations.cases) == 20,
            "ocr_confirmed_before_qwen": confirmations.signed_at < grading_result.first_call_at,
            "exactly_eighteen_qwen_calls": grading_result.qwen_call_count == 18,
            "two_blank_policy_refusals": grading_result.blank_refusal_count == 2,
            "zero_provider_retries": grading_result.retry_count == 0,
            "zero_fallback_calls": grading_result.fallback_call_count == 0,
            "zero_cloud_calls": grading_result.cloud_call_count == 0,
            "teacher_review_after_grading": review.signed_at > grading_result.completed_at,
            "all_confirmed_text_matches_ground_truth": all(
                normalize_text(confirmation_by_id[case_id].confirmed_text)
                == normalize_text(truth.teacher_transcription)
                for case_id, truth in truth_by_id.items()
            ),
            "all_qwen_hashes_match_confirmed_text": all(
                grading_by_id[case_id].confirmed_text_sha256
                == confirmation.confirmed_text_sha256
                for case_id, confirmation in confirmation_by_id.items()
            ),
            "twenty_signed_reviews": len(review.cases) == 20,
            "no_dataset_errors_reported": all(
                case.disagreement_reason != "dataset_error" for case in review.cases
            ),
            "no_unfixed_ocr_errors_reported": all(
                case.disagreement_reason != "ocr_unfixed" for case in review.cases
            ),
        }
    )
    verdict, reasons = evaluate_verdict(
        manifest,
        process_checks=process_checks,
        ocr_metrics=ocr_metrics,
        grading_metrics=grading_metrics,
    )
    report = {
        "schema_version": 2,
        "run_id": manifest.run_id,
        "generated_at": datetime.now(UTC),
        "verdict": verdict.value,
        "verdict_reasons": reasons,
        "integration_commit": manifest.integration_commit,
        "harness_commit": manifest.harness_commit,
        "protocol_version": manifest.protocol_version,
        "dataset": {
            "case_count": 20,
            "synthetic": True,
            "teacher_signed": True,
            "primary_category_distribution": dict(
                Counter(case.primary_category.value for case in manifest.cases)
            ),
        },
        "models": {
            "grading_provider": (
                "llama_cpp_qwen"
                if manifest.expected_qwen_model == "qwen3.6-35b-a3b-q4km"
                else "llama_cpp_qwen38"
            ),
            "grading_alias": manifest.expected_qwen_model,
            "grading_model_sha256": manifest.operator_assets.llama_cpp.model_sha256,
            "visual_alias": manifest.operator_assets.qwen38_vision.model_alias,
            "visual_model_sha256": manifest.operator_assets.qwen38_vision.model_sha256,
            "visual_mmproj_sha256": manifest.operator_assets.qwen38_vision.mmproj_sha256,
            "visual_reasoning_mode": environment["visual_reasoning_mode"],
            "llama_cpp_build": manifest.operator_assets.qwen38_vision.build,
        },
        "call_counts": {
            "visual_transcription": OCR_CALL_LIMIT,
            "grading": QWEN_CALL_LIMIT,
            "retries": 0,
        },
        "process_checks": process_checks,
        "ocr_metrics": ocr_metrics,
        "grading_metrics": grading_metrics,
        "review_summary": dict(
            Counter(case.disagreement_reason for case in review.cases)
        ),
        "limitations": [
            "Synthetic cases do not establish production accuracy.",
            "RQ crash recovery was not evaluated because dispatch used direct_host_eval.",
            "PASS permits only a supervised Custom Controlled pilot.",
        ],
    }
    report_path = run_dir / "report.json"
    markdown_path = run_dir / "report.md"
    write_json(report_path, report)
    markdown_path.write_text(_report_markdown(_jsonable(report)), encoding="utf-8")
    terminal_state = "invalid" if verdict == EvaluationVerdict.INVALID_RUN else "reported"
    append_state(
        run_dir,
        terminal_state,
        locked_artifacts={
            "report.json": sha256_file(report_path),
            "report.md": sha256_file(markdown_path),
        },
        metadata={"verdict": verdict.value},
    )
    return report


def _edit_distance(left: list[str] | str, right: list[str] | str) -> int:
    if len(left) < len(right):
        left, right = right, left
    previous = list(range(len(right) + 1))
    for index, left_item in enumerate(left, start=1):
        current = [index]
        for right_index, right_item in enumerate(right, start=1):
            current.append(
                min(
                    current[-1] + 1,
                    previous[right_index] + 1,
                    previous[right_index - 1] + (left_item != right_item),
                )
            )
        previous = current
    return previous[-1]


def character_error_rate(reference: str, candidate: str) -> Decimal:
    normalized_reference = normalize_text(reference)
    normalized_candidate = normalize_text(candidate)
    if not normalized_reference:
        return Decimal("0") if not normalized_candidate else Decimal("1")
    return Decimal(_edit_distance(normalized_reference, normalized_candidate)) / Decimal(
        len(normalized_reference)
    )


def word_error_rate(reference: str, candidate: str) -> Decimal:
    reference_words = normalize_text(reference).split()
    candidate_words = normalize_text(candidate).split()
    if not reference_words:
        return Decimal("0") if not candidate_words else Decimal("1")
    return Decimal(_edit_distance(reference_words, candidate_words)) / Decimal(
        len(reference_words)
    )


def _semantic_tokens(value: str) -> set[str]:
    normalized = normalize_text(value).lower()
    normalized = normalized.replace("²", "^2").replace("⁄", "/")
    tokens = set(
        re.findall(r"[a-z]+\^?\d*|\d+(?:\.\d+)?(?:/\d+)?|[+\-=/|]", normalized)
    )
    for token in list(tokens):
        match = re.fullmatch(r"([a-z]+)\^(\d+)", token)
        if match is None or len(match.group(1)) <= 1:
            continue
        variable_product, exponent = match.groups()
        tokens.update(variable_product[:-1])
        tokens.add(f"{variable_product[-1]}^{exponent}")
    return tokens


def critical_token_recall(reference_tokens: list[str], candidate: str) -> Decimal:
    if not reference_tokens:
        return Decimal("1")
    candidate_tokens = _semantic_tokens(candidate)
    expected: set[str] = set()
    for token in reference_tokens:
        expected.update(_semantic_tokens(token))
    if not expected:
        return Decimal("1")
    return Decimal(len(expected & candidate_tokens)) / Decimal(len(expected))


def _block_order_valid(blocks: list[dict[str, Any]]) -> bool:
    keys: list[tuple[int, int]] = []
    for block in blocks:
        page = block.get("page")
        order = block.get("order")
        if not isinstance(page, int) or page < 1 or not isinstance(order, int) or order < 1:
            return False
        keys.append((page, order))
    return keys == sorted(keys) and len(keys) == len(set(keys))


def severe_error_threshold(max_score: Decimal) -> Decimal:
    return min(Decimal("2"), max_score * Decimal("0.40"))


def percentile(values: list[int], percentile_value: Decimal) -> Decimal:
    if not values:
        return Decimal("0")
    ordered = sorted(values)
    position = (Decimal(len(ordered) - 1) * percentile_value)
    lower = int(position)
    upper = min(lower + 1, len(ordered) - 1)
    fraction = position - Decimal(lower)
    return Decimal(ordered[lower]) + (Decimal(ordered[upper] - ordered[lower]) * fraction)


def calculate_ocr_metrics(
    manifest: LocalCuratedEvaluationManifest,
    results: list[OcrCaseResult],
    confirmations: list[OcrConfirmationCase] | None = None,
) -> dict[str, Any]:
    result_by_id = {result.case_id: result for result in results}
    confirmation_by_id = {
        confirmation.case_id: confirmation for confirmation in confirmations or []
    }
    rows: list[dict[str, Any]] = []
    for case in manifest.cases:
        result = result_by_id[case.case_id]
        normalized_reference = normalize_text(case.authored_transcription)
        normalized_candidate = normalize_text(result.draft_text)
        reference_words = normalized_reference.split()
        candidate_words = normalized_candidate.split()
        cer = character_error_rate(case.authored_transcription, result.draft_text)
        wer = word_error_rate(case.authored_transcription, result.draft_text)
        token_recall = critical_token_recall(case.critical_tokens, result.draft_text)
        confirmation = confirmation_by_id.get(case.case_id)
        edit_burden = None
        if confirmation is not None:
            edit_burden = character_error_rate(confirmation.confirmed_text, result.draft_text)
        rows.append(
            {
                "case_id": case.case_id,
                "primary_category": case.primary_category.value,
                "answer_quality": case.answer_quality.value,
                "cer": cer,
                "wer": wer,
                "character_edits": _edit_distance(
                    normalized_reference,
                    normalized_candidate,
                ),
                "reference_characters": len(normalized_reference),
                "word_edits": _edit_distance(reference_words, candidate_words),
                "reference_words": len(reference_words),
                "critical_token_recall": token_recall,
                "edit_burden": edit_burden,
                "latency_ms": result.latency_ms,
                "warning_count": len(result.warnings),
                "block_order_valid": _block_order_valid(result.blocks),
                "blank_semantic_hallucination": (
                    case.answer_quality == AnswerQuality.BLANK
                    and bool(_semantic_tokens(result.draft_text))
                ),
            }
        )
    nonblank = [row for row in rows if row["answer_quality"] != AnswerQuality.BLANK.value]
    clean = [
        row
        for row in rows
        if next(case for case in manifest.cases if case.case_id == row["case_id"]).clean_typed
    ]
    handwriting = [
        row
        for row in rows
        if row["primary_category"] == PrimaryCategory.DIFFICULT_HANDWRITING.value
    ]
    formula = [
        row for row in rows if row["primary_category"] == PrimaryCategory.FORMULA_HEAVY.value
    ]
    latencies = [int(row["latency_ms"]) for row in rows]
    nonblank_reference_characters = sum(
        int(row["reference_characters"]) for row in nonblank
    )
    nonblank_reference_words = sum(int(row["reference_words"]) for row in nonblank)
    clean_reference_characters = sum(int(row["reference_characters"]) for row in clean)
    return {
        "case_count": len(rows),
        "overall_nonblank_cer": Decimal(
            sum(int(row["character_edits"]) for row in nonblank)
        )
        / Decimal(nonblank_reference_characters),
        "overall_nonblank_macro_cer": sum(
            (row["cer"] for row in nonblank), Decimal("0")
        )
        / Decimal(len(nonblank)),
        "overall_nonblank_wer": Decimal(sum(int(row["word_edits"]) for row in nonblank))
        / Decimal(nonblank_reference_words),
        "clean_typed_cer": Decimal(sum(int(row["character_edits"]) for row in clean))
        / Decimal(clean_reference_characters),
        "handwriting_mean_cer": sum((row["cer"] for row in handwriting), Decimal("0"))
        / Decimal(len(handwriting)),
        "handwriting_max_cer": max(row["cer"] for row in handwriting),
        "formula_critical_token_recall": min(
            row["critical_token_recall"] for row in formula
        ),
        "blank_semantic_hallucination_count": sum(
            1 for row in rows if row["blank_semantic_hallucination"]
        ),
        "mean_edit_burden": (
            sum(
                (row["edit_burden"] for row in rows if row["edit_burden"] is not None),
                Decimal("0"),
            )
            / Decimal(sum(1 for row in rows if row["edit_burden"] is not None))
            if any(row["edit_burden"] is not None for row in rows)
            else None
        ),
        "warning_count": sum(int(row["warning_count"]) for row in rows),
        "block_order_issue_count": sum(
            1 for row in rows if not row["block_order_valid"]
        ),
        "median_latency_ms": Decimal(str(median(latencies))),
        "p95_latency_ms": percentile(latencies, Decimal("0.95")),
        "cases": rows,
    }


def calculate_grading_metrics(
    manifest: LocalCuratedEvaluationManifest,
    results: list[GradingCaseResult],
) -> dict[str, Any]:
    case_by_id = {case.case_id: case for case in manifest.cases}
    suggested = [result for result in results if result.outcome == "suggested"]
    rows: list[dict[str, Any]] = []
    for result in suggested:
        case = case_by_id[result.case_id]
        if result.ai_score is None or result.confidence is None:
            raise LocalCuratedEvaluationError("Suggested grading result lacks score or confidence")
        absolute_error = abs(result.ai_score - case.expected_score)
        normalized_error = absolute_error / case.max_score
        severe = absolute_error >= severe_error_threshold(case.max_score)
        rows.append(
            {
                "case_id": case.case_id,
                "primary_category": case.primary_category.value,
                "answer_quality": case.answer_quality.value,
                "expected_score": case.expected_score,
                "ai_score": result.ai_score,
                "max_score": case.max_score,
                "confidence": result.confidence,
                "absolute_error": absolute_error,
                "normalized_error": normalized_error,
                "exact": absolute_error == 0,
                "within_one": absolute_error <= 1,
                "severe": severe,
                "severe_false_confident": severe and result.confidence >= Decimal("0.80"),
            }
        )
    if len(rows) != 18:
        raise LocalCuratedEvaluationError("Grading metrics require exactly 18 suggestions")
    total = Decimal(len(rows))
    formula_and_multistep = [
        row
        for row in rows
        if row["primary_category"]
        in {PrimaryCategory.FORMULA_HEAVY.value, PrimaryCategory.MULTI_STEP.value}
    ]
    irrelevant = [row for row in rows if row["answer_quality"] == AnswerQuality.IRRELEVANT.value]
    wrong = [row for row in rows if row["answer_quality"] == AnswerQuality.WRONG.value]
    zero_reference_overscores = [
        row for row in rows if row["expected_score"] == 0 and row["ai_score"] > 0
    ]
    return {
        "suggested_case_count": len(rows),
        "exact_count": sum(1 for row in rows if row["exact"]),
        "within_one_count": sum(1 for row in rows if row["within_one"]),
        "mean_absolute_error": sum((row["absolute_error"] for row in rows), Decimal("0"))
        / total,
        "mean_normalized_absolute_error": sum(
            (row["normalized_error"] for row in rows), Decimal("0")
        )
        / total,
        "severe_false_confident_count": sum(
            1 for row in rows if row["severe_false_confident"]
        ),
        "severe_low_confidence_count": sum(
            1 for row in rows if row["severe"] and not row["severe_false_confident"]
        ),
        "formula_multistep_within_one": all(row["within_one"] for row in formula_and_multistep),
        "irrelevant_over_limit_count": sum(
            1 for row in irrelevant if row["ai_score"] > Decimal("0.50")
        ),
        "wrong_over_half_count": sum(
            1 for row in wrong if row["ai_score"] > (row["max_score"] / 2)
        ),
        "zero_reference_overscore_count": len(zero_reference_overscores),
        "cases": rows,
    }


def evaluate_verdict(
    manifest: LocalCuratedEvaluationManifest,
    *,
    process_checks: dict[str, bool],
    ocr_metrics: dict[str, Any],
    grading_metrics: dict[str, Any],
) -> tuple[EvaluationVerdict, list[str]]:
    reasons: list[str] = []
    failed_process = sorted(key for key, passed in process_checks.items() if not passed)
    if failed_process:
        return (
            EvaluationVerdict.INVALID_RUN,
            [f"Process or safety check failed: {key}" for key in failed_process],
        )
    thresholds = manifest.thresholds
    quality_checks = {
        "overall_nonblank_cer": (
            ocr_metrics["overall_nonblank_cer"] <= thresholds.overall_nonblank_cer_max
        ),
        "clean_typed_cer": ocr_metrics["clean_typed_cer"] <= thresholds.clean_typed_cer_max,
        "handwriting_mean_cer": (
            ocr_metrics["handwriting_mean_cer"] <= thresholds.handwriting_mean_cer_max
        ),
        "handwriting_max_cer": (
            ocr_metrics["handwriting_max_cer"] <= thresholds.handwriting_case_cer_max
        ),
        "formula_token_recall": (
            ocr_metrics["formula_critical_token_recall"]
            >= thresholds.formula_critical_token_recall_min
        ),
        "blank_hallucination": ocr_metrics["blank_semantic_hallucination_count"] == 0,
        "grading_exact": grading_metrics["exact_count"] >= thresholds.grading_exact_min,
        "grading_within_one": (
            grading_metrics["within_one_count"] >= thresholds.grading_within_one_min
        ),
        "grading_mae": grading_metrics["mean_absolute_error"] <= thresholds.grading_mae_max,
        "grading_normalized_mae": (
            grading_metrics["mean_normalized_absolute_error"]
            <= thresholds.grading_normalized_mae_max
        ),
        "formula_multistep_within_one": grading_metrics["formula_multistep_within_one"],
        "severe_false_confident": grading_metrics["severe_false_confident_count"] == 0,
        "severe_low_confidence": (
            grading_metrics["severe_low_confidence_count"]
            <= thresholds.severe_low_confidence_max
        ),
        "irrelevant_over_limit": grading_metrics["irrelevant_over_limit_count"] == 0,
        "wrong_over_half": grading_metrics["wrong_over_half_count"] == 0,
        "repeated_zero_reference_overscore": (
            grading_metrics["zero_reference_overscore_count"] < 2
        ),
    }
    for key, passed in quality_checks.items():
        if not passed:
            reasons.append(f"Quality threshold failed: {key}")
    if reasons:
        return EvaluationVerdict.NO_GO_QUALITY, reasons
    return EvaluationVerdict.PASS, []


def _resolve_run_dir(root: Path, run_id: str) -> Path:
    if not _RUN_ID_PATTERN.fullmatch(run_id):
        raise LocalCuratedEvaluationError("Invalid run ID")
    resolved_root = root.resolve()
    run_dir = (resolved_root / run_id).resolve()
    if run_dir.parent != resolved_root:
        raise LocalCuratedEvaluationError("Evaluation run path escaped the output root")
    if not run_dir.is_dir():
        raise LocalCuratedEvaluationError("Evaluation run directory does not exist")
    return run_dir


def _default_run_id() -> str:
    return datetime.now(UTC).strftime("lc_%Y%m%dt%H%M%Sz").lower()


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Run the 20-case local curated evaluation")
    parser.add_argument(
        "--root",
        type=Path,
        default=default_evaluation_root(),
        help="Ignored local evaluation artifact root.",
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    prepare_parser = subparsers.add_parser("prepare")
    prepare_parser.add_argument("--run-id", default=_default_run_id())
    prepare_parser.add_argument("--seed", type=int, default=DEFAULT_SEED)
    prepare_parser.add_argument("--integration-commit", required=True)
    prepare_parser.add_argument("--harness-commit", default=None)
    prepare_parser.add_argument(
        "--grading-model",
        required=True,
        choices=SUPPORTED_GRADING_MODELS,
        help="Grading model alias this run pins; recorded in the manifest.",
    )

    lock_parser = subparsers.add_parser("lock-ground-truth")
    lock_parser.add_argument("--run-id", required=True)
    lock_parser.add_argument("--reviewer-id", required=True)
    lock_parser.add_argument("--confirm-teacher-signoff", action="store_true")

    ocr_parser = subparsers.add_parser("run-ocr")
    ocr_parser.add_argument("--run-id", required=True)
    ocr_parser.add_argument("--allow-local-ocr", action="store_true")
    ocr_parser.add_argument("--max-ocr-calls", type=int, default=OCR_CALL_LIMIT)
    ocr_parser.add_argument("--database-url", default=os.environ.get("DATABASE_URL", ""))
    ocr_parser.add_argument("--local-ai-env", type=Path, default=None)

    confirm_parser = subparsers.add_parser("lock-ocr-confirmations")
    confirm_parser.add_argument("--run-id", required=True)
    confirm_parser.add_argument("--reviewer-id", required=True)
    confirm_parser.add_argument("--confirm-teacher-signoff", action="store_true")
    confirm_parser.add_argument(
        "--database-url", default=os.environ.get("DATABASE_URL", "")
    )
    confirm_parser.add_argument("--local-ai-env", type=Path, default=None)

    no_go_parser = subparsers.add_parser("record-ocr-no-go")
    no_go_parser.add_argument("--run-id", required=True)
    no_go_parser.add_argument("--reviewer-id", required=True)
    no_go_parser.add_argument("--confirm-teacher-signoff", action="store_true")
    no_go_parser.add_argument(
        "--database-url", default=os.environ.get("DATABASE_URL", "")
    )
    no_go_parser.add_argument("--local-ai-env", type=Path, default=None)

    grading_parser = subparsers.add_parser("run-grading")
    grading_parser.add_argument("--run-id", required=True)
    grading_parser.add_argument("--allow-local-qwen", action="store_true")
    grading_parser.add_argument("--max-qwen-calls", type=int, default=QWEN_CALL_LIMIT)
    grading_parser.add_argument("--expected-model", required=True)
    grading_parser.add_argument(
        "--database-url", default=os.environ.get("DATABASE_URL", "")
    )
    grading_parser.add_argument("--local-ai-env", type=Path, default=None)

    review_parser = subparsers.add_parser("lock-review")
    review_parser.add_argument("--run-id", required=True)
    review_parser.add_argument("--reviewer-id", required=True)
    review_parser.add_argument("--confirm-teacher-signoff", action="store_true")

    report_parser = subparsers.add_parser("report")
    report_parser.add_argument("--run-id", required=True)

    verify_parser = subparsers.add_parser("verify")
    verify_parser.add_argument("--run-id", required=True)

    return parser


def main() -> None:
    args = _build_parser().parse_args()
    try:
        if args.command == "prepare":
            harness_commit = args.harness_commit or current_git_commit()
            require_clean_git_worktree()
            require_commit_lineage(
                integration_commit=args.integration_commit,
                harness_commit=harness_commit,
            )
            _load_local_ai_environment(None)
            operator_assets = _operator_asset_metadata(args.grading_model)
            run_dir = prepare_evaluation(
                run_id=args.run_id,
                output_root=args.root,
                integration_commit=args.integration_commit,
                harness_commit=harness_commit,
                operator_assets=operator_assets,
                seed=args.seed,
            )
            print(
                json.dumps(
                    {
                        "run_dir": str(run_dir),
                        "state": current_state(run_dir),
                        "model_calls": 0,
                        "next_gate": "teacher completes ground_truth_review.xlsx",
                    },
                    indent=2,
                )
            )
            return
        run_dir = _resolve_run_dir(args.root, args.run_id)
        if args.command == "lock-ground-truth":
            lock = lock_ground_truth(
                run_dir,
                reviewer_id=args.reviewer_id,
                confirm_teacher_signoff=args.confirm_teacher_signoff,
            )
            print(json.dumps(_jsonable(lock), indent=2, sort_keys=True))
            return
        if args.command == "run-ocr":
            result = run_ocr_stage(
                run_dir,
                allow_local_ocr=args.allow_local_ocr,
                max_ocr_calls=args.max_ocr_calls,
                database_url=args.database_url,
                local_ai_env=args.local_ai_env,
            )
            print(
                json.dumps(
                    {
                        "run_id": result.run_id,
                        "state": current_state(run_dir),
                        "ocr_call_count": result.call_count,
                        "retry_count": result.retry_count,
                        "next_gate": "teacher completes ocr_review.xlsx",
                    },
                    indent=2,
                )
            )
            return
        if args.command == "lock-ocr-confirmations":
            lock = lock_ocr_confirmations(
                run_dir,
                reviewer_id=args.reviewer_id,
                confirm_teacher_signoff=args.confirm_teacher_signoff,
                database_url=args.database_url,
                local_ai_env=args.local_ai_env,
            )
            print(json.dumps(_jsonable(lock), indent=2, sort_keys=True))
            return
        if args.command == "record-ocr-no-go":
            lock = record_ocr_quality_no_go(
                run_dir,
                reviewer_id=args.reviewer_id,
                confirm_teacher_signoff=args.confirm_teacher_signoff,
                database_url=args.database_url,
                local_ai_env=args.local_ai_env,
            )
            print(
                json.dumps(
                    {
                        "run_id": lock.run_id,
                        "state": current_state(run_dir),
                        "rejected_case_count": len(lock.rejected_cases),
                        "grading_calls": 0,
                        "next_gate": "report (NO_GO_QUALITY; grading is blocked)",
                    },
                    indent=2,
                )
            )
            return
        if args.command == "run-grading":
            result = run_grading_stage(
                run_dir,
                allow_local_qwen=args.allow_local_qwen,
                max_qwen_calls=args.max_qwen_calls,
                expected_model=args.expected_model,
                database_url=args.database_url,
                local_ai_env=args.local_ai_env,
            )
            print(
                json.dumps(
                    {
                        "run_id": result.run_id,
                        "state": current_state(run_dir),
                        "qwen_call_count": result.qwen_call_count,
                        "blank_refusal_count": result.blank_refusal_count,
                        "next_gate": "teacher completes grading_review.xlsx",
                    },
                    indent=2,
                )
            )
            return
        if args.command == "lock-review":
            lock = lock_grading_review(
                run_dir,
                reviewer_id=args.reviewer_id,
                confirm_teacher_signoff=args.confirm_teacher_signoff,
            )
            print(json.dumps(_jsonable(lock), indent=2, sort_keys=True))
            return
        if args.command == "report":
            report = generate_report(run_dir)
            print(
                json.dumps(
                    {
                        "run_id": report["run_id"],
                        "verdict": report["verdict"],
                        "state": current_state(run_dir),
                        "report": str(run_dir / "report.md"),
                    },
                    indent=2,
                )
            )
            return
        if args.command == "verify":
            entries = read_ledger(run_dir)
            verify_locked_artifacts(run_dir, entries)
            print(
                json.dumps(
                    {
                        "run_id": args.run_id,
                        "state": entries[-1].state if entries else None,
                        "ledger_entries": len(entries),
                        "integrity": "valid",
                    },
                    indent=2,
                )
            )
            return
    except LocalCuratedEvaluationError as exc:
        raise SystemExit(f"Evaluation refused: {exc}") from exc


if __name__ == "__main__":
    main()
