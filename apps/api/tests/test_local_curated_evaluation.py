from __future__ import annotations

from collections import Counter
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from PIL import Image
from pydantic import ValidationError

from packages.evaluation import local_curated_evaluation as evaluation_module
from packages.evaluation.local_curated_evaluation import (
    AnswerQuality,
    EvaluationVerdict,
    GradingCaseResult,
    GradingRunResult,
    GroundTruthCase,
    GroundTruthLock,
    LocalCuratedEvaluationError,
    OcrCaseResult,
    OcrConfirmationCase,
    OcrRunResult,
    OperatorAssetMetadata,
    PrimaryCategory,
    append_state,
    build_case_blueprint,
    calculate_grading_metrics,
    calculate_ocr_metrics,
    character_error_rate,
    create_grading_review_workbook,
    create_ocr_review_workbook,
    critical_token_recall,
    current_state,
    evaluate_verdict,
    load_manifest,
    lock_ground_truth,
    normalize_text,
    prepare_evaluation,
    read_grading_review_workbook,
    read_ledger,
    read_ocr_review_workbook,
    run_grading_stage,
    run_ocr_stage,
    sha256_file,
    sha256_text,
    validate_database_url,
    verify_locked_artifacts,
    word_error_rate,
)

INTEGRATION_COMMIT = "d" * 40
HARNESS_COMMIT = "e" * 40


def _operator_assets() -> OperatorAssetMetadata:
    return OperatorAssetMetadata.model_validate(
        {
            "llama_cpp": {
                "build": "10249",
                "model_alias": "qwen3.6-35b-a3b-q4km",
                "model_sha256": "1" * 64,
                "model_size_bytes": 1024,
                "device": "gpu_hybrid",
            },
        }
    )


def _prepare(tmp_path: Path, run_id: str = "local_eval_test") -> Path:
    return prepare_evaluation(
        run_id=run_id,
        output_root=tmp_path,
        integration_commit=INTEGRATION_COMMIT,
        harness_commit=HARNESS_COMMIT,
        operator_assets=_operator_assets(),
        seed=12345,
    )


def _complete_ground_truth_workbook(run_dir: Path) -> None:
    manifest = load_manifest(run_dir)
    workbook_path = run_dir / "ground_truth_review.xlsx"
    workbook = load_workbook(workbook_path)
    sheet = workbook["Cases"]
    headers = {str(cell.value): index for index, cell in enumerate(sheet[1], start=1)}
    for row, case in enumerate(manifest.cases, start=2):
        sheet.cell(row, headers["teacher_transcription"], case.authored_transcription)
        sheet.cell(row, headers["teacher_score"], str(case.expected_score))
        sheet.cell(row, headers["teacher_notes"], "Independent synthetic review")
        if case.primary_category == PrimaryCategory.DIFFICULT_HANDWRITING:
            sheet.cell(row, headers["handwriting_acceptable"], "yes")
        sheet.cell(row, headers["teacher_approved"], "yes")
    workbook.save(workbook_path)


def _ground_truth(manifest: object, run_dir: Path) -> GroundTruthLock:
    typed_manifest = load_manifest(run_dir)
    del manifest
    return GroundTruthLock(
        run_id=typed_manifest.run_id,
        reviewer_id="teacher-reviewer",
        signed_at=datetime.now(UTC) - timedelta(minutes=10),
        manifest_sha256=sha256_file(run_dir / "manifest.json"),
        workbook_sha256=sha256_file(run_dir / "ground_truth_review.xlsx"),
        cases=[
            GroundTruthCase(
                case_id=case.case_id,
                teacher_transcription=case.authored_transcription,
                teacher_score=case.expected_score,
                teacher_notes="locked",
                handwriting_acceptable=(
                    True
                    if case.primary_category == PrimaryCategory.DIFFICULT_HANDWRITING
                    else None
                ),
                approved=True,
            )
            for case in typed_manifest.cases
        ],
    )


def _ocr_result(run_dir: Path) -> OcrRunResult:
    manifest = load_manifest(run_dir)
    now = datetime.now(UTC)
    return OcrRunResult(
        run_id=manifest.run_id,
        first_call_at=now - timedelta(minutes=8),
        completed_at=now - timedelta(minutes=5),
        call_count=20,
        service_status_before={"qwen": {"available": True}, "ocr": {"available": True}},
        service_status_after={"qwen": {"available": True}, "ocr": {"available": True}},
        database_name=f"teacher_assistant_eval_{manifest.run_id}",
        assessment_id=1,
        grading_run_id=1,
        owner_teacher_id=1,
        intruder_teacher_id=2,
        question_ids={pack: index for index, pack in enumerate("ABCDE", start=1)},
        rubric_ids={pack: index for index, pack in enumerate("ABCDE", start=11)},
        cases=[
            OcrCaseResult(
                case_id=case.case_id,
                answer_region_id=index,
                ocr_run_id=100 + index,
                status="succeeded",
                draft_text=case.authored_transcription,
                markdown=case.authored_transcription,
                blocks=[],
                warnings=[],
                latency_ms=1000 + index,
                provider="local_paddle_qwen",
                model="PaddleOCR-VL-1.6",
                layout_model="PP-DocLayoutV3",
                draft_text_sha256=sha256_text(case.authored_transcription),
            )
            for index, case in enumerate(manifest.cases, start=1)
        ],
    )


def _confirmations(run_dir: Path) -> list[OcrConfirmationCase]:
    manifest = load_manifest(run_dir)
    return [
        OcrConfirmationCase(
            case_id=case.case_id,
            answer_region_id=index,
            ocr_run_id=100 + index,
            confirmed_text=case.authored_transcription,
            confirmed_text_sha256=sha256_text(case.authored_transcription),
            teacher_approved=True,
            evidence_status=(
                "blank" if case.answer_quality == AnswerQuality.BLANK else "complete"
            ),
            full_answer_confirmed=case.answer_quality != AnswerQuality.BLANK,
        )
        for index, case in enumerate(manifest.cases, start=1)
    ]


def _grading_result(run_dir: Path) -> GradingRunResult:
    manifest = load_manifest(run_dir)
    results: list[GradingCaseResult] = []
    next_suggestion_id = 1
    for index, case in enumerate(manifest.cases, start=1):
        if case.answer_quality == AnswerQuality.BLANK:
            results.append(
                GradingCaseResult(
                    case_id=case.case_id,
                    answer_region_id=index,
                    outcome="not_called_blank_safety_gate",
                    max_score=case.max_score,
                    confirmed_text_sha256=sha256_text(case.authored_transcription),
                )
            )
            continue
        results.append(
            GradingCaseResult(
                case_id=case.case_id,
                answer_region_id=index,
                outcome="suggested",
                dispatch_run_id=((ord(case.pack_id) - ord("A")) + 1),
                dispatch_item_id=next_suggestion_id,
                grading_job_id=next_suggestion_id,
                grade_suggestion_id=next_suggestion_id,
                ai_score=case.expected_score,
                max_score=case.max_score,
                confidence=Decimal("0.70"),
                needs_review=True,
                rubric_breakdown=[
                    {
                        "criterion_id": "test",
                        "criterion": "Synthetic test criterion",
                        "max_marks": str(case.max_score),
                        "awarded_marks": str(case.expected_score),
                    }
                ],
                review_flags=[
                    "image_input_disabled",
                    "local_provider",
                    "teacher_review_required",
                ],
                model_provider="llama_cpp_qwen",
                model_name="qwen3.6-35b-a3b-q4km",
                prompt_version="real-grading-v2",
                marking_policy="general",
                token_usage={
                    "prompt_tokens": 60,
                    "completion_tokens": 40,
                    "total_tokens": 100,
                },
                latency_ms=100,
                cost_estimate=Decimal("0"),
                confirmed_text_sha256=sha256_text(case.authored_transcription),
            )
        )
        next_suggestion_id += 1
    return GradingRunResult(
        run_id=manifest.run_id,
        first_call_at=datetime.now(UTC) - timedelta(minutes=4),
        completed_at=datetime.now(UTC) - timedelta(minutes=1),
        qwen_call_count=18,
        blank_refusal_count=2,
        dispatch_run_ids=[1, 2, 3, 4, 5],
        cases=results,
        safety_checks={
            key: True for key in evaluation_module._REQUIRED_GRADING_SAFETY_CHECKS
        },
    )


def test_blueprint_has_exact_case_and_category_allocation() -> None:
    cases = build_case_blueprint()

    assert [case.case_id for case in cases] == [
        f"{pack}{index}" for pack in "ABCDE" for index in range(1, 5)
    ]
    assert Counter(case.primary_category for case in cases) == Counter(
        {
            PrimaryCategory.CORRECT: 3,
            PrimaryCategory.PARTIAL: 3,
            PrimaryCategory.WRONG: 3,
            PrimaryCategory.BLANK: 2,
            PrimaryCategory.IRRELEVANT: 2,
            PrimaryCategory.DIFFICULT_HANDWRITING: 3,
            PrimaryCategory.FORMULA_HEAVY: 2,
            PrimaryCategory.MULTI_STEP: 2,
        }
    )
    assert Counter(case.answer_quality for case in cases) == Counter(
        {
            AnswerQuality.CORRECT: 7,
            AnswerQuality.PARTIAL: 5,
            AnswerQuality.WRONG: 4,
            AnswerQuality.BLANK: 2,
            AnswerQuality.IRRELEVANT: 2,
        }
    )
    assert all(
        sum((criterion.max_marks for criterion in case.rubric), Decimal("0"))
        == case.max_score
        for case in cases
    )
    assert {case.case_id: (case.expected_score, case.max_score) for case in cases} == {
        "A1": (Decimal("3"), Decimal("3")),
        "A2": (Decimal("1"), Decimal("3")),
        "A3": (Decimal("0"), Decimal("3")),
        "A4": (Decimal("0"), Decimal("3")),
        "B1": (Decimal("4"), Decimal("4")),
        "B2": (Decimal("2"), Decimal("4")),
        "B3": (Decimal("0"), Decimal("4")),
        "B4": (Decimal("4"), Decimal("4")),
        "C1": (Decimal("1"), Decimal("5")),
        "C2": (Decimal("5"), Decimal("5")),
        "C3": (Decimal("4"), Decimal("5")),
        "C4": (Decimal("0"), Decimal("5")),
        "D1": (Decimal("6"), Decimal("6")),
        "D2": (Decimal("0"), Decimal("6")),
        "D3": (Decimal("5"), Decimal("6")),
        "D4": (Decimal("0"), Decimal("6")),
        "E1": (Decimal("3"), Decimal("5")),
        "E2": (Decimal("2"), Decimal("5")),
        "E3": (Decimal("5"), Decimal("5")),
        "E4": (Decimal("5"), Decimal("5")),
    }
    assert {
        case.case_id for case in cases if not case.authored_transcription
    } == {"A4", "C4"}


def test_prepare_is_deterministic_and_blanks_have_no_dark_text(tmp_path: Path) -> None:
    first = _prepare(tmp_path, "deterministic_one")
    second = _prepare(tmp_path, "deterministic_two")
    first_manifest = load_manifest(first)
    second_manifest = load_manifest(second)

    assert [case.image_sha256 for case in first_manifest.cases] == [
        case.image_sha256 for case in second_manifest.cases
    ]
    for case_id in ("A4", "C4"):
        case = next(case for case in first_manifest.cases if case.case_id == case_id)
        with Image.open(first / case.image_relative_path) as image:
            channel_minimums = [channel[0] for channel in image.getextrema()]
        assert min(channel_minimums) > 180
        assert case.authored_transcription == ""


def test_prepare_workbook_does_not_prefill_teacher_truth(tmp_path: Path) -> None:
    run_dir = _prepare(tmp_path)
    workbook = load_workbook(run_dir / "ground_truth_review.xlsx", data_only=True)
    sheet = workbook["Cases"]
    headers = {str(cell.value): index for index, cell in enumerate(sheet[1], start=1)}

    assert sheet.max_row == 21
    for row in range(2, 22):
        assert sheet.cell(row, headers["teacher_transcription"]).value is None
        assert sheet.cell(row, headers["teacher_score"]).value is None
        assert sheet.cell(row, headers["teacher_approved"]).value is None


def test_ground_truth_lock_requires_explicit_complete_human_signoff(tmp_path: Path) -> None:
    run_dir = _prepare(tmp_path)
    _complete_ground_truth_workbook(run_dir)

    with pytest.raises(LocalCuratedEvaluationError, match="Explicit teacher sign-off"):
        lock_ground_truth(
            run_dir,
            reviewer_id="teacher-1",
            confirm_teacher_signoff=False,
        )

    lock = lock_ground_truth(
        run_dir,
        reviewer_id="teacher-1",
        confirm_teacher_signoff=True,
    )

    assert len(lock.cases) == 20
    assert current_state(run_dir) == "ground_truth_locked"
    assert len(read_ledger(run_dir)) == 2
    verify_locked_artifacts(run_dir)


def test_hash_chain_refuses_skipped_state_and_tampering(tmp_path: Path) -> None:
    run_dir = _prepare(tmp_path)

    with pytest.raises(LocalCuratedEvaluationError, match="Cannot transition"):
        append_state(run_dir, "ocr_completed", locked_artifacts={})

    manifest_path = run_dir / "manifest.json"
    manifest_path.write_text(manifest_path.read_text(encoding="utf-8") + " ", encoding="utf-8")
    with pytest.raises(LocalCuratedEvaluationError, match="Locked evaluation artifact"):
        verify_locked_artifacts(run_dir)
    append_state(run_dir, "invalid", locked_artifacts={})
    assert current_state(run_dir) == "invalid"


def test_invalid_state_is_irreversible(tmp_path: Path) -> None:
    run_dir = _prepare(tmp_path)
    append_state(run_dir, "invalid", locked_artifacts={})

    with pytest.raises(LocalCuratedEvaluationError, match="already terminal"):
        append_state(run_dir, "invalid", locked_artifacts={})
    with pytest.raises(LocalCuratedEvaluationError, match="already terminal"):
        append_state(run_dir, "ground_truth_locked", locked_artifacts={})


def test_ocr_stage_refuses_because_no_tier1_engine_is_wired(tmp_path: Path) -> None:
    """The OCR stage must fail closed while it has no engine behind it.

    This replaces a test that drove a tampered artifact through run_ocr_stage to
    prove integrity failures mark a run invalid before any provider call. That
    path cannot be exercised now: run_ocr_stage refuses up front, and the
    grading stage - which carries the same verify/mark-invalid guard - requires
    the ocr_confirmed state that only the OCR stage can produce. The
    integrity-to-invalid behaviour of a real stage is therefore UNTESTED until
    the OCR stage is rewired, and must be re-covered then.
    """
    run_dir = _prepare(tmp_path)
    _complete_ground_truth_workbook(run_dir)
    lock_ground_truth(
        run_dir,
        reviewer_id="teacher-1",
        confirm_teacher_signoff=True,
    )
    image_path = run_dir / "images" / "A1.png"
    image_path.write_bytes(image_path.read_bytes() + b"tampered")

    with pytest.raises(LocalCuratedEvaluationError, match="not wired"):
        run_ocr_stage(
            run_dir,
            allow_local_ocr=True,
            max_ocr_calls=20,
            database_url="",
        )

    # Refusing must not poison an otherwise valid prepared run.
    assert current_state(run_dir) == "ground_truth_locked"


def test_asset_versions_and_result_hashes_are_schema_locked(tmp_path: Path) -> None:
    # The OCR engine's asset block is gone with the PaddleOCR stack, so the
    # locked-version check now covers the grading model alias instead.
    assets = _operator_assets().model_dump(mode="json")
    assets["llama_cpp"]["model_alias"] = "not-a-supported-model"
    with pytest.raises(ValidationError):
        OperatorAssetMetadata.model_validate(assets)

    run_dir = _prepare(tmp_path)
    manifest = load_manifest(run_dir)
    first = manifest.cases[0]
    with pytest.raises(ValidationError, match="draft hash"):
        OcrCaseResult(
            case_id=first.case_id,
            answer_region_id=1,
            ocr_run_id=1,
            status="succeeded",
            draft_text=first.authored_transcription,
            markdown="",
            latency_ms=1,
            provider="local_paddle_qwen",
            model="PaddleOCR-VL-1.6",
            layout_model="PP-DocLayoutV3",
            draft_text_sha256="0" * 64,
        )


def test_operator_asset_metadata_hashes_local_files_without_recording_paths(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    model = tmp_path / "model.gguf"
    binary = tmp_path / "llama-server.exe"
    model.write_bytes(b"qwen-model")
    binary.write_bytes(b"llama-binary")
    monkeypatch.setenv("LOCAL_QWEN_MODEL_PATH", str(model))
    monkeypatch.setenv("LOCAL_QWEN_BINARY_PATH", str(binary))

    def fake_run(command: list[str], **_kwargs: object) -> object:
        assert command[-1] == "--version"
        return type("Result", (), {"stdout": "llama.cpp build 10249", "stderr": ""})()

    monkeypatch.setattr(evaluation_module.subprocess, "run", fake_run)
    metadata = evaluation_module._operator_asset_metadata("qwen3.6-35b-a3b-q4km")
    payload = metadata.model_dump(mode="json")

    assert metadata.llama_cpp.model_sha256 == sha256_file(model)
    assert metadata.llama_cpp.model_alias == "qwen3.6-35b-a3b-q4km"
    # Local filesystem paths must never leak into the recorded manifest.
    assert str(tmp_path) not in str(payload)

    with pytest.raises(ValidationError, match="hidden provider metadata"):
        GradingCaseResult(
            case_id="A4",
            answer_region_id=4,
            outcome="not_called_blank_safety_gate",
            max_score=Decimal("3"),
            token_usage={"total_tokens": 1},
            confirmed_text_sha256=sha256_text(""),
        )


def test_database_name_guard_is_exact() -> None:
    run_id = "local_eval_20260807"
    expected = f"teacher_assistant_eval_{run_id}"

    assert (
        validate_database_url(f"postgresql+psycopg://postgres@127.0.0.1:55432/{expected}", run_id)
        == expected
    )
    with pytest.raises(LocalCuratedEvaluationError, match="named exactly"):
        validate_database_url(
            "postgresql+psycopg://postgres@127.0.0.1:55432/teacher_assistant_test",
            run_id,
        )


def test_real_stage_kill_switches_caps_and_model_alias_refuse_before_calls(
    tmp_path: Path,
) -> None:
    run_dir = _prepare(tmp_path)

    # The OCR stage's own --allow-local-ocr and call-cap guards are not asserted
    # here any more: it refuses before reaching them while no engine is wired.
    # Those guards must be re-covered when the stage is rewired.
    with pytest.raises(LocalCuratedEvaluationError, match="not wired"):
        run_ocr_stage(
            run_dir,
            allow_local_ocr=False,
            max_ocr_calls=20,
            database_url="",
        )
    with pytest.raises(LocalCuratedEvaluationError, match="allow-local-qwen"):
        run_grading_stage(
            run_dir,
            allow_local_qwen=False,
            max_qwen_calls=18,
            expected_model="qwen3.6-35b-a3b-q4km",
            database_url="",
        )
    with pytest.raises(LocalCuratedEvaluationError, match="exactly 18"):
        run_grading_stage(
            run_dir,
            allow_local_qwen=True,
            max_qwen_calls=17,
            expected_model="qwen3.6-35b-a3b-q4km",
            database_url="",
        )
    with pytest.raises(LocalCuratedEvaluationError, match="supported grading models"):
        run_grading_stage(
            run_dir,
            allow_local_qwen=True,
            max_qwen_calls=18,
            expected_model="wrong-model",
            database_url="",
        )
    # Both candidate aliases are accepted while the grading-model bake-off runs;
    # they fail later on run state, not on the alias check.
    for candidate in ("qwen3.6-35b-a3b-q4km", "qwen3.8-27b-q4km"):
        with pytest.raises(LocalCuratedEvaluationError, match="confirmed OCR evidence"):
            run_grading_stage(
                run_dir,
                allow_local_qwen=True,
                max_qwen_calls=18,
                expected_model=candidate,
                database_url="",
            )


def test_ocr_metrics_normalize_symbols_and_measure_formula_tokens(tmp_path: Path) -> None:
    run_dir = _prepare(tmp_path)
    manifest = load_manifest(run_dir)
    result = _ocr_result(run_dir)

    assert normalize_text(" 2 × 3 \n = 6 ") == "2 x 3 = 6"
    assert character_error_rate("2 × 3", "2 x 3") == 0
    assert word_error_rate("alpha beta", "alpha gamma") == Decimal("0.5")
    assert critical_token_recall(["KE", "1/2", "9", "J"], "KE = 1/2 mv^2 = 9 J") == 1

    metrics = calculate_ocr_metrics(manifest, result.cases, _confirmations(run_dir))
    assert metrics["overall_nonblank_cer"] == 0
    assert metrics["clean_typed_cer"] == 0
    assert metrics["handwriting_mean_cer"] == 0
    assert metrics["formula_critical_token_recall"] == 1
    assert metrics["blank_semantic_hallucination_count"] == 0
    assert metrics["block_order_issue_count"] == 0

    result.cases[0].blocks = [
        {"page": 1, "order": 2, "label": "text", "text": "second"},
        {"page": 1, "order": 1, "label": "text", "text": "first"},
    ]
    metrics = calculate_ocr_metrics(manifest, result.cases, _confirmations(run_dir))
    assert metrics["block_order_issue_count"] == 1


def test_ocr_review_requires_all_confirmed_text_to_match_locked_truth(tmp_path: Path) -> None:
    run_dir = _prepare(tmp_path)
    manifest = load_manifest(run_dir)
    ground_truth = _ground_truth(manifest, run_dir)
    ocr_result = _ocr_result(run_dir)
    workbook_path = create_ocr_review_workbook(manifest, ground_truth, ocr_result, run_dir)
    workbook = load_workbook(workbook_path)
    sheet = workbook["OCR Review"]
    headers = {str(cell.value): index for index, cell in enumerate(sheet[1], start=1)}
    truth_by_id = {case.case_id: case for case in ground_truth.cases}
    for row in range(2, 22):
        case_id = str(sheet.cell(row, headers["case_id"]).value)
        sheet.cell(row, headers["confirmed_text"], truth_by_id[case_id].teacher_transcription)
        sheet.cell(row, headers["teacher_approved"], "yes")
    workbook.save(workbook_path)

    confirmations = read_ocr_review_workbook(
        workbook_path,
        manifest,
        ocr_result,
        ground_truth,
    )
    assert len(confirmations) == 20
    assert [case.evidence_status for case in confirmations].count("blank") == 2

    workbook = load_workbook(workbook_path)
    sheet = workbook["OCR Review"]
    sheet.cell(2, headers["confirmed_text"], "changed answer")
    workbook.save(workbook_path)
    with pytest.raises(LocalCuratedEvaluationError, match="must match"):
        read_ocr_review_workbook(workbook_path, manifest, ocr_result, ground_truth)


def test_grading_metrics_and_all_three_verdicts(tmp_path: Path) -> None:
    run_dir = _prepare(tmp_path)
    manifest = load_manifest(run_dir)
    ocr_metrics = calculate_ocr_metrics(
        manifest,
        _ocr_result(run_dir).cases,
        _confirmations(run_dir),
    )
    grading_result = _grading_result(run_dir)
    grading_metrics = calculate_grading_metrics(manifest, grading_result.cases)

    assert grading_metrics["exact_count"] == 18
    assert grading_metrics["within_one_count"] == 18
    verdict, reasons = evaluate_verdict(
        manifest,
        process_checks={"safety": True},
        ocr_metrics=ocr_metrics,
        grading_metrics=grading_metrics,
    )
    assert verdict == EvaluationVerdict.PASS
    assert reasons == []

    first_suggested = next(
        case for case in grading_result.cases if case.outcome == "suggested"
    )
    first_suggested.ai_score = Decimal("0")
    first_suggested.confidence = Decimal("0.95")
    bad_metrics = calculate_grading_metrics(manifest, grading_result.cases)
    verdict, reasons = evaluate_verdict(
        manifest,
        process_checks={"safety": True},
        ocr_metrics=ocr_metrics,
        grading_metrics=bad_metrics,
    )
    assert verdict == EvaluationVerdict.NO_GO_QUALITY
    assert any("severe_false_confident" in reason for reason in reasons)

    verdict, reasons = evaluate_verdict(
        manifest,
        process_checks={"ownership": False},
        ocr_metrics=ocr_metrics,
        grading_metrics=grading_metrics,
    )
    assert verdict == EvaluationVerdict.INVALID_RUN
    assert reasons == ["Process or safety check failed: ownership"]


def test_quality_verdict_enforces_weak_answer_and_blank_protections(
    tmp_path: Path,
) -> None:
    run_dir = _prepare(tmp_path)
    manifest = load_manifest(run_dir)
    ocr_result = _ocr_result(run_dir)
    ocr_metrics = calculate_ocr_metrics(
        manifest,
        ocr_result.cases,
        _confirmations(run_dir),
    )
    grading_result = _grading_result(run_dir)
    grading_by_id = {case.case_id: case for case in grading_result.cases}
    grading_by_id["B3"].ai_score = Decimal("0.25")
    grading_by_id["D4"].ai_score = Decimal("0.25")
    grading_metrics = calculate_grading_metrics(manifest, grading_result.cases)
    verdict, reasons = evaluate_verdict(
        manifest,
        process_checks={"safety": True},
        ocr_metrics=ocr_metrics,
        grading_metrics=grading_metrics,
    )
    assert verdict == EvaluationVerdict.NO_GO_QUALITY
    assert any("repeated_zero_reference_overscore" in reason for reason in reasons)

    grading_by_id["B3"].ai_score = Decimal("0")
    grading_by_id["D4"].ai_score = Decimal("0")
    grading_by_id["A3"].ai_score = Decimal("1.51")
    grading_metrics = calculate_grading_metrics(manifest, grading_result.cases)
    verdict, reasons = evaluate_verdict(
        manifest,
        process_checks={"safety": True},
        ocr_metrics=ocr_metrics,
        grading_metrics=grading_metrics,
    )
    assert verdict == EvaluationVerdict.NO_GO_QUALITY
    assert any("wrong_over_half" in reason for reason in reasons)

    grading_by_id["A3"].ai_score = Decimal("0")
    blank_result = next(case for case in ocr_result.cases if case.case_id == "A4")
    blank_result.draft_text = "invented answer"
    hallucination_metrics = calculate_ocr_metrics(
        manifest,
        ocr_result.cases,
        _confirmations(run_dir),
    )
    verdict, reasons = evaluate_verdict(
        manifest,
        process_checks={"safety": True},
        ocr_metrics=hallucination_metrics,
        grading_metrics=calculate_grading_metrics(manifest, grading_result.cases),
    )
    assert verdict == EvaluationVerdict.NO_GO_QUALITY
    assert any("blank_hallucination" in reason for reason in reasons)


def test_grading_review_requires_classified_disagreements_and_blank_refusals(
    tmp_path: Path,
) -> None:
    run_dir = _prepare(tmp_path)
    manifest = load_manifest(run_dir)
    grading_result = _grading_result(run_dir)
    workbook_path = create_grading_review_workbook(manifest, grading_result, run_dir)
    workbook = load_workbook(workbook_path)
    sheet = workbook["Grading Review"]
    headers = {str(cell.value): index for index, cell in enumerate(sheet[1], start=1)}
    result_by_id = {case.case_id: case for case in grading_result.cases}
    for row in range(2, 22):
        case_id = str(sheet.cell(row, headers["case_id"]).value)
        result = result_by_id[case_id]
        sheet.cell(row, headers["disagreement_reason"], "none")
        sheet.cell(row, headers["useful_draft"], "no" if result.ai_score is None else "yes")
        sheet.cell(row, headers["approved_review"], "yes")
    workbook.save(workbook_path)

    reviews = read_grading_review_workbook(workbook_path, manifest, grading_result)
    assert len(reviews) == 20

    workbook = load_workbook(workbook_path)
    sheet = workbook["Grading Review"]
    sheet.cell(2, headers["disagreement_reason"], "model_error")
    workbook.save(workbook_path)
    with pytest.raises(LocalCuratedEvaluationError, match="exact scores require"):
        read_grading_review_workbook(workbook_path, manifest, grading_result)
